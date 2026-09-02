# -*- coding: utf-8 -*-
"""
SQL / Excel / JSON / CSV 数据互转工具 (PySide6 / Qt GUI)
====================================================
功能(输入 × 输出全组合):
    输入:  SQL 数据库(自由勾选 TableName + ColumnName)
           Excel (.xlsx) / JSON (.json) / CSV (.csv)
    输出:  Excel (.xlsx) / JSON (.json) / CSV (.csv) / SQL 数据库

命名约定(库-表-字段三层一一对应):
    数据库名  == 输出文件名(Excel/JSON/CSV)
    表名      == Excel Sheet 名 == JSON 一级参数
    字段名    == Excel 首行单元格 == JSON 二级参数 == CSV 表头

设计说明:
    * 数据库适配:MySQL 使用 PyMySQL,PostgreSQL 使用 psycopg2,统一 executemany 批量写入;
    * Excel 读写使用 openpyxl 的 read_only / write_only 流式模式,大批量数据内存占用低;
    * JSON 采用“列数组”结构(一级键=表名,二级键=字段名,值为数组);
    * CSV 使用标准库 csv,支持分隔符/编码选择,写文件前检测同名文件并支持覆盖/合并;
    * UI 采用 PySide6(Qt)+ QSS 彩色主题,输入/输出分区动态展开,数据库连接面板按需显示。
"""

import csv
import json
import os
import re
import sys
import threading
import traceback
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path

import html
from PySide6.QtCore import Qt, QObject, Signal
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QLabel, QLineEdit, QComboBox,
    QPushButton, QFrame, QGroupBox, QGridLayout, QHBoxLayout, QVBoxLayout,
    QPlainTextEdit, QFileDialog, QMessageBox, QDialog, QTreeWidget,
    QTreeWidgetItem, QSpinBox,
)

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

try:
    import pymysql
except Exception:  # pragma: no cover
    pymysql = None

try:
    import psycopg2
except Exception:  # pragma: no cover
    psycopg2 = None

# ---------------------------------------------------------------------------
# 常量 / 配色
# ---------------------------------------------------------------------------
APP_TITLE = "SQL ↔ Excel / JSON / CSV 数据转换工具"
SAMPLE_SIZE = 200            # 建表时用于推断字段类型的采样行数
DEFAULT_BATCH = 1000         # 默认每批读写行数
FONT = ("Microsoft YaHei", 10)
FONT_BOLD = ("Microsoft YaHei", 10, "bold")

# UI 配色(clam 主题)
BG = "#eef2f8"               # 窗体背景
CARD = "#ffffff"             # 卡片背景
NAVY = "#16294d"             # 顶部横幅
TEXT = "#1f2733"             # 主文字
MUTED = "#5b6b7f"            # 次要文字
BORDER = "#c9d4e4"           # 卡片边框
PRIMARY = "#1f6feb"          # 主蓝色(SQL)
GREEN = "#1a9e5c"            # Excel
ORANGE = "#e8833a"           # JSON
PURPLE = "#8e5bd8"           # CSV
RED = "#d64545"              # 错误
TEAL = "#0e9aa7"             # 青色(辅助)
AMBER = "#f0a12e"            # 琥珀(辅助)
INDIGO = "#5a67d8"           # 靛蓝(辅助)
MAGENTA = "#c256b1"          # 品红(辅助)
CORAL = "#e56b4f"            # 珊瑚(辅助)
CYAN = "#22b8cf"             # 湖蓝(辅助)
STATUS_OK = "#1a9e5c"
STATUS_OFF = "#9aa7b8"

ACCENT_STRIP = [PRIMARY, CYAN, TEAL, GREEN, AMBER, ORANGE, CORAL, MAGENTA, PURPLE]

MODE_COLORS = {"SQL": PRIMARY, "Excel": GREEN, "JSON": ORANGE, "CSV": PURPLE}

INPUT_LABELS = {
    "SQL": "SQL 数据库",
    "Excel": "Excel 文件 (.xlsx)",
    "JSON": "JSON 文件 (.json)",
    "CSV": "CSV 文件 (.csv)",
}
OUTPUT_LABELS = {
    "Excel": "Excel 文件 (.xlsx)",
    "JSON": "JSON 文件 (.json)",
    "CSV": "CSV 文件 (.csv)",
    "SQL": "SQL 数据库",
}

MODE_APPEND = "追加到已有表 (append)"
MODE_CREATE = "不存在则创建,存在则追加 (create if missing)"
MODE_REPLACE = "删除重建 (replace)"
MODE_LABELS = [MODE_APPEND, MODE_CREATE, MODE_REPLACE]

DELIM_LABELS = {"逗号 (,)": ",", "分号 (;)": ";", "制表符 (TAB)": "\t", "竖线 (|)": "|"}
ENC_LABELS = {
    "UTF-8 with BOM (Excel 友好)": "utf-8-sig",
    "UTF-8": "utf-8",
    "GBK (中文 Excel)": "gbk",
}

IDENT_RE = re.compile(r"^[^\W\d]\w*$")  # 标识符:字母/下划线开头,允许中文、数字、下划线
CSV_SECTION_RE = re.compile(r"^\[Sheet:(.+?)\]$")  # 多表 CSV 分节标记,如 [Sheet:departments]


def config_file_path():
    """连接配置记忆文件路径(Windows: %APPDATA%\\DataTransformer\\config.json)。"""
    base = os.environ.get("APPDATA") or str(Path.home())
    return Path(base) / "DataTransformer" / "config.json"


# ---------------------------------------------------------------------------
# 数据库适配层:MySQL / PostgreSQL
# ---------------------------------------------------------------------------
class DBAdapter:
    """数据库适配器基类,MySQL 与 PostgreSQL 各自实现差异点。"""

    NAME = ""
    DEFAULT_PORT = 3306
    QUOTE = '"'

    def connect(self, cfg):
        raise NotImplementedError

    def open_query_cursor(self, conn):
        """打开一个流式查询游标(避免一次性把全部结果缓冲到客户端内存)。"""
        raise NotImplementedError

    def list_tables(self, conn):
        raise NotImplementedError

    def list_columns(self, conn, table):
        raise NotImplementedError

    def is_missing_table_error(self, exc):
        raise NotImplementedError

    # -- 通用能力 ----------------------------------------------------------
    def quote_ident(self, name):
        """校验并引用标识符,支持 schema.table 形式,防止 SQL 注入。"""
        name = str(name).strip()
        if not name:
            raise ValueError("表名/字段名不能为空")
        parts = [p.strip() for p in name.split(".") if p.strip()]
        for part in parts:
            if not IDENT_RE.match(part):
                raise ValueError(f"非法标识符(仅允许字母/下划线/数字): {name!r}")
        q = self.QUOTE
        return ".".join(f"{q}{p}{q}" for p in parts)

    def table_exists(self, conn, quoted_table):
        """通过 SELECT 探测表是否存在(跨库、跨 schema 均可靠)。"""
        try:
            with conn.cursor() as cur:
                cur.execute(f"SELECT 1 FROM {quoted_table} WHERE 1=0")
            return True
        except Exception as exc:
            # PostgreSQL 中任何出错语句都会使当前事务进入 aborted 状态,
            # 必须先 rollback 才能继续后续 SQL。
            try:
                conn.rollback()
            except Exception:
                pass
            if self.is_missing_table_error(exc):
                return False
            raise

    def server_version(self, conn):
        with conn.cursor() as cur:
            cur.execute("SELECT VERSION()")
            return str(cur.fetchone()[0])


class MySQLAdapter(DBAdapter):
    NAME = "MySQL"
    DEFAULT_PORT = 3306
    QUOTE = "`"

    def connect(self, cfg):
        if pymysql is None:
            raise RuntimeError("未安装 PyMySQL,请先执行: pip install pymysql")
        return pymysql.connect(
            host=cfg["host"],
            port=int(cfg["port"]),
            user=cfg["user"],
            password=cfg["password"],
            database=cfg["database"],
            charset=cfg.get("charset") or "utf8mb4",
            connect_timeout=int(cfg.get("timeout", 10)),
        )

    def open_query_cursor(self, conn):
        # SSCursor = 服务端无缓冲游标,大数据量导出不占客户端内存
        return conn.cursor(pymysql.cursors.SSCursor)

    def list_tables(self, conn):
        with conn.cursor() as cur:
            cur.execute("SHOW TABLES")
            return [str(row[0]) for row in cur.fetchall()]

    def list_columns(self, conn, table):
        with conn.cursor() as cur:
            cur.execute(f"SHOW COLUMNS FROM {self.quote_ident(table)}")
            return [str(row[0]) for row in cur.fetchall()]

    def is_missing_table_error(self, exc):
        errno = exc.args[0] if exc.args else None
        return isinstance(exc, pymysql.err.ProgrammingError) and errno == 1146


class PostgresAdapter(DBAdapter):
    NAME = "PostgreSQL"
    DEFAULT_PORT = 5432
    QUOTE = '"'

    def connect(self, cfg):
        if psycopg2 is None:
            raise RuntimeError("未安装 psycopg2,请先执行: pip install psycopg2-binary")
        return psycopg2.connect(
            host=cfg["host"],
            port=int(cfg["port"]),
            user=cfg["user"],
            password=cfg["password"],
            dbname=cfg["database"],
            connect_timeout=int(cfg.get("timeout", 10)),
        )

    def open_query_cursor(self, conn):
        # 服务端命名游标:结果按 itersize 分批从服务器拉取
        cur = conn.cursor(name=f"sql_excel_cur_{os.getpid()}_{id(conn)}")
        cur.itersize = 5000
        return cur

    def list_tables(self, conn):
        with conn.cursor() as cur:
            cur.execute(
                "SELECT table_schema, table_name FROM information_schema.tables "
                "WHERE table_schema NOT IN ('pg_catalog', 'information_schema') "
                "ORDER BY table_schema, table_name"
            )
            return [f"{s}.{t}" if s != "public" else str(t) for s, t in cur.fetchall()]

    def list_columns(self, conn, table):
        parts = table.split(".")
        schema, tname = (parts[0], parts[-1]) if len(parts) == 2 else ("public", parts[-1])
        with conn.cursor() as cur:
            cur.execute(
                "SELECT column_name FROM information_schema.columns "
                "WHERE table_schema = %s AND table_name = %s "
                "ORDER BY ordinal_position",
                (schema, tname),
            )
            return [str(row[0]) for row in cur.fetchall()]

    def is_missing_table_error(self, exc):
        return getattr(exc, "pgcode", None) == "42P01"


ADAPTERS = {
    "MySQL": MySQLAdapter(),
    "PostgreSQL": PostgresAdapter(),
}


# ---------------------------------------------------------------------------
# Excel 工具函数(openpyxl 流式模式)
# ---------------------------------------------------------------------------
def safe_sheet_title(name):
    """工作表名:去掉非法字符,长度 <= 31。"""
    name = re.sub(r'[\\/*?:\[\]]', "_", str(name)).strip()
    return (name[:31] or "Sheet1")


def display_width(text):
    """按显示宽度估算列宽(中文等全角字符算 2)。"""
    return sum(2 if ord(ch) > 0x7F else 1 for ch in str(text))


def complex_to_json_text(value):
    """dict / list / tuple 等复合值(如 PostgreSQL jsonb 列)→ JSON 文本。"""
    return json.dumps(value, ensure_ascii=False, default=str)


def to_excel_value(value):
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, (dict, list, tuple)):   # json/jsonb 列
        return complex_to_json_text(value)
    if isinstance(value, bytes):
        try:
            return value.decode("utf-8")
        except UnicodeDecodeError:
            return value.decode("utf-8", "replace")
    return value  # int / float / Decimal / str / bool 等


def write_rows_to_worksheet(ws, header, row_iterator, log=None):
    """把数据行逐行写入工作表,顺便自动计算列宽。"""
    ws.append([to_excel_value(h) for h in header])
    widths = {i + 1: min(display_width(h) + 2, 60) for i, h in enumerate(header)}
    total = 0
    for row in row_iterator:
        values = [to_excel_value(v) for v in row]
        ws.append(values)
        for idx, value in enumerate(values, start=1):
            if value is None:
                continue
            width = min(display_width(value) + 2, 60)
            if width > widths.get(idx, 0):
                widths[idx] = width
        total += 1
        if log and total % 20000 == 0:
            log(f"  已写出 {total:,} 行 ...")
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width
    return total


def is_blank_row(row):
    return all(
        v is None or (isinstance(v, str) and not v.strip())
        for v in row
    )


def normalize_row(row, ncols):
    """把行对齐到表头列数:短了补 None,长了截断。"""
    row = list(row)
    if len(row) > ncols:
        return row[:ncols]
    return row + [None] * (ncols - len(row))


def to_db_value(value):
    if value is None:
        return None
    if isinstance(value, str):
        s = value.strip()
        return s or None
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, (dict, list, tuple)):   # json/jsonb 列 → JSON 文本入库
        return complex_to_json_text(value)
    return value  # datetime / date / Decimal / bool / int / float 直接交给驱动


def parse_header(raw_row):
    """解析表头:去掉尾部空列,空列自动命名,重名自动加后缀。"""
    cells = list(raw_row)
    last = len(cells)
    while last > 0 and (
        cells[last - 1] is None
        or (isinstance(cells[last - 1], str) and not cells[last - 1].strip())
    ):
        last -= 1
    if last == 0:
        raise ValueError("数据表头为空,无法转换")

    header, counts = [], {}
    for i in range(last):
        name = str(cells[i]).strip() if cells[i] is not None else ""
        if not name:
            name = f"column_{i + 1}"
        if name in counts:
            counts[name] += 1
            name = f"{name}_{counts[name]}"
        else:
            counts[name] = 1
        header.append(name)
    return header


def collect_stats(header, sample_rows):
    """统计采样行中各列出现的数据类型,用于 CREATE TABLE 的类型推断。"""
    stats = {
        i: {
            "str": False, "max_len": 0, "int": False, "big": False,
            "float": False, "bool": False, "datetime": False, "date": False,
        }
        for i in range(len(header))
    }
    for row in sample_rows:
        for i, value in enumerate(row):
            st = stats[i]
            if value is None:
                continue
            if isinstance(value, bool):
                st["bool"] = True
            elif isinstance(value, int):
                st["int"] = True
                if value > 2 ** 31 - 1 or value < -(2 ** 31):
                    st["big"] = True
            elif isinstance(value, (float, Decimal)):
                st["float"] = True
            elif isinstance(value, datetime):
                st["datetime"] = True
            elif isinstance(value, date):
                st["date"] = True
            else:
                st["str"] = True
                st["max_len"] = max(st["max_len"], len(str(value)))
    return stats


def decide_type(adapter, st):
    is_mysql = adapter.NAME == "MySQL"
    if st.get("str"):                   # 含文本(混合类型也按文本兜底)
        # 字符串列统一建为 TEXT:采样只覆盖前 SAMPLE_SIZE 行,若按采样长度
        # 建 VARCHAR(n),后续出现更长的值会触发
        # "value too long for type character varying(n)" 写入失败。
        # TEXT(MySQL 上限 64KB,PostgreSQL 无上限)可彻底规避该问题。
        return "TEXT"
    if st.get("datetime"):
        return "TIMESTAMP"
    if st.get("date"):
        return "DATE"
    if st.get("float"):
        return "DOUBLE" if is_mysql else "DOUBLE PRECISION"
    if st.get("bool") and not st.get("int"):
        return "TINYINT(1)" if is_mysql else "BOOLEAN"
    if st.get("int"):
        return "BIGINT" if st.get("big") else "INT"
    return "TEXT"


def build_create_table(adapter, table, header, stats):
    columns = [
        f"{adapter.quote_ident(col)} {decide_type(adapter, stats[idx])}"
        for idx, col in enumerate(header)
    ]
    return f"CREATE TABLE {adapter.quote_ident(table)} ({', '.join(columns)})"


# ---------------------------------------------------------------------------
# 文件写出工具(含同名文件冲突处理)
# ---------------------------------------------------------------------------
def safe_filename(name):
    """去掉文件名中 Windows 不允许的字符。"""
    name = re.sub(r'[\\/:*?"<>|\r\n]+', "_", str(name)).strip()
    return name or "export"


def to_json_value(value):
    """把数据库返回值转换成 JSON 兼容值(时间统一为字符串,与示例文件一致)。"""
    if value is None:
        return None
    if isinstance(value, dict):                   # json/jsonb 列:递归转换,保留嵌套结构
        return {str(k): to_json_value(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        return [to_json_value(v) for v in value]
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, bytes):
        try:
            return value.decode("utf-8")
        except UnicodeDecodeError:
            return value.decode("utf-8", "replace")
    return value  # int / float / str / bool


def to_csv_value(value):
    """把数据值转换成 CSV 单元格文本(时间与 JSON 规则一致)。"""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, (dict, list, tuple)):   # json/jsonb 列
        return complex_to_json_text(value)
    if isinstance(value, bytes):
        try:
            return value.decode("utf-8")
        except UnicodeDecodeError:
            return value.decode("utf-8", "replace")
    return value


def build_json_document(top_key, columns, rows):
    """构造 {表名: {字段名: [值...]}} 的列数组式 JSON 结构。"""
    doc = {top_key: {col: [] for col in columns}}
    for row in rows:
        norm = normalize_row(row, len(columns))
        for col, value in zip(columns, norm):
            doc[top_key][col].append(to_json_value(value))
    return doc


def write_json_file(path, top_key, columns, rows, mode="replace", log=None):
    """写出 JSON 文件。mode=merge 时保留原文件其他一级键,仅覆盖同名键。"""
    doc = build_json_document(top_key, columns, rows)
    if mode == "merge" and os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as fp:
                old = json.load(fp)
        except Exception:
            old = None
        if isinstance(old, dict):
            old[top_key] = doc[top_key]
            doc = old
    with open(path, "w", encoding="utf-8") as fp:
        json.dump(doc, fp, ensure_ascii=False, indent=4)
    return len(rows)


def write_excel_multi_file(path, tables, mode="replace", log=None):
    """多表写出 Excel:tables 为可迭代的 (sheet_name, columns, row_iter)。

    每个表一个 Sheet;mode=merge 时保留原工作簿其他 Sheet,覆盖同名 Sheet。
    """
    if mode == "replace":
        wb = Workbook(write_only=True)
        try:
            total = 0
            for sheet, columns, rows in tables:
                ws = wb.create_sheet(title=sheet)
                total += write_rows_to_worksheet(ws, columns, rows, log)
            wb.save(path)
        finally:
            try:
                wb.close()
            except Exception:
                pass
        return total

    wb = load_workbook(path)
    try:
        total = 0
        for sheet, columns, rows in tables:
            if sheet in wb.sheetnames:
                del wb[sheet]
            ws = wb.create_sheet(title=sheet)
            total += write_rows_to_worksheet(ws, columns, rows, log)
        wb.save(path)
    finally:
        wb.close()
    return total


def write_json_multi_file(path, tables, mode="replace", log=None):
    """多表写出 JSON:tables 为可迭代的 (top_key, columns, rows)。

    每个表一个一级键;mode=merge 时保留原文件其他一级键,覆盖同名键。
    """
    payload = {}
    total = 0
    for key, columns, rows in tables:
        sub = build_json_document(key, columns, rows)[key]
        payload[key] = sub
        total += len(sub[columns[0]]) if columns else 0
    if mode == "merge" and os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as fp:
                old = json.load(fp)
        except Exception:
            old = None
        if isinstance(old, dict):
            old.update(payload)
            payload = old
    with open(path, "w", encoding="utf-8") as fp:
        json.dump(payload, fp, ensure_ascii=False, indent=4)
    return total


def write_excel_file(path, sheet, columns, row_iterator, mode="replace", log=None):
    """写出 Excel 文件。

    mode=replace : 用流式 write_only 模式重建整个工作簿(大数据量首选);
    mode=merge   : 保留原工作簿其他 Sheet,覆盖同名 Sheet(无同名则新增)。
    """
    if mode == "replace":
        wb = Workbook(write_only=True)
        try:
            ws = wb.create_sheet(title=sheet)
            total = write_rows_to_worksheet(ws, columns, row_iterator, log)
            wb.save(path)
        finally:
            try:
                wb.close()
            except Exception:
                pass
        return total

    # merge:打开已有工作簿(常规模式),只动同名 Sheet
    wb = load_workbook(path)
    try:
        if sheet in wb.sheetnames:
            del wb[sheet]
        ws = wb.create_sheet(title=sheet)
        total = write_rows_to_worksheet(ws, columns, row_iterator, log)
        wb.save(path)
    finally:
        wb.close()
    return total


def write_csv_file(path, columns, row_iterator, delimiter=",", encoding="utf-8-sig",
                   mode="replace", log=None):
    """写出 CSV 文件。

    mode=replace : 重建文件(含表头);
    mode=merge   : 向已有文件追加数据行(不重复写表头)。
    """
    append = mode == "merge" and os.path.exists(path)
    fp = open(path, "a" if append else "w", newline="", encoding=encoding)
    writer = csv.writer(fp, delimiter=delimiter, lineterminator="\n")
    total = 0
    try:
        if not append:
            writer.writerow([to_csv_value(h) for h in columns])
        for row in row_iterator:
            writer.writerow(
                [to_csv_value(v) for v in normalize_row(row, len(columns))])
            total += 1
            if log and total % 50000 == 0:
                log(f"  已写出 {total:,} 行 ...")
    finally:
        fp.close()
    return total


# ---------------------------------------------------------------------------
# 多表 CSV:目录模式(目录名 = 数据库名,每个表一个 {表名}.csv);
# 以及旧版"分节 CSV"文件的读取(向后兼容)
# ---------------------------------------------------------------------------
def write_csv_directory(dir_path, tables, delimiter=",", encoding="utf-8-sig",
                        mode="replace", log=None):
    """多表写出 CSV:tables 为可迭代的 (表名, columns, row_iterator)。

    在 dir_path 目录下为每个表生成一个 {表名}.csv 文件(表名 = 文件名)。
    """
    os.makedirs(dir_path, exist_ok=True)
    total = 0
    for sheet, columns, rows in tables:
        path = os.path.join(dir_path, f"{safe_filename(sheet)}.csv")
        total += write_csv_file(path, columns, rows, delimiter=delimiter,
                                encoding=encoding, mode=mode, log=log)
    return total


def read_csv_sections(path, delimiter, encoding):
    """解析旧版分节 CSV,返回 {分节名: (表头, 数据行列表)};

    普通(无分节标记)的 CSV 返回 None。
    """
    with open(path, "r", newline="", encoding=encoding) as fp:
        rows_all = list(csv.reader(fp, delimiter=delimiter))
    if not any(len(r) == 1 and CSV_SECTION_RE.match(r[0].strip()) for r in rows_all):
        return None
    sections = {}
    current = None
    header = None
    rows = []
    for row in rows_all:
        m = CSV_SECTION_RE.match(row[0].strip()) if (row and len(row) == 1) else None
        if m:
            if current is not None:
                sections[current] = (header or ["column_1"], rows)
            current = m.group(1)
            header = None
            rows = []
        elif current is not None:
            if header is None:
                header = parse_header(row)
            elif not is_blank_row(row):
                rows.append(row)
    if current is not None:
        sections[current] = (header or ["column_1"], rows)
    return sections


def load_csv_section(path, section, delimiter, encoding, parse=False):
    """读取分节 CSV 中的某一节,返回 (表头, 行生成器)。"""
    sections = read_csv_sections(path, delimiter, encoding)
    if not sections or section not in sections:
        raise ValueError(
            f"分节 CSV 中不存在分节 {section!r},"
            f"可用: {', '.join(sections or [])}")
    header, rows = sections[section]

    def gen():
        for row in rows:
            norm = normalize_row(row, len(header))
            if parse:
                yield [parse_csv_cell(c) for c in norm]
            else:
                yield norm

    return header, gen()


# ---------------------------------------------------------------------------
# 数据读取工具(统一为: 表头 + 行生成器)
# ---------------------------------------------------------------------------
def load_excel_stream(path, sheet):
    """读取 xlsx 工作表,返回 (表头, 行生成器)。"""
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        raise ValueError(f"工作簿中不存在工作表 {sheet!r},可用: {', '.join(wb.sheetnames)}")
    ws = wb[sheet]
    it = ws.iter_rows(values_only=True)
    try:
        raw_header = next(it)
    except StopIteration:
        wb.close()
        raise ValueError("工作表为空,无法读取")
    header = parse_header(raw_header)

    def gen():
        try:
            for row in it:
                if is_blank_row(row):
                    continue
                yield normalize_row(row, len(header))
        finally:
            wb.close()

    return header, gen()


def filter_columns(header, rows_iter, cols):
    """按勾选字段过滤行流,返回 (新表头, 行生成器)。"""
    idxs = [header.index(c) for c in cols if c in header]
    new_header = [header[i] for i in idxs]

    def gen():
        for row in rows_iter:
            yield [row[i] for i in idxs]

    return new_header, gen()


def read_csv_header(path, delimiter, encoding):
    with open(path, "r", newline="", encoding=encoding) as fp:
        reader = csv.reader(fp, delimiter=delimiter)
        try:
            raw = next(reader)
        except StopIteration:
            raise ValueError("CSV 文件为空,无法读取")
    return parse_header(raw)


def parse_csv_cell(value):
    """CSV 单元格文本 → 自动推断 int / float / date / datetime,失败保持字符串。"""
    if not isinstance(value, str):
        return value
    s = value.strip()
    if s == "":
        return None
    if re.fullmatch(r"-?\d{1,9}", s):          # 短整数(身份证/长号码保留字符串)
        return int(s)
    if re.fullmatch(r"-?\d+\.\d+", s):
        return float(s)
    try:
        return datetime.strptime(s, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        pass
    try:
        return date.fromisoformat(s)
    except ValueError:
        pass
    return s


def iter_csv_rows(path, delimiter, encoding, header_len, parse=False):
    """逐行生成 CSV 数据(跳过表头与空行),parse=True 时按类型解析单元格。"""
    with open(path, "r", newline="", encoding=encoding) as fp:
        reader = csv.reader(fp, delimiter=delimiter)
        next(reader, None)  # 跳过表头
        for row in reader:
            if is_blank_row(row):
                continue
            norm = normalize_row(row, header_len)
            if parse:
                yield [parse_csv_cell(c) for c in norm]
            else:
                yield norm


def load_json_stream(path, top_key):
    """读取 JSON 列数组结构,返回 (表头, 行生成器)。"""
    with open(path, "r", encoding="utf-8") as fp:
        doc = json.load(fp)
    if not isinstance(doc, dict):
        raise ValueError("JSON 顶层必须是对象")
    if top_key not in doc:
        raise ValueError(f"JSON 中不存在一级键 {top_key!r},可用: {list(doc.keys())}")
    table = doc[top_key]
    if not isinstance(table, dict):
        raise ValueError(f"一级键 {top_key!r} 的值必须是 {{字段: [值...]}} 对象")
    columns = list(table.keys())
    arrays = []
    for col in columns:
        arr = table.get(col)
        if not isinstance(arr, list):
            raise ValueError(f"字段 {col!r} 的值必须是数组")
        arrays.append(arr)
    nrows = max((len(a) for a in arrays), default=0)

    def gen():
        for i in range(nrows):
            yield [arr[i] if i < len(arr) else None for arr in arrays]

    return columns, gen()


def load_sql_stream(conn, adapter, table, columns=None):
    """读取数据库表,返回 (表头, 行生成器, 游标)。调用方负责关闭游标。"""
    col_sql = ", ".join(adapter.quote_ident(c) for c in columns) if columns else "*"
    cur = adapter.open_query_cursor(conn)
    cur.execute(f"SELECT {col_sql} FROM {adapter.quote_ident(table)}")
    # PostgreSQL 命名游标的 description 要等首次 fetch 后才就绪
    first = cur.fetchmany(DEFAULT_BATCH)
    header = [str(desc[0]) for desc in cur.description]

    def gen():
        for row in first:
            yield list(row)
        while True:
            rows = cur.fetchmany(DEFAULT_BATCH)
            if not rows:
                break
            for row in rows:
                yield list(row)

    return header, gen(), cur


def write_sql_table(conn, adapter, table, header, rows_iter, mode, batch, log,
                    sample=None):
    """把行流写入数据库表。

    sample=None 时单遍采样前 SAMPLE_SIZE 行推断类型(适合文件输入);
    调用方也可自行提供 sample 并传入未消费的行流(适合 SQL 输入,避免游标与 DDL 冲突)。
    """
    if sample is None:
        # 单遍采样:先缓存前 SAMPLE_SIZE 行用于类型推断,随后接续剩余行
        sample, buffered = [], []
        for row in rows_iter:
            norm = normalize_row(row, len(header))
            buffered.append(norm)
            if len(sample) < SAMPLE_SIZE:
                sample.append(norm)
            if len(buffered) >= SAMPLE_SIZE:
                break

        source_iter = rows_iter  # 固定引用,避免闭包晚绑定

        def chained():
            for row in buffered:
                yield row
            for row in source_iter:
                yield normalize_row(row, len(header))

        rows_iter = chained()

    quoted_table = adapter.quote_ident(table)
    exists = adapter.table_exists(conn, quoted_table)
    need_create = (mode == "replace") or (mode == "create_if_missing" and not exists)
    if mode == "append" and not exists:
        raise ValueError(
            f"目标表【{table}】不存在。请改用“{MODE_CREATE}”或“{MODE_REPLACE}”模式")
    if need_create:
        stats = collect_stats(header, sample)
        if mode == "replace" and exists:
            with conn.cursor() as cur:
                cur.execute(f"DROP TABLE {quoted_table}")
            log(f"已删除旧表: {table}")
        ddl = build_create_table(adapter, table, header, stats)
        with conn.cursor() as cur:
            cur.execute(ddl)
        log(f"已创建表: {table}")

    insert_sql = (
        f"INSERT INTO {quoted_table} "
        f"({', '.join(adapter.quote_ident(c) for c in header)}) "
        f"VALUES ({', '.join(['%s'] * len(header))})"
    )
    total = 0
    chunk = []
    try:
        with conn.cursor() as cur:
            for row in rows_iter:
                chunk.append(tuple(to_db_value(v) for v in row))
                if len(chunk) >= batch:
                    cur.executemany(insert_sql, chunk)
                    total += len(chunk)
                    chunk = []
                    log(f"  已写入 {total:,} 行 ...")
            if chunk:
                cur.executemany(insert_sql, chunk)
                total += len(chunk)
        conn.commit()
    except Exception as exc:
        try:
            conn.rollback()
        except Exception:
            pass
        low = str(exc).lower()
        if any(k in low for k in ("too long", "truncat", "太长了", "data too long")):
            # 长度截断类错误:给出可操作的提示再抛出
            raise RuntimeError(
                f"写入失败:{exc}\n"
                "提示:目标表字符串列长度不足。若目标是数据库中已存在的表,"
                "请将对应列改宽或改为 TEXT;若由本工具自动建表,"
                "请使用最新版本(自动建表已统一使用 TEXT 类型)。"
            ) from exc
        raise
    return total


# ===========================================================================
# Qt(PySide6)界面层:由构建脚本拼接到 data_transformer.py
# ===========================================================================

LOG_COLORS = {
    "info": "#3a4a5c", "ok": GREEN, "warn": "#b8741a", "err": RED,
    "head": PRIMARY, "section": TEAL, "db": INDIGO,
}

QSS = f"""
* {{ font-family: "Microsoft YaHei"; font-size: 10pt; }}
QMainWindow, QWidget {{ background: {BG}; color: {TEXT}; }}
QLabel#appTitle {{ background: {NAVY}; color: white; font-size: 15pt;
                   font-weight: bold; padding: 10px 16px; }}
QFrame#bar {{ background: #e8f0fb; border-radius: 6px; }}
QLabel#stepIn {{ background: {INDIGO}; color: white; font-weight: bold;
                 padding: 6px 14px; border-radius: 4px; }}
QLabel#stepOut {{ background: {CORAL}; color: white; font-weight: bold;
                  padding: 6px 14px; border-radius: 4px; }}
QLabel#arrow {{ color: {TEAL}; font-weight: bold; font-size: 11pt; }}
QLabel#muted {{ color: {MUTED}; }}
QLabel#chip {{ background: {STATUS_OFF}; color: white; padding: 4px 12px;
               border-radius: 10px; }}
QLabel#chip[state="ok"] {{ background: {STATUS_OK}; }}
QLabel#chip[state="busy"] {{ background: {ORANGE}; }}
QGroupBox {{ background: {CARD}; border: 2px solid {BORDER}; border-radius: 8px;
             margin-top: 14px; padding: 8px; font-weight: bold; }}
QGroupBox::title {{ subcontrol-origin: margin; subcontrol-position: top left;
                    left: 12px; padding: 0 6px; background: {CARD}; color: {TEXT}; }}
QGroupBox[mode="SQL"] {{ border-color: {PRIMARY}; }}
QGroupBox[mode="SQL"]::title {{ color: {PRIMARY}; }}
QGroupBox[mode="Excel"] {{ border-color: {GREEN}; }}
QGroupBox[mode="Excel"]::title {{ color: {GREEN}; }}
QGroupBox[mode="JSON"] {{ border-color: {ORANGE}; }}
QGroupBox[mode="JSON"]::title {{ color: {ORANGE}; }}
QGroupBox[mode="CSV"] {{ border-color: {PURPLE}; }}
QGroupBox[mode="CSV"]::title {{ color: {PURPLE}; }}
QGroupBox#log {{ border-color: {BORDER}; }}
QGroupBox#log::title {{ color: {MUTED}; }}
QPushButton {{ background: #e7edf6; border: 1px solid {BORDER}; border-radius: 5px;
               padding: 6px 14px; }}
QPushButton:hover {{ background: #d5e2f5; }}
QPushButton:disabled {{ background: #f0f3f8; color: {STATUS_OFF}; }}
QPushButton#primary {{ background: {GREEN}; color: white; font-weight: bold;
                       padding: 8px 18px; }}
QPushButton#primary:hover {{ background: #21b36b; }}
QPushButton#blue {{ background: {PRIMARY}; color: white; }}
QPushButton#blue:hover {{ background: #3b82f6; }}
QPushButton#teal {{ background: {TEAL}; color: white; }}
QPushButton#teal:hover {{ background: #12b6c4; }}
QPushButton#amber {{ background: {AMBER}; color: white; }}
QPushButton#amber:hover {{ background: #f7b64e; }}
QPushButton#indigo {{ background: {INDIGO}; color: white; }}
QPushButton#indigo:hover {{ background: #6f7deb; }}
QPushButton#warn {{ background: {ORANGE}; color: white; }}
QPushButton#warn:hover {{ background: #f09645; }}
QLineEdit, QComboBox, QSpinBox {{ background: white; border: 1px solid {BORDER};
                                  border-radius: 4px; padding: 4px 8px; }}
QLineEdit:focus, QComboBox:focus {{ border: 1px solid {PRIMARY}; }}
QComboBox::drop-down {{ border: none; width: 20px; }}
QComboBox QAbstractItemView {{ background: white;
                               selection-background-color: {PRIMARY};
                               selection-color: white; }}
QPlainTextEdit {{ background: #fbfcfe; border: 1px solid {BORDER};
                  border-radius: 4px; font-family: Consolas; font-size: 9pt; }}
QTreeWidget {{ background: #fbfcfe; border: 1px solid {BORDER}; border-radius: 4px; }}
QTreeWidget::item {{ height: 24px; }}
QTreeWidget::item:selected {{ background: #dbe7ff; color: {NAVY}; }}
QDialog {{ background: {CARD}; }}
"""


class Bridge(QObject):
    """worker 线程 → 主线程的信号桥。"""
    log_s = Signal(str, str)
    status_s = Signal(str, bool)
    tree_s = Signal(dict)
    sheets_s = Signal(list)
    excel_tree_s = Signal(dict)
    keys_s = Signal(list)
    csv_sections_s = Signal(list)
    error_s = Signal(str)
    info_s = Signal(str)
    done_s = Signal()


class HierTableSelect(QWidget):
    """层级勾选控件(表 → 字段):按钮 + 弹窗树形勾选。"""

    def __init__(self, parent=None, text="选择数据表与字段"):
        super().__init__(parent)
        self.title_text = text
        self.tables = {}
        self.checked = {}
        self.command = None
        self._updating = False
        lay = QHBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        self.button = QPushButton(f"{text}: (未加载)")
        self.button.clicked.connect(self._open)
        lay.addWidget(self.button, 1)

    # -- 对外接口 -----------------------------------------------------------
    def set_tables(self, tables, checked=None):
        """tables: {表名: [字段...]};checked: {表名: [勾选字段...]}(None=全选)。"""
        self.tables = {str(t): [str(c) for c in cols] for t, cols in tables.items()}
        if checked is None:
            self.checked = {t: set(cols) for t, cols in self.tables.items()}
        else:
            self.checked = {}
            for t, cols in checked.items():
                if t in self.tables:
                    self.checked[t] = {c for c in cols if c in self.tables[t]}
        self.button.setEnabled(True)
        self._refresh_text()

    def set_hint(self, text):
        self.tables = {}
        self.checked = {}
        self.button.setText(str(text))
        self.button.setEnabled(False)

    def get_selection(self):
        return {
            t: [c for c in self.tables[t] if c in self.checked.get(t, set())]
            for t in self.tables if self.checked.get(t)
        }

    def get_tables(self):
        return [t for t in self.tables if self.checked.get(t)]

    def set_command(self, fn):
        self.command = fn

    # -- 内部实现 -----------------------------------------------------------
    def _refresh_text(self):
        if not self.tables:
            self.button.setText(f"{self.title_text}: (未加载)")
            return
        n_tables = len(self.get_tables())
        n_cols = sum(len(cs) for cs in self.checked.values())
        self.button.setText(f"{self.title_text}: {n_tables}/{len(self.tables)} 表 · {n_cols} 字段")

    def _rebuild_tree(self):
        self._updating = True
        tree = self._tree
        tree.clear()
        for table, cols in self.tables.items():
            top = QTreeWidgetItem([table])
            # 注意:不要加 Qt.ItemIsAutoTristate。它会让 Qt 在子节点变化时
            # 自动重算父节点三态并再次触发 itemChanged,与下面的手动同步逻辑
            # 互相覆盖,导致用户点击字段后勾选状态被立刻还原。
            top.setFlags(top.flags() | Qt.ItemIsUserCheckable)
            sel = self.checked.get(table, set())
            if sel == set(cols):
                state = Qt.Checked
            elif sel:
                state = Qt.PartiallyChecked
            else:
                state = Qt.Unchecked
            top.setCheckState(0, state)
            top.setData(0, Qt.ItemDataRole.UserRole, ("t", table))
            tree.addTopLevelItem(top)
            for col in cols:
                child = QTreeWidgetItem([col])
                child.setFlags(child.flags() | Qt.ItemIsUserCheckable)
                child.setCheckState(0, Qt.Checked if col in sel else Qt.Unchecked)
                child.setData(0, Qt.ItemDataRole.UserRole, ("c", table, col))
                top.addChild(child)
            top.setExpanded(True)
        self._updating = False

    def _on_item_changed(self, item, _col):
        if self._updating or self._tree is None:
            return
        self._updating = True
        kind = item.data(0, Qt.ItemDataRole.UserRole)
        if kind and kind[0] == "t":
            table = kind[1]
            state = item.checkState(0)
            if state == Qt.PartiallyChecked:
                state = Qt.Checked
            self.checked[table] = (set(self.tables[table]) if state == Qt.Checked
                                   else set())
            for i in range(item.childCount()):
                item.child(i).setCheckState(0, state)
        elif kind and kind[0] == "c":
            _, table, col = kind
            sel = set(self.checked.get(table, set()))
            if item.checkState(0) == Qt.Checked:
                sel.add(col)
            else:
                sel.discard(col)
            self.checked[table] = sel
            parent = item.parent()
            all_cols = set(self.tables[table])
            if sel == all_cols:
                parent.setCheckState(0, Qt.Checked)
            elif sel:
                parent.setCheckState(0, Qt.PartiallyChecked)
            else:
                parent.setCheckState(0, Qt.Unchecked)
        self._updating = False

    def _open(self):
        if not self.tables:
            return
        dlg = QDialog(self.window())
        dlg.setWindowTitle(self.title_text)
        dlg.resize(580, 500)
        lay = QVBoxLayout(dlg)
        hint = QLabel("第一层:数据表(勾选整表)   ·   第二层:字段(每张表独立勾选)")
        hint.setObjectName("muted")
        lay.addWidget(hint)
        self._tree = QTreeWidget()
        self._tree.setHeaderHidden(True)
        lay.addWidget(self._tree, 1)
        btn_row = QHBoxLayout()
        b_all = QPushButton("全部勾选"); b_all.setObjectName("teal")
        b_none = QPushButton("全部取消"); b_none.setObjectName("warn")
        b_ok = QPushButton("确定"); b_ok.setObjectName("primary")
        b_all.clicked.connect(self._check_all)
        b_none.clicked.connect(self._check_none)
        b_ok.clicked.connect(dlg.accept)
        btn_row.addWidget(b_all)
        btn_row.addWidget(b_none)
        btn_row.addStretch(1)
        btn_row.addWidget(b_ok)
        lay.addLayout(btn_row)
        self._rebuild_tree()
        self._tree.itemChanged.connect(self._on_item_changed)
        dlg.exec()
        self._tree = None
        self._refresh_text()
        if self.command:
            self.command(self.get_selection())

    def _check_all(self):
        self.checked = {t: set(cols) for t, cols in self.tables.items()}
        self._rebuild_tree()

    def _check_none(self):
        self.checked = {}
        self._rebuild_tree()


class App(QMainWindow):
    """Qt 主窗口:输入/输出双面板 + 数据库连接卡片 + 彩色主题。"""

    def __init__(self):
        super().__init__()
        self.setWindowTitle(APP_TITLE)
        self.resize(1160, 840)
        self.setMinimumSize(1040, 740)

        self.bridge = Bridge()
        self.conn = None
        self.adapter = ADAPTERS["MySQL"]
        self._busy = False
        self.selected_table = None
        self.selected_tables = []
        self.selected_columns = {}
        self.excel_selection = {}   # {Sheet名: [勾选字段...]}(Excel→SQL 层级勾选)
        self.excel_sheets = []
        self._auto_sheet = ""
        self._auto_table = ""

        self._build_vars()
        self._build_ui()
        self._connect_signals()
        self._load_config()
        self.on_input_mode_changed()
        self.on_output_mode_changed()

    # ------------------------------------------------------------------ 变量
    def _build_vars(self):
        self.input_mode = INPUT_LABELS["SQL"]
        self.output_mode = OUTPUT_LABELS["Excel"]
        self.db_type = "MySQL"
        self.host_text = "127.0.0.1"
        self.port_text = ""
        self.user_text = ""
        self.pwd_text = ""
        self.database_text = ""
        self.timeout_text = "10"
        self.in_file = ""
        self.in_sheet = ""
        self.in_key = ""
        self.in_section = ""
        self.in_delim = list(DELIM_LABELS)[0]
        self.in_enc = list(ENC_LABELS)[0]
        self.out_table = ""
        self.out_mode = MODE_APPEND
        self.out_sheet = ""
        self.out_delim = list(DELIM_LABELS)[0]
        self.out_enc = list(ENC_LABELS)[0]
        self.batch_text = str(DEFAULT_BATCH)

    # ------------------------------------------------------------------ UI
    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(10, 0, 10, 8)
        root.setSpacing(8)

        # 顶部横幅 + 九色彩带
        title = QLabel(APP_TITLE)
        title.setObjectName("appTitle")
        root.addWidget(title)
        strip = QWidget()
        strip.setFixedHeight(5)
        strip_lay = QHBoxLayout(strip)
        strip_lay.setContentsMargins(0, 0, 0, 0)
        strip_lay.setSpacing(0)
        for color in ACCENT_STRIP:
            seg = QFrame()
            seg.setStyleSheet(f"background: {color}; border: none;")
            strip_lay.addWidget(seg, 1)
        root.addWidget(strip)

        # 模式栏
        mode_bar = QFrame()
        mode_bar.setObjectName("bar")
        bar_lay = QHBoxLayout(mode_bar)
        step_in = QLabel("① 输入格式"); step_in.setObjectName("stepIn")
        self.input_combo = QComboBox()
        self.input_combo.addItems(list(INPUT_LABELS.values()))
        self.input_combo.setFixedWidth(200)
        arrow = QLabel("▶ 转换"); arrow.setObjectName("arrow")
        step_out = QLabel("② 输出格式"); step_out.setObjectName("stepOut")
        self.output_combo = QComboBox()
        self.output_combo.addItems(list(OUTPUT_LABELS.values()))
        self.output_combo.setFixedWidth(200)
        self.convert_btn = QPushButton("开始转换 ▶")
        self.convert_btn.setObjectName("primary")
        self.convert_btn.clicked.connect(self.on_convert)
        bar_lay.addWidget(step_in)
        bar_lay.addWidget(self.input_combo)
        bar_lay.addSpacing(14)
        bar_lay.addWidget(arrow)
        bar_lay.addSpacing(14)
        bar_lay.addWidget(step_out)
        bar_lay.addWidget(self.output_combo)
        bar_lay.addStretch(1)
        bar_lay.addWidget(self.convert_btn)
        root.addWidget(mode_bar)

        # 内容区
        content = QWidget()
        content_lay = QVBoxLayout(content)
        content_lay.setContentsMargins(0, 0, 0, 0)
        content_lay.setSpacing(8)
        self._build_input_card(content_lay)
        self._build_db_card(content_lay)
        self._build_output_card(content_lay)
        root.addWidget(content, 3)

        # 日志
        log_group = QGroupBox("④ 运行日志")
        log_group.setObjectName("log")
        log_lay = QVBoxLayout(log_group)
        log_head = QHBoxLayout()
        log_head.addStretch(1)
        clear_btn = QPushButton("一键清空")
        clear_btn.setObjectName("blue")
        clear_btn.clicked.connect(self.on_clear_log)
        log_head.addWidget(clear_btn)
        log_lay.addLayout(log_head)
        self.log_text = QPlainTextEdit()
        self.log_text.setReadOnly(True)
        log_lay.addWidget(self.log_text, 1)
        root.addWidget(log_group, 2)

        # 状态胶囊
        self.status_chip = QLabel("未连接数据库 | 就绪")
        self.status_chip.setObjectName("chip")
        self.status_chip.setProperty("state", "idle")
        root.addWidget(self.status_chip)

    def _build_input_card(self, parent_lay):
        self.input_card = QGroupBox("① 输入设置")
        self.input_card.setProperty("mode", "SQL")
        lay = QVBoxLayout(self.input_card)

        # ---- 输入 = SQL:层级勾选(第一层表,第二层字段) ----
        self.src_sql_frame = QWidget()
        sql_lay = QGridLayout(self.src_sql_frame)
        sql_lay.setContentsMargins(0, 0, 0, 0)
        sql_lay.addWidget(QLabel("表/字段:"), 0, 0)
        self.tree_select = HierTableSelect(text="选择数据表与字段(层级勾选)")
        self.tree_select.set_command(self.on_tree_confirmed)
        sql_lay.addWidget(self.tree_select, 0, 1)
        refresh_btn = QPushButton("刷新表")
        refresh_btn.setObjectName("teal")
        refresh_btn.clicked.connect(self.on_refresh_tables)
        sql_lay.addWidget(refresh_btn, 0, 2)
        hint = QLabel("第一层勾选数据表(可多选),第二层分别勾选每张表的导出字段")
        hint.setObjectName("muted")
        sql_lay.addWidget(hint, 1, 1, 1, 2)
        sql_lay.setColumnStretch(1, 1)
        lay.addWidget(self.src_sql_frame)

        # ---- 输入 = 文件 ----
        self.src_file_frame = QWidget()
        file_lay = QGridLayout(self.src_file_frame)
        file_lay.setContentsMargins(0, 0, 0, 0)
        file_lay.addWidget(QLabel("输入文件:"), 0, 0)
        self.in_file_edit = QLineEdit()
        file_lay.addWidget(self.in_file_edit, 0, 1)
        pick_btn = QPushButton("选择文件...")
        pick_btn.setObjectName("teal")
        pick_btn.clicked.connect(self.on_pick_input_file)
        file_lay.addWidget(pick_btn, 0, 2)
        file_lay.setColumnStretch(1, 1)
        lay.addWidget(self.src_file_frame)

        # 文件类型专属参数行(按输入格式切换显示)
        self.file_excel_opts = QWidget()
        ex_lay = QVBoxLayout(self.file_excel_opts)
        ex_lay.setContentsMargins(0, 0, 0, 0)
        ex_lay.setSpacing(4)
        row1 = QHBoxLayout()
        row1.addWidget(QLabel("工作表(单表转换):"))
        self.in_sheet_combo = QComboBox()
        self.in_sheet_combo.setMinimumWidth(320)
        row1.addWidget(self.in_sheet_combo, 1)
        row1.addStretch(1)
        ex_lay.addLayout(row1)
        row2 = QHBoxLayout()
        self.excel_tree_label = QLabel("SQL 导入(层级勾选):")
        self.excel_tree_label.setObjectName("muted")
        row2.addWidget(self.excel_tree_label)
        self.excel_tree = HierTableSelect(text="选择工作表与字段")
        self.excel_tree.set_command(self.on_excel_tree_confirmed)
        row2.addWidget(self.excel_tree, 1)
        ex_lay.addLayout(row2)
        lay.addWidget(self.file_excel_opts)

        self.file_json_opts = QWidget()
        js_lay = QHBoxLayout(self.file_json_opts)
        js_lay.setContentsMargins(0, 0, 0, 0)
        js_lay.addWidget(QLabel("一级键(表名):"))
        self.in_key_combo = QComboBox()
        self.in_key_combo.setMinimumWidth(320)
        js_lay.addWidget(self.in_key_combo, 1)
        js_lay.addStretch(1)
        lay.addWidget(self.file_json_opts)

        self.file_csv_opts = QWidget()
        csv_lay = QHBoxLayout(self.file_csv_opts)
        csv_lay.setContentsMargins(0, 0, 0, 0)
        csv_lay.addWidget(QLabel("分节(Sheet):"))
        self.in_section_combo = QComboBox()
        self.in_section_combo.setMinimumWidth(220)
        csv_lay.addWidget(self.in_section_combo, 1)
        csv_lay.addWidget(QLabel("分隔符:"))
        self.in_delim_combo = QComboBox()
        self.in_delim_combo.addItems(list(DELIM_LABELS))
        csv_lay.addWidget(self.in_delim_combo)
        csv_lay.addWidget(QLabel("编码:"))
        self.in_enc_combo = QComboBox()
        self.in_enc_combo.addItems(list(ENC_LABELS))
        csv_lay.addWidget(self.in_enc_combo)
        csv_lay.addStretch(1)
        lay.addWidget(self.file_csv_opts)
        parent_lay.addWidget(self.input_card)

    def _build_db_card(self, parent_lay):
        self.db_card = QGroupBox("数据库连接(输入或输出为 SQL 时展开)")
        self.db_card.setProperty("mode", "SQL")
        db_lay = QGridLayout(self.db_card)

        db_lay.addWidget(QLabel("类型:"), 0, 0)
        self.db_type_combo = QComboBox()
        self.db_type_combo.addItems(list(ADAPTERS))
        self.db_type_combo.setFixedWidth(120)
        db_lay.addWidget(self.db_type_combo, 0, 1)

        db_lay.addWidget(QLabel("主机:"), 0, 2)
        self.host_edit = QLineEdit("127.0.0.1")
        self.host_edit.setFixedWidth(150)
        db_lay.addWidget(self.host_edit, 0, 3)

        db_lay.addWidget(QLabel("端口:"), 0, 4)
        self.port_edit = QLineEdit("")
        self.port_edit.setFixedWidth(80)
        db_lay.addWidget(self.port_edit, 0, 5)

        db_lay.addWidget(QLabel("用户名:"), 0, 6)
        self.user_edit = QLineEdit("")
        self.user_edit.setFixedWidth(130)
        db_lay.addWidget(self.user_edit, 0, 7)

        db_lay.addWidget(QLabel("密码:"), 1, 0)
        self.pwd_edit = QLineEdit("")
        self.pwd_edit.setEchoMode(QLineEdit.Password)
        self.pwd_edit.setFixedWidth(120)
        db_lay.addWidget(self.pwd_edit, 1, 1)

        db_lay.addWidget(QLabel("数据库:"), 1, 2)
        self.database_edit = QLineEdit("")
        self.database_edit.setFixedWidth(150)
        db_lay.addWidget(self.database_edit, 1, 3)

        db_lay.addWidget(QLabel("超时(秒):"), 1, 4)
        self.timeout_edit = QLineEdit("10")
        self.timeout_edit.setFixedWidth(80)
        db_lay.addWidget(self.timeout_edit, 1, 5)

        btn_box = QHBoxLayout()
        test_btn = QPushButton("测试连接"); test_btn.setObjectName("amber")
        test_btn.clicked.connect(self.on_test_connection)
        conn_btn = QPushButton("连接并加载表"); conn_btn.setObjectName("blue")
        conn_btn.clicked.connect(self.on_connect)
        disc_btn = QPushButton("断开")
        disc_btn.clicked.connect(self.disconnect)
        btn_box.addWidget(test_btn)
        btn_box.addWidget(conn_btn)
        btn_box.addWidget(disc_btn)
        btn_box.addStretch(1)
        db_lay.addLayout(btn_box, 1, 6, 1, 2)
        db_lay.setColumnStretch(3, 1)
        parent_lay.addWidget(self.db_card)

    def _build_output_card(self, parent_lay):
        self.output_card = QGroupBox("③ 输出设置")
        self.output_card.setProperty("mode", "Excel")
        lay = QVBoxLayout(self.output_card)

        # ---- 输出 = SQL ----
        self.out_db_frame = QWidget()
        out_db_lay = QHBoxLayout(self.out_db_frame)
        out_db_lay.setContentsMargins(0, 0, 0, 0)
        out_db_lay.addWidget(QLabel("目标表名:"))
        self.out_table_edit = QLineEdit("")
        self.out_table_edit.setFixedWidth(240)
        out_db_lay.addWidget(self.out_table_edit)
        out_db_lay.addWidget(QLabel("写入模式:"))
        self.out_mode_combo = QComboBox()
        self.out_mode_combo.addItems(MODE_LABELS)
        self.out_mode_combo.setMinimumWidth(300)
        out_db_lay.addWidget(self.out_mode_combo)
        out_db_lay.addWidget(QLabel("每批行数:"))
        self.batch_spin = QSpinBox()
        self.batch_spin.setRange(1, 1000000)
        self.batch_spin.setValue(DEFAULT_BATCH)
        out_db_lay.addWidget(self.batch_spin)
        out_db_lay.addStretch(1)
        lay.addWidget(self.out_db_frame)

        # ---- 输出 = 文件 ----
        self.out_file_frame = QWidget()
        out_file_lay = QHBoxLayout(self.out_file_frame)
        out_file_lay.setContentsMargins(0, 0, 0, 0)
        out_file_lay.addWidget(QLabel("Sheet名/JSON键:"))
        self.out_sheet_edit = QLineEdit("")
        self.out_sheet_edit.setFixedWidth(240)
        out_file_lay.addWidget(self.out_sheet_edit)
        hint = QLabel("(Excel 用 Sheet 名,JSON 用一级键,CSV 忽略此项)")
        hint.setObjectName("muted")
        out_file_lay.addWidget(hint)
        out_file_lay.addStretch(1)
        lay.addWidget(self.out_file_frame)

        self.out_csv_opts = QWidget()
        out_csv_lay = QHBoxLayout(self.out_csv_opts)
        out_csv_lay.setContentsMargins(0, 0, 0, 0)
        out_csv_lay.addWidget(QLabel("分隔符:"))
        self.out_delim_combo = QComboBox()
        self.out_delim_combo.addItems(list(DELIM_LABELS))
        out_csv_lay.addWidget(self.out_delim_combo)
        out_csv_lay.addWidget(QLabel("编码:"))
        self.out_enc_combo = QComboBox()
        self.out_enc_combo.addItems(list(ENC_LABELS))
        out_csv_lay.addWidget(self.out_enc_combo)
        hint = QLabel("(SQL→CSV 时:目录名 = 数据库名,每个表一个 .csv 文件)")
        hint.setObjectName("muted")
        out_csv_lay.addWidget(hint)
        out_csv_lay.addStretch(1)
        lay.addWidget(self.out_csv_opts)
        parent_lay.addWidget(self.output_card)

    def _connect_signals(self):
        b = self.bridge
        b.log_s.connect(self._on_log)
        b.status_s.connect(self._on_status)
        b.tree_s.connect(self._on_tree_data)
        b.sheets_s.connect(self._on_sheets)
        b.excel_tree_s.connect(self._on_excel_tree_data)
        b.keys_s.connect(self._on_keys)
        b.csv_sections_s.connect(self._on_csv_sections)
        b.error_s.connect(lambda m: QMessageBox.critical(self, "错误", m))
        b.info_s.connect(lambda m: QMessageBox.information(self, "提示", m))
        b.done_s.connect(self._on_done)
        self.input_combo.currentTextChanged.connect(self.on_input_mode_changed)
        self.output_combo.currentTextChanged.connect(self.on_output_mode_changed)
        self.db_type_combo.currentTextChanged.connect(self.on_db_type_changed)
        self.in_sheet_combo.currentTextChanged.connect(self.on_in_sheet_selected)
        self.in_key_combo.currentTextChanged.connect(self.on_in_key_selected)
        self.in_section_combo.currentTextChanged.connect(self.on_in_section_selected)

    # ------------------------------------------------------- 线程 / 消息桥
    def log(self, message, level="info"):
        self.bridge.log_s.emit(str(message), level)

    def set_status(self, message, connected=False):
        self.bridge.status_s.emit(str(message), bool(connected))

    def _repolish(self, widget):
        widget.style().unpolish(widget)
        widget.style().polish(widget)

    def _set_chip_state(self, state):
        self.status_chip.setProperty("state", state)
        self._repolish(self.status_chip)

    def _on_log(self, message, level):
        color = LOG_COLORS.get(level, LOG_COLORS["info"])
        esc = html.escape(message)
        self.log_text.appendHtml(f'<span style="color:{color};">{esc}</span>')

    def _on_status(self, message, connected):
        self.status_chip.setText(message)
        self._set_chip_state("ok" if connected else "idle")

    def on_clear_log(self):
        self.log_text.clear()
        self.log("日志已清空", "info")

    def run_task(self, fn):
        if self._busy:
            QMessageBox.information(self, "提示", "已有任务正在执行,请等待完成后再试。")
            return False
        self._busy = True
        self.status_chip.setText("运行中...")
        self._set_chip_state("busy")

        def worker():
            try:
                fn()
            except Exception as exc:  # noqa: BLE001
                self.log(f"发生错误: {exc}", "err")
                self.log(traceback.format_exc(), "err")
                self.bridge.error_s.emit(str(exc))
            finally:
                self.bridge.done_s.emit()

        threading.Thread(target=worker, daemon=True).start()
        return True

    def _on_done(self):
        self._busy = False
        connected = self.conn is not None
        self.status_chip.setText(
            f"{'已连接数据库' if connected else '未连接数据库'} | 就绪")
        self._set_chip_state("ok" if connected else "idle")

    def _on_tree_data(self, table_columns):
        self.tree_select.set_tables(table_columns)
        self.on_tree_confirmed(self.tree_select.get_selection())

    def _on_sheets(self, names):
        self.in_sheet_combo.blockSignals(True)
        self.in_sheet_combo.clear()
        self.in_sheet_combo.addItems(names)
        self.in_sheet_combo.blockSignals(False)
        if names:
            self.in_sheet_combo.setCurrentIndex(0)
            self.on_in_sheet_selected()

    def _on_excel_tree_data(self, sheet_headers):
        """Excel 文件解析完成:{Sheet名: [表头字段...]}。"""
        self.excel_tree.set_tables(sheet_headers)
        self.on_excel_tree_confirmed(self.excel_tree.get_selection())

    def _on_keys(self, keys):
        self.in_key_combo.blockSignals(True)
        self.in_key_combo.clear()
        self.in_key_combo.addItems(keys)
        self.in_key_combo.blockSignals(False)
        if keys:
            self.in_key_combo.setCurrentIndex(0)
            self.on_in_key_selected()

    def _on_csv_sections(self, names):
        self.in_section_combo.blockSignals(True)
        self.in_section_combo.clear()
        if names:
            self.in_section_combo.addItems(names)
        else:
            self.in_section_combo.addItem("(普通CSV,无分节)")
        self.in_section_combo.blockSignals(False)
        self.in_section_combo.setCurrentIndex(0)
        self.on_in_section_selected()

    # ------------------------------------------------------------- 模式切换
    @staticmethod
    def _key_of(value, mapping):
        return next((k for k, v in mapping.items() if v == value), None)

    def input_key(self):
        return self._key_of(self.input_combo.currentText(), INPUT_LABELS)

    def output_key(self):
        return self._key_of(self.output_combo.currentText(), OUTPUT_LABELS)

    def on_input_mode_changed(self, *_):
        key = self.input_key()
        self.input_card.setProperty("mode", key)
        self._repolish(self.input_card)
        sql_mode = key == "SQL"
        self.src_sql_frame.setVisible(sql_mode)
        self.src_file_frame.setVisible(not sql_mode)
        self.file_excel_opts.setVisible(key == "Excel")
        self.file_json_opts.setVisible(key == "JSON")
        self.file_csv_opts.setVisible(key == "CSV")
        self._sync_db_card()
        self._sync_excel_tree()
        self._auto_fill_output_names()
        self.log(f"输入格式切换为: {INPUT_LABELS[key]}", "info")

    def on_output_mode_changed(self, *_):
        key = self.output_key()
        self.output_card.setProperty("mode", key)
        self._repolish(self.output_card)
        sql_mode = key == "SQL"
        self.out_db_frame.setVisible(sql_mode)
        self.out_file_frame.setVisible(not sql_mode)
        self.out_csv_opts.setVisible(key == "CSV")
        self._sync_db_card()
        self._sync_excel_tree()
        self._auto_fill_output_names()
        self.log(f"输出格式切换为: {OUTPUT_LABELS[key]}", "info")

    def _sync_db_card(self):
        need_db = self.input_key() == "SQL" or self.output_key() == "SQL"
        self.db_card.setVisible(need_db)

    def _auto_fill_output_names(self):
        in_key = self.input_key()
        base = ""
        if in_key == "SQL":
            if len(self.selected_tables) > 1:
                cur = self.out_sheet_edit.text()
                if not cur or cur == self._auto_sheet:
                    self.out_sheet_edit.setText("")
                    self._auto_sheet = ""
                cur = self.out_table_edit.text()
                if not cur or cur == self._auto_table:
                    self.out_table_edit.setText("")
                    self._auto_table = ""
                return
            base = (self.selected_table or "").split(".")[-1]
        elif in_key == "Excel":
            base = self.in_sheet_combo.currentText().strip() or Path(
                self.in_file_edit.text()).stem
        elif in_key == "JSON":
            base = self.in_key_combo.currentText().strip() or Path(
                self.in_file_edit.text()).stem
        elif in_key == "CSV":
            sec = self.in_section_combo.currentText().strip()
            if sec and not sec.startswith("("):
                base = sec
            else:
                base = Path(self.in_file_edit.text()).stem
        if not base:
            return
        cur = self.out_sheet_edit.text()
        if not cur or cur == self._auto_sheet:
            self.out_sheet_edit.setText(base)
            self._auto_sheet = base
        cur = self.out_table_edit.text()
        if not cur or cur == self._auto_table:
            self.out_table_edit.setText(base)
            self._auto_table = base

    # ------------------------------------------------------------- 连接管理
    def get_cfg(self):
        return {
            "host": self.host_edit.text().strip(),
            "port": self.port_edit.text().strip() or str(self.adapter.DEFAULT_PORT),
            "user": self.user_edit.text().strip(),
            "password": self.pwd_edit.text(),
            "database": self.database_edit.text().strip(),
            "charset": "utf8mb4",
            "timeout": self.timeout_edit.text().strip() or "10",
        }

    def connect_db(self):
        cfg = self.get_cfg()
        if not cfg["host"] or not cfg["user"] or not cfg["database"]:
            raise ValueError("请填写【主机】【用户名】【数据库】后再连接")
        self._disconnect_conn()
        self.conn = self.adapter.connect(cfg)
        self._save_config()
        return self.adapter.server_version(self.conn)

    def _load_config(self):
        try:
            data = json.loads(config_file_path().read_text(encoding="utf-8"))
        except Exception:
            return
        if isinstance(data, dict):
            if data.get("password"):
                self.pwd_edit.setText(str(data["password"]))
            if data.get("database"):
                self.database_edit.setText(str(data["database"]))
            self.log(f"已载入记忆的连接配置(密码/数据库名): {config_file_path()}", "info")

    def _save_config(self):
        try:
            path = config_file_path()
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_text(
                json.dumps({"password": self.pwd_edit.text(),
                            "database": self.database_edit.text()},
                           ensure_ascii=False, indent=2),
                encoding="utf-8")
        except Exception:
            pass

    def _disconnect_conn(self):
        if self.conn is not None:
            try:
                self.conn.close()
            except Exception:
                pass
            self.conn = None

    def disconnect(self):
        self._disconnect_conn()
        self.status_chip.setText("未连接数据库 | 就绪")
        self._set_chip_state("idle")

    def _ping(self):
        try:
            if self.adapter.NAME == "MySQL":
                self.conn.ping(reconnect=False)
                return True
            return not self.conn.closed
        except Exception:
            return False

    def ensure_conn(self):
        try:
            if self.conn is not None and self._ping():
                return self.conn
        except Exception:
            pass
        self._disconnect_conn()
        version = self.connect_db()
        self.log(f"已自动连接 {self.adapter.NAME} {version}", "ok")
        self.set_status(f"已连接 {self.adapter.NAME} {version}", connected=True)
        return self.conn

    # ------------------------------------------------------------- UI 事件
    def on_db_type_changed(self, *_):
        name = self.db_type_combo.currentText()
        self.adapter = ADAPTERS[name]
        self.disconnect()
        self.log(f"已切换到 {name}(端口留空时自动使用默认端口 "
                 f"{self.adapter.DEFAULT_PORT})", "info")

    def on_test_connection(self):
        def work():
            version = self.connect_db()
            self.log(f"连接成功: {self.adapter.NAME} {version}", "ok")
            self.bridge.info_s.emit(f"连接成功!\n{self.adapter.NAME} {version}")
            self._disconnect_conn()
        self.run_task(work)

    def _load_table_columns(self, conn, tables):
        table_columns = {}
        for table in tables:
            try:
                table_columns[table] = self.adapter.list_columns(conn, table)
            except Exception as exc:
                self.log(f"读取表【{table}】字段失败: {exc}", "warn")
                table_columns[table] = []
        return table_columns

    def on_connect(self):
        def work():
            version = self.connect_db()
            tables = self.adapter.list_tables(self.conn)
            self.log(f"连接成功: {self.adapter.NAME} {version},共 {len(tables)} 张表", "ok")
            self.log("正在加载字段结构 ...", "info")
            table_columns = self._load_table_columns(self.conn, tables)
            self.bridge.tree_s.emit(table_columns)
            self.set_status(f"已连接 {self.adapter.NAME} {version}", connected=True)
        self.run_task(work)

    def on_refresh_tables(self):
        def work():
            conn = self.ensure_conn()
            tables = self.adapter.list_tables(conn)
            self.log(f"共发现 {len(tables)} 张表,正在加载字段结构 ...", "info")
            table_columns = self._load_table_columns(conn, tables)
            self.bridge.tree_s.emit(table_columns)
        self.run_task(work)

    def on_tree_confirmed(self, selection):
        self.selected_columns = {t: list(cols) for t, cols in selection.items() if cols}
        self.selected_tables = [t for t in self.selected_columns]
        self.selected_table = self.selected_tables[0] if self.selected_tables else None
        self._auto_fill_output_names()
        if self.selected_tables:
            self.log(f"已选择 {len(self.selected_tables)} 张表、"
                     f"{sum(len(c) for c in self.selected_columns.values())} 个字段", "ok")
        else:
            self.log("未勾选任何表/字段", "warn")

    def on_pick_input_file(self):
        key = self.input_key()
        filters = {
            "Excel": "Excel 工作簿 (*.xlsx);;所有文件 (*.*)",
            "JSON": "JSON 文件 (*.json);;所有文件 (*.*)",
            "CSV": "CSV 文件 (*.csv);;所有文件 (*.*)",
        }[key]
        path, _ = QFileDialog.getOpenFileName(
            self, f"选择输入文件 ({INPUT_LABELS[key]})", "", filters)
        if not path:
            return
        self.in_file_edit.setText(path)
        self._auto_fill_output_names()

        if key == "Excel":
            def work():
                wb = load_workbook(path, read_only=True, data_only=True)
                try:
                    names = wb.sheetnames
                    sheet_headers = {}
                    for name in names:
                        ws = wb[name]
                        it = ws.iter_rows(values_only=True)
                        try:
                            raw = next(it)
                        except StopIteration:
                            raw = None
                        sheet_headers[name] = parse_header(raw) if raw else []
                finally:
                    wb.close()
                self.log(f"工作簿包含工作表: {', '.join(names)}", "info")
                self.bridge.sheets_s.emit(list(names))
                self.bridge.excel_tree_s.emit(sheet_headers)
            self.run_task(work)
        elif key == "JSON":
            def work():
                with open(path, "r", encoding="utf-8") as fp:
                    doc = json.load(fp)
                keys = list(doc.keys()) if isinstance(doc, dict) else []
                self.log(f"JSON 包含一级键: {', '.join(map(str, keys))}", "info")
                self.bridge.keys_s.emit(keys)
            self.run_task(work)
        elif key == "CSV":
            def work():
                delim = DELIM_LABELS[self.in_delim_combo.currentText()]
                enc = ENC_LABELS[self.in_enc_combo.currentText()]
                sections = read_csv_sections(path, delim, enc)
                if sections is None:
                    header = read_csv_header(path, delim, enc)
                    self.log(f"普通 CSV,表头列: {header}", "info")
                    self.bridge.csv_sections_s.emit([])
                else:
                    self.log(f"分节 CSV,共 {len(sections)} 个分节(Sheet): "
                             f"{', '.join(sections)}", "section")
                    self.bridge.csv_sections_s.emit(list(sections))
            self.run_task(work)

    def on_in_sheet_selected(self, *_):
        self._auto_fill_output_names()

    def on_in_key_selected(self, *_):
        self._auto_fill_output_names()

    def on_in_section_selected(self, *_):
        self._auto_fill_output_names()

    def on_excel_tree_confirmed(self, selection):
        """Excel→SQL 层级勾选确认:selection = {Sheet名: [勾选字段...]}。"""
        self.excel_selection = {s: list(cols) for s, cols in selection.items() if cols}
        self.excel_sheets = [s for s in self.excel_selection]
        if self.excel_sheets:
            cur = self.out_table_edit.text()
            if not cur or cur == self._auto_table:
                self.out_table_edit.setText(self.excel_sheets[0])
                self._auto_table = self.excel_sheets[0]
        if self.excel_selection:
            self.log(f"Excel 已选择 {len(self.excel_sheets)} 个 Sheet、"
                     f"{sum(len(c) for c in self.excel_selection.values())} 个字段", "ok")
        else:
            self.log("Excel 未勾选任何 Sheet/字段", "warn")

    def _sync_excel_tree(self):
        """仅当 输入=Excel 且 输出=SQL 时显示层级勾选控件。"""
        visible = self.input_key() == "Excel" and self.output_key() == "SQL"
        self.excel_tree.setVisible(visible)
        self.excel_tree_label.setVisible(visible)

    def ask_conflict_mode(self, existing_paths):
        """同名文件冲突纠错对话框,返回 'replace' / 'merge' / None(取消)。"""
        dlg = QDialog(self)
        dlg.setWindowTitle("检测到同名文件")
        lay = QVBoxLayout(dlg)
        lines = "\n".join(f"  • {p}" for p in existing_paths)
        label = QLabel(f"以下文件已存在:\n{lines}\n\n请选择处理方式:")
        label.setWordWrap(True)
        lay.addWidget(label)
        btn_row = QHBoxLayout()
        result = {"mode": None}

        def pick(mode):
            result["mode"] = mode
            dlg.accept()

        b1 = QPushButton("覆盖整个文件"); b1.setObjectName("warn")
        b1.clicked.connect(lambda: pick("replace"))
        b2 = QPushButton("合并写入(覆盖同名Sheet/JSON键;CSV追加/覆盖分节)")
        b2.setObjectName("blue")
        b2.clicked.connect(lambda: pick("merge"))
        b3 = QPushButton("取消")
        b3.clicked.connect(lambda: pick(None))
        btn_row.addWidget(b1)
        btn_row.addWidget(b2)
        btn_row.addWidget(b3)
        lay.addLayout(btn_row)
        dlg.exec()
        return result["mode"]

    def _ask_sheet_mode(self, sheet):
        """检测到数据库同名表时,询问 覆盖写入 / 追加写入 / 跳过该表。"""
        dlg = QDialog(self)
        dlg.setWindowTitle("检测到同名数据表")
        lay = QVBoxLayout(dlg)
        label = QLabel(f"数据库已存在与 Sheet 同名的表【{sheet}】。\n请选择写入方式:")
        label.setWordWrap(True)
        lay.addWidget(label)
        btn_row = QHBoxLayout()
        result = {"mode": "skip"}

        def pick(mode):
            result["mode"] = mode
            dlg.accept()

        b1 = QPushButton("覆盖写入(删除原表并重建导入)")
        b1.setObjectName("warn")
        b1.clicked.connect(lambda: pick("replace"))
        b2 = QPushButton("追加写入(保留原表,追加数据行)")
        b2.setObjectName("blue")
        b2.clicked.connect(lambda: pick("append"))
        b3 = QPushButton("跳过该表")
        b3.clicked.connect(lambda: pick("skip"))
        btn_row.addWidget(b1)
        btn_row.addWidget(b2)
        btn_row.addWidget(b3)
        lay.addLayout(btn_row)
        dlg.exec()
        return result["mode"]

    # ------------------------------------------------------------- 主转换流程
    def _ask_output_path(self, out_key, default_base):
        ext = {"Excel": ".xlsx", "JSON": ".json", "CSV": ".csv"}[out_key]
        label = OUTPUT_LABELS[out_key]
        initial = str(Path.home() / "Documents" / f"{safe_filename(default_base)}{ext}")
        path, _ = QFileDialog.getSaveFileName(
            self, f"保存输出文件 ({label})", initial, f"{label} (*{ext})")
        if not path:
            return None, None
        if os.path.exists(path):
            mode = self.ask_conflict_mode([path])
            if mode is None:
                self.log("已取消转换(存在同名文件)", "warn")
                return None, None
        else:
            mode = "replace"
        return path, mode

    def _default_output_base(self, in_key):
        if in_key == "SQL":
            return self.database_edit.text().strip() or (self.selected_table or "export")
        return Path(self.in_file_edit.text()).stem or "export"

    def _output_mode_key(self):
        label = self.out_mode_combo.currentText()
        if MODE_REPLACE in label:
            return "replace"
        if MODE_CREATE in label:
            return "create_if_missing"
        return "append"

    # ----------------------------------------------------- 多表 SQL 导出
    def _iter_sql_sources(self, conn, adapter, sources):
        for src in sources:
            self.log(f"读取数据表: {src['table']}", "head")
            header, rows, cur = load_sql_stream(
                conn, adapter, src["table"], src["columns"])
            yield src, header, rows, cur

    @staticmethod
    def _close_cursor(conn, cur):
        if cur is None:
            return
        try:
            cur.close()
        except Exception:
            pass
        try:
            conn.rollback()
        except Exception:
            pass

    def _run_multi_sql(self, conn, adapter, sources, out_key, output):
        total_all = 0
        if out_key == "Excel":
            def tables():
                for src, header, rows, cur in self._iter_sql_sources(conn, adapter, sources):
                    try:
                        yield (safe_sheet_title(src["table"]), header, rows)
                    finally:
                        self._close_cursor(conn, cur)
            total_all = write_excel_multi_file(
                output["path"], tables(), mode=output["mode"], log=self.log)
            self.log(f"✅ Excel 导出完成: {output['path']}"
                     f"({len(sources)} 个 Sheet,共 {total_all:,} 行)", "ok")
            self.bridge.info_s.emit(
                f"导出完成,{len(sources)} 个 Sheet,共 {total_all:,} 行:\n{output['path']}")
        elif out_key == "JSON":
            def tables():
                for src, header, rows, cur in self._iter_sql_sources(conn, adapter, sources):
                    try:
                        yield (src["table"], header, list(rows))
                    finally:
                        self._close_cursor(conn, cur)
            total_all = write_json_multi_file(
                output["path"], tables(), mode=output["mode"], log=self.log)
            self.log(f"✅ JSON 导出完成: {output['path']}"
                     f"({len(sources)} 个一级键,共 {total_all:,} 行)", "ok")
            self.bridge.info_s.emit(
                f"导出完成,{len(sources)} 个一级键,共 {total_all:,} 行:\n{output['path']}")
        elif out_key == "CSV":
            # 目录模式:目录名 = 数据库名,每个表一个 {表名}.csv
            os.makedirs(os.path.dirname(output["paths"][0]), exist_ok=True)
            for idx, (src, header, rows, cur) in enumerate(
                    self._iter_sql_sources(conn, adapter, sources)):
                try:
                    n = write_csv_file(
                        output["paths"][idx], header, rows,
                        delimiter=output["delimiter"],
                        encoding=output["encoding"],
                        mode=output["mode"], log=self.log)
                    total_all += n
                    self.log(f"✅ {output['paths'][idx]}: 共 {n:,} 行", "ok")
                finally:
                    self._close_cursor(conn, cur)
            self.bridge.info_s.emit(
                f"导出完成,共 {total_all:,} 行,{len(sources)} 个 CSV 文件")
        elif out_key == "SQL":
            mode, batch = output["mode"], output["batch"]
            conn_out = adapter.connect(self.get_cfg())
            try:
                for src, header, rows, cur in self._iter_sql_sources(
                        conn, adapter, sources):
                    sample, buffered = [], []
                    for row in rows:
                        buffered.append(list(row))
                        if len(sample) < SAMPLE_SIZE:
                            sample.append(list(row))
                        if len(buffered) >= SAMPLE_SIZE:
                            break

                    def chained():
                        for r in buffered:
                            yield r
                        for r in rows:
                            yield list(r)

                    target = src["table"]
                    n = write_sql_table(conn_out, adapter, target, header,
                                        chained(), mode, batch, self.log,
                                        sample=sample)
                    total_all += n
                    self._close_cursor(conn, cur)
                    self.log(f"✅ 表【{target}】写入 {n:,} 行", "ok")
            finally:
                conn_out.close()
            self.log(f"✅ 多表写入完成: 共 {total_all:,} 行", "ok")
            self.bridge.info_s.emit(
                f"写入完成,共 {total_all:,} 行 → {len(sources)} 张表")
        return total_all

    def _run_excel_to_sql(self, conn, adapter, path, selection, decisions, batch):
        """Excel→SQL:层级勾选 Sheet/字段;目标表 = Sheet 名。

        无同名表 → 按 Sheet 名新建(create_if_missing);
        同名表 → 按主线程预检的 decisions(覆盖 replace / 追加 append / 跳过 skip)。
        """
        total_all = 0
        for sheet, cols in selection.items():
            mode = decisions.get(sheet)
            if mode == "skip":
                self.log(f"⏭ 已按选择跳过同名表【{sheet}】", "warn")
                continue
            self.log(f"读取工作表: {sheet}(字段: {', '.join(cols)})", "head")
            header, rows = load_excel_stream(path, sheet)
            if list(cols) != header:
                header, rows = filter_columns(header, rows, cols)
            mode = mode or "create_if_missing"
            n = write_sql_table(conn, adapter, sheet, header, rows, mode, batch,
                                self.log)
            total_all += n
            self.log(f"✅ 表【{sheet}】写入 {n:,} 行(模式: {mode})", "ok")
        self.bridge.info_s.emit(
            f"导入完成,共 {total_all:,} 行 → {len(selection)} 个 Sheet")
        return total_all

    def on_convert(self):
        if self._busy:
            QMessageBox.information(self, "提示", "已有任务正在执行,请等待完成后再试。")
            return
        in_key = self.input_key()
        out_key = self.output_key()
        adapter = self.adapter
        batch = self.batch_spin.value()

        # ---------- 校验输入 ----------
        sources = []
        if in_key == "SQL":
            selection = {t: cols for t, cols in self.selected_columns.items() if cols}
            if not selection:
                QMessageBox.warning(
                    self, "提示",
                    "请连接数据库,并在“选择数据表与字段”层级下拉框中勾选表和字段")
                return
            sources = [{"table": t, "columns": cols} for t, cols in selection.items()]
        elif in_key == "Excel":
            if not self.in_file_edit.text() or not os.path.isfile(self.in_file_edit.text()):
                QMessageBox.warning(self, "提示", "请先选择有效的 Excel 输入文件")
                return
            if out_key == "SQL":
                if not self.excel_selection:
                    QMessageBox.warning(
                        self, "提示",
                        "请在“选择工作表与字段”层级下拉框中勾选要导入的 Sheet 与字段")
                    return
            elif not self.in_sheet_combo.currentText():
                QMessageBox.warning(self, "提示", "请选择要读取的工作表")
                return
        elif in_key == "JSON":
            if not self.in_file_edit.text() or not os.path.isfile(self.in_file_edit.text()):
                QMessageBox.warning(self, "提示", "请先选择有效的 JSON 输入文件")
                return
            if not self.in_key_combo.currentText():
                QMessageBox.warning(self, "提示", "请选择 JSON 一级键(表名)")
                return
        elif in_key == "CSV":
            if not self.in_file_edit.text() or not os.path.isfile(self.in_file_edit.text()):
                QMessageBox.warning(self, "提示", "请先选择有效的 CSV 输入文件")
                return

        multi_sql = in_key == "SQL" and len(sources) > 1
        # SQL → CSV 一律使用目录模式:目录名 = 数据库名,每个表一个 {表名}.csv
        dir_csv = in_key == "SQL" and out_key == "CSV"
        # Excel → SQL:层级勾选 Sheet/字段,目标表 = Sheet 名
        excel_to_sql = in_key == "Excel" and out_key == "SQL"

        # ---------- 校验/询问输出 ----------
        output = {}
        if out_key == "SQL":
            if excel_to_sql:
                # Excel→SQL:目标表 = Sheet 名;同名表在主线程逐个询问 覆盖/追加/跳过
                try:
                    conn = self.ensure_conn()
                except Exception as exc:
                    QMessageBox.critical(self, "错误", f"数据库连接失败: {exc}")
                    return
                decisions = {}
                for sheet in self.excel_sheets:
                    if adapter.table_exists(conn, adapter.quote_ident(sheet)):
                        choice = self._ask_sheet_mode(sheet)
                        if choice in (None, "skip"):
                            decisions[sheet] = "skip"
                            self.log(f"已选择跳过同名表【{sheet}】", "warn")
                        else:
                            decisions[sheet] = choice
                output = {"decisions": decisions, "batch": batch}
            else:
                table_out = self.out_table_edit.text().strip()
                if multi_sql:
                    if table_out:
                        QMessageBox.warning(
                            self, "提示", "多表拷贝时目标表名请留空(将按源表同名建表)")
                        return
                elif not table_out:
                    QMessageBox.warning(self, "提示", "请填写目标表名")
                    return
                mode = self._output_mode_key()
                if mode == "replace":
                    msg = (f"将先删除表【{table_out}】(若存在)再重建并写入,确定继续吗?"
                           if not multi_sql else
                           "多表同名拷贝:将先删除并重建每张勾选的源表再重新写入,确定继续吗?")
                    if QMessageBox.question(self, "危险操作", msg,
                                            QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
                        return
                output = {"table": table_out or None, "mode": mode, "batch": batch}
        elif out_key == "CSV" and dir_csv:
            # SQL → CSV:创建目录(目录名 = 数据库名),每个表一个 {表名}.csv
            default_dir = safe_filename(self._default_output_base(in_key))
            parent_dir = QFileDialog.getExistingDirectory(
                self, f"选择 CSV 输出的父目录(将创建目录 {default_dir},"
                      f"其中每个表一个 .csv 文件)")
            if not parent_dir:
                return
            target_dir = os.path.join(parent_dir, default_dir)
            delim = DELIM_LABELS[self.out_delim_combo.currentText()]
            enc = ENC_LABELS[self.out_enc_combo.currentText()]
            paths = [os.path.join(target_dir, f"{safe_filename(t['table'])}.csv")
                     for t in sources]
            existing = [p for p in paths if os.path.exists(p)]
            conflict_mode = "replace"
            if existing:
                conflict_mode = self.ask_conflict_mode(existing)
                if conflict_mode is None:
                    self.log("已取消转换(存在同名文件)", "warn")
                    return
            output = {"paths": paths, "mode": conflict_mode,
                      "delimiter": delim, "encoding": enc}
        else:
            default_base = self._default_output_base(in_key)
            path, conflict_mode = self._ask_output_path(out_key, default_base)
            if path is None:
                return
            sheet = safe_sheet_title(self.out_sheet_edit.text().strip() or default_base)
            key_name = self.out_sheet_edit.text().strip() or default_base
            delim = DELIM_LABELS[self.out_delim_combo.currentText()]
            enc = ENC_LABELS[self.out_enc_combo.currentText()]
            output = {"path": path, "mode": conflict_mode, "sheet": sheet,
                      "key": key_name, "delimiter": delim, "encoding": enc}

        # ---------- 后台执行管道 ----------
        def work():
            conn = self.ensure_conn() if (in_key == "SQL" or out_key == "SQL") else None
            cur = None
            header = rows = None
            try:
                if excel_to_sql:
                    self._run_excel_to_sql(conn, adapter, self.in_file_edit.text(),
                                           self.excel_selection,
                                           output["decisions"], output["batch"])
                    return
                if multi_sql or dir_csv:
                    self._run_multi_sql(conn, adapter, sources, out_key, output)
                    return

                # 1) 载入单输入源
                if in_key == "SQL":
                    self.log(f"读取数据表: {sources[0]['table']}"
                             f"(字段: {', '.join(sources[0]['columns'] or ['*'])})", "head")
                    header, rows, cur = load_sql_stream(
                        conn, adapter, sources[0]["table"], sources[0]["columns"])
                elif in_key == "Excel":
                    self.log(f"读取 Excel: {self.in_file_edit.text()} "
                             f"→ [{self.in_sheet_combo.currentText()}]", "head")
                    header, rows = load_excel_stream(
                        self.in_file_edit.text(), self.in_sheet_combo.currentText())
                elif in_key == "JSON":
                    self.log(f"读取 JSON: {self.in_file_edit.text()} "
                             f"→ 键[{self.in_key_combo.currentText()}]", "head")
                    header, rows = load_json_stream(
                        self.in_file_edit.text(), self.in_key_combo.currentText())
                elif in_key == "CSV":
                    delim = DELIM_LABELS[self.in_delim_combo.currentText()]
                    enc = ENC_LABELS[self.in_enc_combo.currentText()]
                    section = self.in_section_combo.currentText().strip()
                    if section and not section.startswith("("):
                        self.log(f"读取分节 CSV: {self.in_file_edit.text()} "
                                 f"→ [Sheet:{section}]", "head")
                        header, rows = load_csv_section(
                            self.in_file_edit.text(), section, delim, enc,
                            parse=(out_key == "SQL"))
                    else:
                        self.log(f"读取 CSV: {self.in_file_edit.text()}"
                                 f"(分隔符 {delim!r},编码 {enc})", "head")
                        header = read_csv_header(self.in_file_edit.text(), delim, enc)
                        rows = iter_csv_rows(self.in_file_edit.text(), delim, enc,
                                             len(header), parse=(out_key == "SQL"))
                self.log(f"共 {len(header)} 列: {header}", "info")

                # 2) 写出输出目标
                if out_key == "SQL":
                    if in_key == "SQL":
                        sample, buffered = [], []
                        for row in rows:
                            buffered.append(list(row))
                            if len(sample) < SAMPLE_SIZE:
                                sample.append(list(row))
                            if len(buffered) >= SAMPLE_SIZE:
                                break

                        def chained():
                            for row in buffered:
                                yield row
                            for row in rows:
                                yield list(row)

                        conn_out = adapter.connect(self.get_cfg())
                        try:
                            total = write_sql_table(
                                conn_out, adapter, output["table"], header,
                                chained(), output["mode"], output["batch"],
                                self.log, sample=sample)
                        finally:
                            conn_out.close()
                    else:
                        total = write_sql_table(
                            conn, adapter, output["table"], header, rows,
                            output["mode"], output["batch"], self.log)
                    self.log(f"✅ 写入数据库完成: 表【{output['table']}】共 {total:,} 行", "ok")
                    self.bridge.info_s.emit(
                        f"写入完成,共 {total:,} 行 → 表【{output['table']}】")
                elif out_key == "Excel":
                    total = write_excel_file(
                        output["path"], output["sheet"], header, rows,
                        mode=output["mode"], log=self.log)
                    self.log(f"✅ Excel 导出完成: {output['path']}"
                             f"(Sheet [{output['sheet']}],共 {total:,} 行)", "ok")
                    self.bridge.info_s.emit(
                        f"导出完成,共 {total:,} 行:\n{output['path']}")
                elif out_key == "JSON":
                    row_list = list(rows)
                    write_json_file(output["path"], output["key"], header, row_list,
                                    mode=output["mode"], log=self.log)
                    self.log(f"✅ JSON 导出完成: {output['path']}"
                             f"(键 [{output['key']}],共 {len(row_list):,} 行)", "ok")
                    self.bridge.info_s.emit(
                        f"导出完成,共 {len(row_list):,} 行:\n{output['path']}")
                elif out_key == "CSV":
                    total = write_csv_file(
                        output["path"], header, rows, delimiter=output["delimiter"],
                        encoding=output["encoding"], mode=output["mode"], log=self.log)
                    self.log(f"✅ CSV 导出完成: {output['path']}(共 {total:,} 行)", "ok")
                    self.bridge.info_s.emit(
                        f"导出完成,共 {total:,} 行:\n{output['path']}")
            finally:
                if cur is not None:
                    try:
                        cur.close()
                    except Exception:
                        pass
                    try:
                        conn.rollback()
                    except Exception:
                        pass

        self.run_task(work)

    # ------------------------------------------------------------- 退出
    def closeEvent(self, event):
        self._save_config()
        self._disconnect_conn()
        event.accept()


def run_self_test():
    """--selftest: 无界面自检(打包后验证驱动/引擎是否齐全),结果写入当前目录 selftest.log。"""
    log_path = Path.cwd() / "selftest.log"
    lines = []

    for module_name in ("pymysql", "psycopg2", "openpyxl"):
        try:
            __import__(module_name)
            lines.append(f"[OK]   import {module_name}")
        except Exception as exc:
            lines.append(f"[FAIL] import {module_name}: {exc}")

    # Excel 写读往返(覆盖 openpyxl + et_xmlfile + jsonb 复合值)
    try:
        tmp = log_path.parent / "_selftest.xlsx"
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title="自检")
        write_rows_to_worksheet(
            ws, ["id", "name", "device"],
            iter([[1, "张三", {"ip": "192.168.1.59", "os": "XiaomiNote8"}],
                  [2, "李四", [1, 2, 3]]]))
        wb.save(tmp)
        wb.close()
        wb2 = load_workbook(tmp, read_only=True, data_only=True)
        rows = list(wb2[wb2.sheetnames[0]].iter_rows(values_only=True))
        wb2.close()
        tmp.unlink(missing_ok=True)
        ok = (len(rows) == 3 and rows[1][0] == 1
              and isinstance(rows[1][2], str) and "192.168.1.59" in rows[1][2])
        lines.append(f"[{'OK' if ok else 'FAIL'}]   openpyxl 写读往返(含jsonb): {len(rows)} 行")
    except Exception as exc:
        lines.append(f"[FAIL] openpyxl 写读往返: {exc}")

    # JSON 写读往返(覆盖列数组结构 + 同名键合并逻辑)
    try:
        tmp = log_path.parent / "_selftest.json"
        rows = [[1, "张三"], [2, "李四"]]
        write_json_file(str(tmp), "info", ["Id", "Name"], rows)
        with open(tmp, "r", encoding="utf-8") as fp:
            doc = json.load(fp)
        ok = doc["info"]["Id"] == [1, 2] and doc["info"]["Name"][-1] == "李四"
        write_json_file(str(tmp), "info", ["Id", "Name"], [[3, "王五"]], mode="merge")
        write_json_file(str(tmp), "other", ["X"], [[9]], mode="merge")
        with open(tmp, "r", encoding="utf-8") as fp:
            doc = json.load(fp)
        ok = ok and doc["info"]["Id"] == [3] and "other" in doc
        tmp.unlink(missing_ok=True)
        lines.append(f"[{'OK' if ok else 'FAIL'}]   JSON 写读/合并往返")
    except Exception as exc:
        lines.append(f"[FAIL] JSON 写读/合并往返: {exc}")

    # CSV 写读往返(覆盖 csv 引擎 + 类型解析 + 复合值)
    try:
        tmp = log_path.parent / "_selftest.csv"
        write_csv_file(str(tmp), ["Id", "Name", "Device"],
                       iter([[1, "张三", {"ip": "192.168.1.59"}], [2, "李四", [4, 5]]]))
        header = read_csv_header(str(tmp), ",", "utf-8-sig")
        rows = list(iter_csv_rows(str(tmp), ",", "utf-8-sig", len(header), parse=True))
        ok = (header == ["Id", "Name", "Device"] and rows[0] == [1, "张三", '{"ip": "192.168.1.59"}']
              and rows[1][0] == 2)
        tmp.unlink(missing_ok=True)
        lines.append(f"[{'OK' if ok else 'FAIL'}]   CSV 写读往返: {len(rows)} 行")
    except Exception as exc:
        lines.append(f"[FAIL] CSV 写读往返: {exc}")

    # 目录模式(多表 CSV)写读往返:目录名 = 数据库名,每个表一个 {表名}.csv
    try:
        tmp_dir = log_path.parent / "_selftest_dir"
        tables = [
            ("NC7Middleschool", ["id", "name"], iter([[1, "张三"], [2, "李四"]])),
            ("departments", ["did", "dname"], iter([[10, "研发部"]])),
        ]
        write_csv_directory(str(tmp_dir), tables)
        p1 = tmp_dir / "NC7Middleschool.csv"
        p2 = tmp_dir / "departments.csv"
        ok = p1.exists() and p2.exists()
        header = read_csv_header(str(p1), ",", "utf-8-sig")
        rows = list(iter_csv_rows(str(p1), ",", "utf-8-sig", len(header), parse=True))
        ok = ok and header == ["id", "name"] and rows[0] == [1, "张三"]
        for p in (p1, p2):
            p.unlink(missing_ok=True)
        tmp_dir.rmdir()
        lines.append(f"[{'OK' if ok else 'FAIL'}]   目录CSV(每表一个文件)写读往返")
    except Exception as exc:
        lines.append(f"[FAIL] 目录CSV 写读往返: {exc}")

    # 旧版分节 CSV 读取(向后兼容:仍可读取 [Sheet:表名] 分节文件)
    try:
        tmp = log_path.parent / "_selftest_sec.csv"
        with open(tmp, "w", newline="", encoding="utf-8-sig") as fp:
            w = csv.writer(fp)
            w.writerow(["[Sheet:info]"])
            w.writerow(["Id", "Name"])
            w.writerow(["1", "张三"])
        sections = read_csv_sections(str(tmp), ",", "utf-8-sig")
        header, gen = load_csv_section(str(tmp), "info", ",", "utf-8-sig")
        rows = list(gen)
        ok = (isinstance(sections, dict) and "info" in sections
              and header == ["Id", "Name"] and rows[0] == ["1", "张三"])
        tmp.unlink(missing_ok=True)
        lines.append(f"[{'OK' if ok else 'FAIL'}]   分节CSV读取(兼容旧文件)")
    except Exception as exc:
        lines.append(f"[FAIL] 分节CSV读取: {exc}")

    # 若提供数据库连接环境变量,则做真实连接自检
    for env_prefix, adapter in (("SELFTEST_MYSQL", ADAPTERS["MySQL"]),
                                ("SELFTEST_PG", ADAPTERS["PostgreSQL"])):
        env = {
            "host": os.environ.get(f"{env_prefix}_HOST", "127.0.0.1"),
            "port": os.environ.get(f"{env_prefix}_PORT") or str(adapter.DEFAULT_PORT),
            "user": os.environ.get(f"{env_prefix}_USER", ""),
            "password": os.environ.get(f"{env_prefix}_PASSWORD", ""),
            "database": os.environ.get(f"{env_prefix}_DATABASE", ""),
            "charset": "utf8mb4",
            "timeout": "10",
        }
        if not env["user"] or not env["database"]:
            lines.append(f"[SKIP] {adapter.NAME} 连接自检(未设置 {env_prefix}_* 环境变量)")
            continue
        try:
            conn = adapter.connect(env)
            with conn.cursor() as cur:
                cur.execute("SELECT 1")
                cur.fetchone()
            conn.close()
            lines.append(f"[OK]   {adapter.NAME} 连接 SELECT 1")
        except Exception as exc:
            lines.append(f"[FAIL] {adapter.NAME} 连接: {exc}")

    log_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return all(not line.startswith("[FAIL]") for line in lines)


def main():
    if "--selftest" in sys.argv:
        sys.exit(0 if run_self_test() else 1)
    app = QApplication(sys.argv)
    app.setStyleSheet(QSS)
    window = App()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
