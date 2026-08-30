# -*- coding: utf-8 -*-
"""
SQL / Excel / JSON / CSV 数据互转工具 (Tkinter GUI)
====================================================
功能(输入 × 输出全组合):
    输入:  SQL 数据库(自由勾选 TableName + ColumnName)
           Excel (.xlsx) / JSON (.json) / CSV (.csv)
    输出:  Excel (.xlsx) / JSON (.json) / CSV (.csv) / SQL 数据库

命名约定(库-表-字段三层一一对应):
    数据库名  == 输出文件名(Excel/JSON/CSV)
    表名      == Excel Sheet 名 == JSON 一级参数
    字段名    == Excel 首行单元格 == JSON 二级参数 == CSV 表头

设计说明:
    * 数据库适配:MySQL 使用 PyMySQL,PostgreSQL 使用 psycopg2,统一 executemany 批量写入;
    * Excel 读写使用 openpyxl 的 read_only / write_only 流式模式,大批量数据内存占用低;
    * JSON 采用“列数组”结构(一级键=表名,二级键=字段名,值为数组);
    * CSV 使用标准库 csv,支持分隔符/编码选择,写文件前检测同名文件并支持覆盖/合并;
    * UI 采用 clam 主题 + 彩色样式,输入/输出分区动态展开,数据库连接面板按需显示。
"""

import csv
import json
import os
import queue
import re
import sys
import threading
import traceback
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path

import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

try:
    import pymysql
except Exception:  # pragma: no cover
    pymysql = None

try:
    import psycopg2
except Exception:  # pragma: no cover
    psycopg2 = None

# ---------------------------------------------------------------------------
# 常量 / 配色
# ---------------------------------------------------------------------------
APP_TITLE = "SQL ↔ Excel / JSON / CSV 数据转换工具"
SAMPLE_SIZE = 200            # 建表时用于推断字段类型的采样行数
DEFAULT_BATCH = 1000         # 默认每批读写行数
FONT = ("Microsoft YaHei", 10)
FONT_BOLD = ("Microsoft YaHei", 10, "bold")

# UI 配色(clam 主题)
BG = "#eef2f8"               # 窗体背景
CARD = "#ffffff"             # 卡片背景
NAVY = "#16294d"             # 顶部横幅
TEXT = "#1f2733"             # 主文字
MUTED = "#5b6b7f"            # 次要文字
BORDER = "#c9d4e4"           # 卡片边框
PRIMARY = "#1f6feb"          # 主蓝色(SQL)
GREEN = "#1a9e5c"            # Excel
ORANGE = "#e8833a"           # JSON
PURPLE = "#8e5bd8"           # CSV
RED = "#d64545"              # 错误
STATUS_OK = "#1a9e5c"
STATUS_OFF = "#9aa7b8"

MODE_COLORS = {"SQL": PRIMARY, "Excel": GREEN, "JSON": ORANGE, "CSV": PURPLE}

INPUT_LABELS = {
    "SQL": "SQL 数据库",
    "Excel": "Excel 文件 (.xlsx)",
    "JSON": "JSON 文件 (.json)",
    "CSV": "CSV 文件 (.csv)",
}
OUTPUT_LABELS = {
    "Excel": "Excel 文件 (.xlsx)",
    "JSON": "JSON 文件 (.json)",
    "CSV": "CSV 文件 (.csv)",
    "SQL": "SQL 数据库",
}

MODE_APPEND = "追加到已有表 (append)"
MODE_CREATE = "不存在则创建,存在则追加 (create if missing)"
MODE_REPLACE = "删除重建 (replace)"
MODE_LABELS = [MODE_APPEND, MODE_CREATE, MODE_REPLACE]

DELIM_LABELS = {"逗号 (,)": ",", "分号 (;)": ";", "制表符 (TAB)": "\t", "竖线 (|)": "|"}
ENC_LABELS = {
    "UTF-8 with BOM (Excel 友好)": "utf-8-sig",
    "UTF-8": "utf-8",
    "GBK (中文 Excel)": "gbk",
}

IDENT_RE = re.compile(r"^[^\W\d]\w*$")  # 标识符:字母/下划线开头,允许中文、数字、下划线


# ---------------------------------------------------------------------------
# 数据库适配层:MySQL / PostgreSQL
# ---------------------------------------------------------------------------
class DBAdapter:
    """数据库适配器基类,MySQL 与 PostgreSQL 各自实现差异点。"""

    NAME = ""
    DEFAULT_PORT = 3306
    QUOTE = '"'

    def connect(self, cfg):
        raise NotImplementedError

    def open_query_cursor(self, conn):
        """打开一个流式查询游标(避免一次性把全部结果缓冲到客户端内存)。"""
        raise NotImplementedError

    def list_tables(self, conn):
        raise NotImplementedError

    def list_columns(self, conn, table):
        raise NotImplementedError

    def is_missing_table_error(self, exc):
        raise NotImplementedError

    # -- 通用能力 ----------------------------------------------------------
    def quote_ident(self, name):
        """校验并引用标识符,支持 schema.table 形式,防止 SQL 注入。"""
        name = str(name).strip()
        if not name:
            raise ValueError("表名/字段名不能为空")
        parts = [p.strip() for p in name.split(".") if p.strip()]
        for part in parts:
            if not IDENT_RE.match(part):
                raise ValueError(f"非法标识符(仅允许字母/下划线/数字): {name!r}")
        q = self.QUOTE
        return ".".join(f"{q}{p}{q}" for p in parts)

    def table_exists(self, conn, quoted_table):
        """通过 SELECT 探测表是否存在(跨库、跨 schema 均可靠)。"""
        try:
            with conn.cursor() as cur:
                cur.execute(f"SELECT 1 FROM {quoted_table} WHERE 1=0")
            return True
        except Exception as exc:
            # PostgreSQL 中任何出错语句都会使当前事务进入 aborted 状态,
            # 必须先 rollback 才能继续后续 SQL。
            try:
                conn.rollback()
            except Exception:
                pass
            if self.is_missing_table_error(exc):
                return False
            raise

    def server_version(self, conn):
        with conn.cursor() as cur:
            cur.execute("SELECT VERSION()")
            return str(cur.fetchone()[0])


class MySQLAdapter(DBAdapter):
    NAME = "MySQL"
    DEFAULT_PORT = 3306
    QUOTE = "`"

    def connect(self, cfg):
        if pymysql is None:
            raise RuntimeError("未安装 PyMySQL,请先执行: pip install pymysql")
        return pymysql.connect(
            host=cfg["host"],
            port=int(cfg["port"]),
            user=cfg["user"],
            password=cfg["password"],
            database=cfg["database"],
            charset=cfg.get("charset") or "utf8mb4",
            connect_timeout=int(cfg.get("timeout", 10)),
        )

    def open_query_cursor(self, conn):
        # SSCursor = 服务端无缓冲游标,大数据量导出不占客户端内存
        return conn.cursor(pymysql.cursors.SSCursor)

    def list_tables(self, conn):
        with conn.cursor() as cur:
            cur.execute("SHOW TABLES")
            return [str(row[0]) for row in cur.fetchall()]

    def list_columns(self, conn, table):
        with conn.cursor() as cur:
            cur.execute(f"SHOW COLUMNS FROM {self.quote_ident(table)}")
            return [str(row[0]) for row in cur.fetchall()]

    def is_missing_table_error(self, exc):
        errno = exc.args[0] if exc.args else None
        return isinstance(exc, pymysql.err.ProgrammingError) and errno == 1146


class PostgresAdapter(DBAdapter):
    NAME = "PostgreSQL"
    DEFAULT_PORT = 5432
    QUOTE = '"'

    def connect(self, cfg):
        if psycopg2 is None:
            raise RuntimeError("未安装 psycopg2,请先执行: pip install psycopg2-binary")
        return psycopg2.connect(
            host=cfg["host"],
            port=int(cfg["port"]),
            user=cfg["user"],
            password=cfg["password"],
            dbname=cfg["database"],
            connect_timeout=int(cfg.get("timeout", 10)),
        )

    def open_query_cursor(self, conn):
        # 服务端命名游标:结果按 itersize 分批从服务器拉取
        cur = conn.cursor(name=f"sql_excel_cur_{os.getpid()}_{id(conn)}")
        cur.itersize = 5000
        return cur

    def list_tables(self, conn):
        with conn.cursor() as cur:
            cur.execute(
                "SELECT table_schema, table_name FROM information_schema.tables "
                "WHERE table_schema NOT IN ('pg_catalog', 'information_schema') "
                "ORDER BY table_schema, table_name"
            )
            return [f"{s}.{t}" if s != "public" else str(t) for s, t in cur.fetchall()]

    def list_columns(self, conn, table):
        parts = table.split(".")
        schema, tname = (parts[0], parts[-1]) if len(parts) == 2 else ("public", parts[-1])
        with conn.cursor() as cur:
            cur.execute(
                "SELECT column_name FROM information_schema.columns "
                "WHERE table_schema = %s AND table_name = %s "
                "ORDER BY ordinal_position",
                (schema, tname),
            )
            return [str(row[0]) for row in cur.fetchall()]

    def is_missing_table_error(self, exc):
        return getattr(exc, "pgcode", None) == "42P01"


ADAPTERS = {
    "MySQL": MySQLAdapter(),
    "PostgreSQL": PostgresAdapter(),
}


# ---------------------------------------------------------------------------
# Excel 工具函数(openpyxl 流式模式)
# ---------------------------------------------------------------------------
def safe_sheet_title(name):
    """工作表名:去掉非法字符,长度 <= 31。"""
    name = re.sub(r'[\\/*?:\[\]]', "_", str(name)).strip()
    return (name[:31] or "Sheet1")


def display_width(text):
    """按显示宽度估算列宽(中文等全角字符算 2)。"""
    return sum(2 if ord(ch) > 0x7F else 1 for ch in str(text))


def to_excel_value(value):
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, bytes):
        try:
            return value.decode("utf-8")
        except UnicodeDecodeError:
            return value.decode("utf-8", "replace")
    return value  # int / float / Decimal / str / bool 等


def write_rows_to_worksheet(ws, header, row_iterator, log=None):
    """把数据行逐行写入工作表,顺便自动计算列宽。"""
    ws.append([to_excel_value(h) for h in header])
    widths = {i + 1: min(display_width(h) + 2, 60) for i, h in enumerate(header)}
    total = 0
    for row in row_iterator:
        values = [to_excel_value(v) for v in row]
        ws.append(values)
        for idx, value in enumerate(values, start=1):
            if value is None:
                continue
            width = min(display_width(value) + 2, 60)
            if width > widths.get(idx, 0):
                widths[idx] = width
        total += 1
        if log and total % 20000 == 0:
            log(f"  已写出 {total:,} 行 ...")
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width
    return total


def is_blank_row(row):
    return all(
        v is None or (isinstance(v, str) and not v.strip())
        for v in row
    )


def normalize_row(row, ncols):
    """把行对齐到表头列数:短了补 None,长了截断。"""
    row = list(row)
    if len(row) > ncols:
        return row[:ncols]
    return row + [None] * (ncols - len(row))


def to_db_value(value):
    if value is None:
        return None
    if isinstance(value, str):
        s = value.strip()
        return s or None
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    return value  # datetime / date / Decimal / bool / int / float 直接交给驱动


def parse_header(raw_row):
    """解析表头:去掉尾部空列,空列自动命名,重名自动加后缀。"""
    cells = list(raw_row)
    last = len(cells)
    while last > 0 and (
        cells[last - 1] is None
        or (isinstance(cells[last - 1], str) and not cells[last - 1].strip())
    ):
        last -= 1
    if last == 0:
        raise ValueError("数据表头为空,无法转换")

    header, counts = [], {}
    for i in range(last):
        name = str(cells[i]).strip() if cells[i] is not None else ""
        if not name:
            name = f"column_{i + 1}"
        if name in counts:
            counts[name] += 1
            name = f"{name}_{counts[name]}"
        else:
            counts[name] = 1
        header.append(name)
    return header


def collect_stats(header, sample_rows):
    """统计采样行中各列出现的数据类型,用于 CREATE TABLE 的类型推断。"""
    stats = {
        i: {
            "str": False, "max_len": 0, "int": False, "big": False,
            "float": False, "bool": False, "datetime": False, "date": False,
        }
        for i in range(len(header))
    }
    for row in sample_rows:
        for i, value in enumerate(row):
            st = stats[i]
            if value is None:
                continue
            if isinstance(value, bool):
                st["bool"] = True
            elif isinstance(value, int):
                st["int"] = True
                if value > 2 ** 31 - 1 or value < -(2 ** 31):
                    st["big"] = True
            elif isinstance(value, (float, Decimal)):
                st["float"] = True
            elif isinstance(value, datetime):
                st["datetime"] = True
            elif isinstance(value, date):
                st["date"] = True
            else:
                st["str"] = True
                st["max_len"] = max(st["max_len"], len(str(value)))
    return stats


def decide_type(adapter, st):
    is_mysql = adapter.NAME == "MySQL"
    if st.get("str"):                   # 含文本(混合类型也按文本兜底)
        max_len = st.get("max_len", 0) or 1
        return "TEXT" if max_len > 4000 else f"VARCHAR({max_len})"
    if st.get("datetime"):
        return "TIMESTAMP"
    if st.get("date"):
        return "DATE"
    if st.get("float"):
        return "DOUBLE" if is_mysql else "DOUBLE PRECISION"
    if st.get("bool") and not st.get("int"):
        return "TINYINT(1)" if is_mysql else "BOOLEAN"
    if st.get("int"):
        return "BIGINT" if st.get("big") else "INT"
    return "VARCHAR(255)"


def build_create_table(adapter, table, header, stats):
    columns = [
        f"{adapter.quote_ident(col)} {decide_type(adapter, stats[idx])}"
        for idx, col in enumerate(header)
    ]
    return f"CREATE TABLE {adapter.quote_ident(table)} ({', '.join(columns)})"


# ---------------------------------------------------------------------------
# 文件写出工具(含同名文件冲突处理)
# ---------------------------------------------------------------------------
def safe_filename(name):
    """去掉文件名中 Windows 不允许的字符。"""
    name = re.sub(r'[\\/:*?"<>|\r\n]+', "_", str(name)).strip()
    return name or "export"


def to_json_value(value):
    """把数据库返回值转换成 JSON 兼容值(时间统一为字符串,与示例文件一致)。"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, bytes):
        try:
            return value.decode("utf-8")
        except UnicodeDecodeError:
            return value.decode("utf-8", "replace")
    return value  # int / float / str / bool


def to_csv_value(value):
    """把数据值转换成 CSV 单元格文本(时间与 JSON 规则一致)。"""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, bytes):
        try:
            return value.decode("utf-8")
        except UnicodeDecodeError:
            return value.decode("utf-8", "replace")
    return value


def build_json_document(top_key, columns, rows):
    """构造 {表名: {字段名: [值...]}} 的列数组式 JSON 结构。"""
    doc = {top_key: {col: [] for col in columns}}
    for row in rows:
        norm = normalize_row(row, len(columns))
        for col, value in zip(columns, norm):
            doc[top_key][col].append(to_json_value(value))
    return doc


def write_json_file(path, top_key, columns, rows, mode="replace", log=None):
    """写出 JSON 文件。mode=merge 时保留原文件其他一级键,仅覆盖同名键。"""
    doc = build_json_document(top_key, columns, rows)
    if mode == "merge" and os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as fp:
                old = json.load(fp)
        except Exception:
            old = None
        if isinstance(old, dict):
            old[top_key] = doc[top_key]
            doc = old
    with open(path, "w", encoding="utf-8") as fp:
        json.dump(doc, fp, ensure_ascii=False, indent=4)
    return len(rows)


def write_excel_file(path, sheet, columns, row_iterator, mode="replace", log=None):
    """写出 Excel 文件。

    mode=replace : 用流式 write_only 模式重建整个工作簿(大数据量首选);
    mode=merge   : 保留原工作簿其他 Sheet,覆盖同名 Sheet(无同名则新增)。
    """
    if mode == "replace":
        wb = Workbook(write_only=True)
        try:
            ws = wb.create_sheet(title=sheet)
            total = write_rows_to_worksheet(ws, columns, row_iterator, log)
            wb.save(path)
        finally:
            try:
                wb.close()
            except Exception:
                pass
        return total

    # merge:打开已有工作簿(常规模式),只动同名 Sheet
    wb = load_workbook(path)
    try:
        if sheet in wb.sheetnames:
            del wb[sheet]
        ws = wb.create_sheet(title=sheet)
        total = write_rows_to_worksheet(ws, columns, row_iterator, log)
        wb.save(path)
    finally:
        wb.close()
    return total


def write_csv_file(path, columns, row_iterator, delimiter=",", encoding="utf-8-sig",
                   mode="replace", log=None):
    """写出 CSV 文件。

    mode=replace : 重建文件(含表头);
    mode=merge   : 向已有文件追加数据行(不重复写表头)。
    """
    append = mode == "merge" and os.path.exists(path)
    fp = open(path, "a" if append else "w", newline="", encoding=encoding)
    writer = csv.writer(fp, delimiter=delimiter, lineterminator="\n")
    total = 0
    try:
        if not append:
            writer.writerow([to_csv_value(h) for h in columns])
        for row in row_iterator:
            writer.writerow(
                [to_csv_value(v) for v in normalize_row(row, len(columns))])
            total += 1
            if log and total % 50000 == 0:
                log(f"  已写出 {total:,} 行 ...")
    finally:
        fp.close()
    return total


# ---------------------------------------------------------------------------
# 数据读取工具(统一为: 表头 + 行生成器)
# ---------------------------------------------------------------------------
def load_excel_stream(path, sheet):
    """读取 xlsx 工作表,返回 (表头, 行生成器)。"""
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        raise ValueError(f"工作簿中不存在工作表 {sheet!r},可用: {', '.join(wb.sheetnames)}")
    ws = wb[sheet]
    it = ws.iter_rows(values_only=True)
    try:
        raw_header = next(it)
    except StopIteration:
        wb.close()
        raise ValueError("工作表为空,无法读取")
    header = parse_header(raw_header)

    def gen():
        try:
            for row in it:
                if is_blank_row(row):
                    continue
                yield normalize_row(row, len(header))
        finally:
            wb.close()

    return header, gen()


def read_csv_header(path, delimiter, encoding):
    with open(path, "r", newline="", encoding=encoding) as fp:
        reader = csv.reader(fp, delimiter=delimiter)
        try:
            raw = next(reader)
        except StopIteration:
            raise ValueError("CSV 文件为空,无法读取")
    return parse_header(raw)


def parse_csv_cell(value):
    """CSV 单元格文本 → 自动推断 int / float / date / datetime,失败保持字符串。"""
    if not isinstance(value, str):
        return value
    s = value.strip()
    if s == "":
        return None
    if re.fullmatch(r"-?\d{1,9}", s):          # 短整数(身份证/长号码保留字符串)
        return int(s)
    if re.fullmatch(r"-?\d+\.\d+", s):
        return float(s)
    try:
        return datetime.strptime(s, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        pass
    try:
        return date.fromisoformat(s)
    except ValueError:
        pass
    return s


def iter_csv_rows(path, delimiter, encoding, header_len, parse=False):
    """逐行生成 CSV 数据(跳过表头与空行),parse=True 时按类型解析单元格。"""
    with open(path, "r", newline="", encoding=encoding) as fp:
        reader = csv.reader(fp, delimiter=delimiter)
        next(reader, None)  # 跳过表头
        for row in reader:
            if is_blank_row(row):
                continue
            norm = normalize_row(row, header_len)
            if parse:
                yield [parse_csv_cell(c) for c in norm]
            else:
                yield norm


def load_json_stream(path, top_key):
    """读取 JSON 列数组结构,返回 (表头, 行生成器)。"""
    with open(path, "r", encoding="utf-8") as fp:
        doc = json.load(fp)
    if not isinstance(doc, dict):
        raise ValueError("JSON 顶层必须是对象")
    if top_key not in doc:
        raise ValueError(f"JSON 中不存在一级键 {top_key!r},可用: {list(doc.keys())}")
    table = doc[top_key]
    if not isinstance(table, dict):
        raise ValueError(f"一级键 {top_key!r} 的值必须是 {{字段: [值...]}} 对象")
    columns = list(table.keys())
    arrays = []
    for col in columns:
        arr = table.get(col)
        if not isinstance(arr, list):
            raise ValueError(f"字段 {col!r} 的值必须是数组")
        arrays.append(arr)
    nrows = max((len(a) for a in arrays), default=0)

    def gen():
        for i in range(nrows):
            yield [arr[i] if i < len(arr) else None for arr in arrays]

    return columns, gen()


def load_sql_stream(conn, adapter, table, columns=None):
    """读取数据库表,返回 (表头, 行生成器, 游标)。调用方负责关闭游标。"""
    col_sql = ", ".join(adapter.quote_ident(c) for c in columns) if columns else "*"
    cur = adapter.open_query_cursor(conn)
    cur.execute(f"SELECT {col_sql} FROM {adapter.quote_ident(table)}")
    # PostgreSQL 命名游标的 description 要等首次 fetch 后才就绪
    first = cur.fetchmany(DEFAULT_BATCH)
    header = [str(desc[0]) for desc in cur.description]

    def gen():
        for row in first:
            yield list(row)
        while True:
            rows = cur.fetchmany(DEFAULT_BATCH)
            if not rows:
                break
            for row in rows:
                yield list(row)

    return header, gen(), cur


def write_sql_table(conn, adapter, table, header, rows_iter, mode, batch, log,
                    sample=None):
    """把行流写入数据库表。

    sample=None 时单遍采样前 SAMPLE_SIZE 行推断类型(适合文件输入);
    调用方也可自行提供 sample 并传入未消费的行流(适合 SQL 输入,避免游标与 DDL 冲突)。
    """
    if sample is None:
        # 单遍采样:先缓存前 SAMPLE_SIZE 行用于类型推断,随后接续剩余行
        sample, buffered = [], []
        for row in rows_iter:
            norm = normalize_row(row, len(header))
            buffered.append(norm)
            if len(sample) < SAMPLE_SIZE:
                sample.append(norm)
            if len(buffered) >= SAMPLE_SIZE:
                break

        source_iter = rows_iter  # 固定引用,避免闭包晚绑定

        def chained():
            for row in buffered:
                yield row
            for row in source_iter:
                yield normalize_row(row, len(header))

        rows_iter = chained()

    quoted_table = adapter.quote_ident(table)
    exists = adapter.table_exists(conn, quoted_table)
    need_create = (mode == "replace") or (mode == "create_if_missing" and not exists)
    if mode == "append" and not exists:
        raise ValueError(
            f"目标表【{table}】不存在。请改用“{MODE_CREATE}”或“{MODE_REPLACE}”模式")
    if need_create:
        stats = collect_stats(header, sample)
        if mode == "replace" and exists:
            with conn.cursor() as cur:
                cur.execute(f"DROP TABLE {quoted_table}")
            log(f"已删除旧表: {table}")
        ddl = build_create_table(adapter, table, header, stats)
        with conn.cursor() as cur:
            cur.execute(ddl)
        log(f"已创建表: {table}")

    insert_sql = (
        f"INSERT INTO {quoted_table} "
        f"({', '.join(adapter.quote_ident(c) for c in header)}) "
        f"VALUES ({', '.join(['%s'] * len(header))})"
    )
    total = 0
    chunk = []
    try:
        with conn.cursor() as cur:
            for row in rows_iter:
                chunk.append(tuple(to_db_value(v) for v in row))
                if len(chunk) >= batch:
                    cur.executemany(insert_sql, chunk)
                    total += len(chunk)
                    chunk = []
                    log(f"  已写入 {total:,} 行 ...")
            if chunk:
                cur.executemany(insert_sql, chunk)
                total += len(chunk)
        conn.commit()
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        raise
    return total


# ---------------------------------------------------------------------------
# 下拉勾选控件:按钮 + 弹出式勾选列表
# ---------------------------------------------------------------------------
class CheckboxDropdown(ttk.Frame):
    """可勾选的下拉列表(multi=False 时为单选,用于选表;multi=True 用于选字段)。"""

    def __init__(self, master, text="请选择", multi=False, width=36, style=None):
        super().__init__(master, style="Card.TFrame")
        self.multi = multi
        self.title_text = text
        self._items = []
        self._checked = set()
        self.command = None
        self._popup = None
        self.button = ttk.Button(
            self, text=f"{text}: (未加载)", width=width,
            command=self._open, style=style or "TButton")
        self.button.pack(fill="x")

    # -- 对外接口 -----------------------------------------------------------
    def set_items(self, items, checked=None):
        self._items = [str(i) for i in items]
        if checked is None:
            checked = self._items[:] if self.multi else []
        checked = [c for c in checked if c in self._items]
        if not self.multi and not checked and self._items:
            checked = [self._items[0]]
        self._checked = set(checked)
        self._refresh_text()

    def get_selected(self):
        return [i for i in self._items if i in self._checked]

    def get_single(self):
        selected = self.get_selected()
        return selected[0] if selected else None

    def set_command(self, fn):
        self.command = fn

    # -- 内部实现 -----------------------------------------------------------
    def _refresh_text(self):
        if not self._items:
            self.button.configure(text=f"{self.title_text}: (未加载)")
        elif self.multi:
            self.button.configure(
                text=f"{self.title_text}: 已选 {len(self._checked)}/{len(self._items)}")
        else:
            self.button.configure(text=f"{self.title_text}: {self.get_single()}")

    def _open(self):
        if not self._items:
            return
        top = tk.Toplevel(self.winfo_toplevel())
        self._popup = top
        top.title(self.title_text)
        top.transient(self.winfo_toplevel())
        top.configure(bg=CARD)
        try:
            x = self.button.winfo_rootx()
            y = self.button.winfo_rooty() + self.button.winfo_height()
            top.geometry(f"+{x}+{y}")
        except tk.TclError:
            pass

        frame = ttk.Frame(top, style="Card.TFrame", padding=6)
        frame.pack(fill="both", expand=True)
        lb = tk.Listbox(
            frame, height=min(12, max(4, len(self._items))), width=46,
            selectmode="single", activestyle="none", font=FONT,
            bg=CARD, fg=TEXT, selectbackground=PRIMARY, selectforeground="white",
            highlightthickness=0, bd=1, relief="solid")
        scroll = ttk.Scrollbar(frame, orient="vertical", command=lb.yview)
        lb.configure(yscrollcommand=scroll.set)
        lb.pack(side="left", fill="both", expand=True)
        scroll.pack(side="left", fill="y")

        def render():
            lb.delete(0, "end")
            for item in self._items:
                mark = "☑ " if item in self._checked else "☐ "
                lb.insert("end", mark + item)

        def on_click(event):
            idx = lb.nearest(event.y)
            if idx < 0 or idx >= len(self._items):
                return
            item = self._items[idx]
            if item in self._checked:
                self._checked.discard(item)
            else:
                if not self.multi:
                    self._checked.clear()
                self._checked.add(item)
            render()

        lb.bind("<Button-1>", on_click)

        btn_row = ttk.Frame(frame, style="Card.TFrame")
        btn_row.pack(fill="x", pady=(6, 0))
        if self.multi:
            ttk.Button(btn_row, text="全选", width=8,
                       command=lambda: (self._checked.update(self._items), render())
                       ).pack(side="left", padx=2)
            ttk.Button(btn_row, text="全不选", width=8,
                       command=lambda: (self._checked.clear(), render())
                       ).pack(side="left", padx=2)
        ttk.Button(btn_row, text="确定", width=8, style="Primary.TButton",
                   command=self._confirm).pack(side="right", padx=2)
        ttk.Label(btn_row, text="点击条目切换勾选",
                  style="Card.TLabel").pack(side="right", padx=8)

        render()
        top.grab_set()

    def _confirm(self):
        popup = getattr(self, "_popup", None)
        selected = self.get_selected()
        self._refresh_text()
        if popup is not None and popup.winfo_exists():
            popup.destroy()
        self._popup = None
        if self.command:
            self.command(selected)


# ---------------------------------------------------------------------------
# Tkinter 主界面
# ---------------------------------------------------------------------------
class App:
    def __init__(self, root):
        self.root = root
        root.title(APP_TITLE)
        root.geometry("1100x780")
        root.minsize(1020, 700)
        root.configure(bg=BG)

        self.conn = None
        self.adapter = ADAPTERS["MySQL"]
        self.msg_queue = queue.Queue()
        self._busy = False
        self.selected_table = None
        self._auto_sheet = ""
        self._auto_table = ""

        self._setup_styles()
        self._build_vars()
        self._build_ui()
        self.on_input_mode_changed()
        self.on_output_mode_changed()
        root.protocol("WM_DELETE_WINDOW", self.on_close)
        root.after(100, self._poll_queue)

    # ------------------------------------------------------------------ 样式
    def _setup_styles(self):
        style = ttk.Style(self.root)
        style.theme_use("clam")
        style.configure(".", font=FONT, background=BG, foreground=TEXT)
        style.configure("TFrame", background=BG)
        style.configure("Card.TFrame", background=CARD)
        style.configure("TLabel", background=BG, foreground=TEXT)
        style.configure("Card.TLabel", background=CARD, foreground=TEXT)
        style.configure("Muted.Card.TLabel", background=CARD, foreground=MUTED)
        style.configure("Section.TLabel", background=CARD, foreground=NAVY,
                        font=("Microsoft YaHei", 11, "bold"))
        style.configure("Header.TLabel", background=NAVY, foreground="white",
                        font=("Microsoft YaHei", 16, "bold"))
        style.configure("SubHeader.TLabel", background=NAVY, foreground="#bcd3f5",
                        font=("Microsoft YaHei", 9))
        style.configure("Step.TLabel", background=PRIMARY, foreground="white",
                        font=("Microsoft YaHei", 10, "bold"), padding=(12, 5))
        style.configure("TButton", background="#e7edf6", foreground=TEXT,
                        padding=(10, 5))
        style.map("TButton", background=[("active", "#d5e2f5")])
        style.configure("Blue.TButton", background=PRIMARY, foreground="white")
        style.map("Blue.TButton", background=[("active", "#3b82f6")])
        style.configure("Primary.TButton", background=GREEN, foreground="white",
                        font=("Microsoft YaHei", 11, "bold"), padding=(16, 8))
        style.map("Primary.TButton", background=[("active", "#21b36b")])
        style.configure("Warn.TButton", background=ORANGE, foreground="white")
        style.map("Warn.TButton", background=[("active", "#f09645")])
        style.configure("TEntry", fieldbackground="white", bordercolor=BORDER,
                        lightcolor=BORDER, darkcolor=BORDER)
        style.configure("TCombobox", fieldbackground="white", bordercolor=BORDER)
        style.configure("TScrollbar", background="#c9d4e4")
        style.configure("TCheckbutton", background=CARD)
        style.configure("Status.TLabel", background=STATUS_OFF, foreground="white",
                        padding=(10, 3))
        style.configure("Ok.Status.TLabel", background=STATUS_OK, foreground="white",
                        padding=(10, 3))
        style.configure("Warn.Status.TLabel", background=ORANGE, foreground="white",
                        padding=(10, 3))
        for key, color in MODE_COLORS.items():
            for prefix in ("CardIn", "CardOut"):
                style.configure(f"{prefix}{key}.TLabelframe", background=CARD,
                                bordercolor=color, relief="solid", borderwidth=2)
                style.configure(f"{prefix}{key}.TLabelframe.Label",
                                background=CARD, foreground=color, font=FONT_BOLD)
        style.configure("DB.TLabelframe", background=CARD, bordercolor=PRIMARY,
                        relief="solid", borderwidth=2)
        style.configure("DB.TLabelframe.Label", background=CARD,
                        foreground=PRIMARY, font=FONT_BOLD)
        style.configure("Log.TLabelframe", background=CARD, bordercolor=BORDER,
                        relief="solid", borderwidth=1)
        style.configure("Log.TLabelframe.Label", background=CARD,
                        foreground=MUTED, font=FONT_BOLD)

    # ------------------------------------------------------------------ 变量
    def _build_vars(self):
        self.input_mode_var = tk.StringVar(value=INPUT_LABELS["SQL"])
        self.output_mode_var = tk.StringVar(value=OUTPUT_LABELS["Excel"])

        self.db_type_var = tk.StringVar(value="MySQL")
        self.host_var = tk.StringVar(value="127.0.0.1")
        self.port_var = tk.StringVar(value="3306")
        self.user_var = tk.StringVar(value="root")
        self.pwd_var = tk.StringVar(value="")
        self.database_var = tk.StringVar(value="")
        self.timeout_var = tk.StringVar(value="10")

        self.in_file_var = tk.StringVar(value="")
        self.in_sheet_var = tk.StringVar(value="")
        self.in_key_var = tk.StringVar(value="")
        self.in_delim_var = tk.StringVar(value=list(DELIM_LABELS)[0])
        self.in_enc_var = tk.StringVar(value=list(ENC_LABELS)[0])

        self.out_table_var = tk.StringVar(value="")
        self.out_mode_var = tk.StringVar(value=MODE_APPEND)
        self.out_sheet_var = tk.StringVar(value="")
        self.out_delim_var = tk.StringVar(value=list(DELIM_LABELS)[0])
        self.out_enc_var = tk.StringVar(value=list(ENC_LABELS)[0])
        self.batch_var = tk.StringVar(value=str(DEFAULT_BATCH))

        self.status_var = tk.StringVar(value="未连接数据库 | 就绪")

    # ------------------------------------------------------------------ UI
    def _build_ui(self):
        pad = {"padx": 4, "pady": 3}

        # ---------- 顶部横幅 ----------
        header = tk.Frame(self.root, bg=NAVY)
        header.grid(row=0, column=0, sticky="ew")
        ttk.Label(header, text=APP_TITLE, style="Header.TLabel").pack(
            side="left", padx=16, pady=(10, 2))
        ttk.Label(header, text="SQL · Excel · JSON · CSV 任意互转,彩色主题界面",
                  style="SubHeader.TLabel").pack(side="right", padx=16)

        # ---------- 模式栏 ----------
        mode_bar = ttk.Frame(self.root, style="Card.TFrame")
        mode_bar.grid(row=1, column=0, sticky="ew", padx=10, pady=(10, 6))
        ttk.Label(mode_bar, text="① 输入格式", style="Step.TLabel").pack(
            side="left", padx=(10, 6), pady=8)
        self.input_combo = ttk.Combobox(
            mode_bar, textvariable=self.input_mode_var,
            values=list(INPUT_LABELS.values()), state="readonly", width=20)
        self.input_combo.pack(side="left", pady=8)
        self.input_combo.bind("<<ComboboxSelected>>", self.on_input_mode_changed)
        ttk.Label(mode_bar, text="▶ 转换", style="Muted.Card.TLabel",
                  font=FONT_BOLD).pack(side="left", padx=14)
        ttk.Label(mode_bar, text="② 输出格式", style="Step.TLabel").pack(
            side="left", padx=(6, 6))
        self.output_combo = ttk.Combobox(
            mode_bar, textvariable=self.output_mode_var,
            values=list(OUTPUT_LABELS.values()), state="readonly", width=20)
        self.output_combo.pack(side="left", pady=8)
        self.output_combo.bind("<<ComboboxSelected>>", self.on_output_mode_changed)
        ttk.Button(mode_bar, text="开始转换 ▶", style="Primary.TButton",
                   command=self.on_convert).pack(side="right", padx=14, pady=8)

        # ---------- 内容区 ----------
        content = ttk.Frame(self.root)
        content.grid(row=2, column=0, sticky="nsew", padx=10)
        content.columnconfigure(0, weight=1)

        self._build_input_card(content)
        self._build_db_card(content)
        self._build_output_card(content)

        # ---------- 日志 ----------
        frm_log = ttk.LabelFrame(self.root, text="④ 运行日志", style="Log.TLabelframe")
        frm_log.grid(row=3, column=0, sticky="nsew", padx=10, pady=(6, 2))
        frm_log.columnconfigure(0, weight=1)
        frm_log.rowconfigure(0, weight=1)
        self.log_text = scrolledtext.ScrolledText(
            frm_log, height=10, font=("Consolas", 9), state="disabled", wrap="word",
            bg="#fbfcfe", fg=TEXT, relief="flat")
        self.log_text.grid(row=0, column=0, sticky="nsew")
        for tag, color in (("info", "#3a4a5c"), ("ok", GREEN), ("warn", "#b8741a"),
                           ("err", RED), ("head", PRIMARY)):
            self.log_text.tag_configure(tag, foreground=color)

        # ---------- 状态栏 ----------
        status_frame = ttk.Frame(self.root)
        status_frame.grid(row=4, column=0, sticky="ew", padx=10, pady=(0, 8))
        self.status_chip = ttk.Label(status_frame, textvariable=self.status_var,
                                     style="Status.TLabel")
        self.status_chip.pack(side="left")

        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(2, weight=1)
        self.root.rowconfigure(3, weight=1)

    def _build_input_card(self, parent):
        self.input_card = ttk.LabelFrame(parent, text="① 输入设置", padding=8)
        self.input_card.grid(row=0, column=0, sticky="ew", pady=(0, 6))
        self.input_card.columnconfigure(0, weight=1)

        # ---- 输入 = SQL:表 / 字段勾选 ----
        self.src_sql_frame = ttk.Frame(self.input_card, style="Card.TFrame")
        self.src_sql_frame.columnconfigure(1, weight=1)
        ttk.Label(self.src_sql_frame, text="数据表:",
                  style="Card.TLabel").grid(row=0, column=0, sticky="e", padx=4, pady=4)
        self.table_dropdown = CheckboxDropdown(
            self.src_sql_frame, text="选择数据表(勾选)", multi=False, width=40)
        self.table_dropdown.grid(row=0, column=1, sticky="ew", padx=4, pady=4)
        self.table_dropdown.set_command(self.on_tables_confirmed)
        ttk.Button(self.src_sql_frame, text="刷新表", style="Blue.TButton",
                   command=self.on_refresh_tables).grid(row=0, column=2, padx=4, pady=4)

        ttk.Label(self.src_sql_frame, text="导出字段:",
                  style="Card.TLabel").grid(row=1, column=0, sticky="e", padx=4, pady=4)
        self.column_dropdown = CheckboxDropdown(
            self.src_sql_frame, text="选择字段(多选勾选)", multi=True, width=40)
        self.column_dropdown.grid(row=1, column=1, sticky="ew", padx=4, pady=4)
        ttk.Label(self.src_sql_frame,
                  text="先勾选数据表,再勾选要导出的字段(默认全选)",
                  style="Muted.Card.TLabel").grid(row=1, column=2, padx=4, pady=4)

        # ---- 输入 = 文件 ----
        self.src_file_frame = ttk.Frame(self.input_card, style="Card.TFrame")
        self.src_file_frame.columnconfigure(1, weight=1)
        ttk.Label(self.src_file_frame, text="输入文件:",
                  style="Card.TLabel").grid(row=0, column=0, sticky="e", padx=4, pady=4)
        ttk.Entry(self.src_file_frame, textvariable=self.in_file_var).grid(
            row=0, column=1, sticky="ew", padx=4, pady=4)
        ttk.Button(self.src_file_frame, text="选择文件...", style="Blue.TButton",
                   command=self.on_pick_input_file).grid(row=0, column=2, padx=4, pady=4)

        # 文件类型专属参数行(按输入格式切换显示)
        self.file_excel_opts = ttk.Frame(self.src_file_frame, style="Card.TFrame")
        ttk.Label(self.file_excel_opts, text="工作表:",
                  style="Card.TLabel").grid(row=0, column=0, sticky="e", padx=4, pady=2)
        self.in_sheet_combo = ttk.Combobox(
            self.file_excel_opts, textvariable=self.in_sheet_var,
            state="readonly", width=40)
        self.in_sheet_combo.grid(row=0, column=1, sticky="w", padx=4, pady=2)
        self.in_sheet_combo.bind("<<ComboboxSelected>>", self.on_in_sheet_selected)

        self.file_json_opts = ttk.Frame(self.src_file_frame, style="Card.TFrame")
        ttk.Label(self.file_json_opts, text="一级键(表名):",
                  style="Card.TLabel").grid(row=0, column=0, sticky="e", padx=4, pady=2)
        self.in_key_combo = ttk.Combobox(
            self.file_json_opts, textvariable=self.in_key_var,
            state="readonly", width=40)
        self.in_key_combo.grid(row=0, column=1, sticky="w", padx=4, pady=2)
        self.in_key_combo.bind("<<ComboboxSelected>>", self.on_in_key_selected)

        self.file_csv_opts = ttk.Frame(self.src_file_frame, style="Card.TFrame")
        ttk.Label(self.file_csv_opts, text="分隔符:",
                  style="Card.TLabel").grid(row=0, column=0, sticky="e", padx=4, pady=2)
        ttk.Combobox(self.file_csv_opts, textvariable=self.in_delim_var,
                     values=list(DELIM_LABELS), state="readonly",
                     width=18).grid(row=0, column=1, sticky="w", padx=4, pady=2)
        ttk.Label(self.file_csv_opts, text="编码:",
                  style="Card.TLabel").grid(row=0, column=2, sticky="e", padx=4, pady=2)
        ttk.Combobox(self.file_csv_opts, textvariable=self.in_enc_var,
                     values=list(ENC_LABELS), state="readonly",
                     width=24).grid(row=0, column=3, sticky="w", padx=4, pady=2)

    def _build_db_card(self, parent):
        self.db_card = ttk.LabelFrame(parent, text="数据库连接(输入或输出为 SQL 时展开)",
                                      style="DB.TLabelframe", padding=8)
        self.db_card.columnconfigure(1, weight=1)
        self.db_card.columnconfigure(4, weight=1)

        ttk.Label(self.db_card, text="类型:", style="Card.TLabel").grid(
            row=0, column=0, sticky="e", padx=4, pady=2)
        self.db_type_combo = ttk.Combobox(
            self.db_card, textvariable=self.db_type_var, values=list(ADAPTERS),
            state="readonly", width=12)
        self.db_type_combo.grid(row=0, column=1, sticky="w", padx=4, pady=2)
        self.db_type_combo.bind("<<ComboboxSelected>>", self.on_db_type_changed)

        ttk.Label(self.db_card, text="主机:", style="Card.TLabel").grid(
            row=0, column=2, sticky="e", padx=4, pady=2)
        ttk.Entry(self.db_card, textvariable=self.host_var, width=15).grid(
            row=0, column=3, sticky="w", padx=4, pady=2)
        ttk.Label(self.db_card, text="端口:", style="Card.TLabel").grid(
            row=0, column=4, sticky="e", padx=4, pady=2)
        ttk.Entry(self.db_card, textvariable=self.port_var, width=8).grid(
            row=0, column=5, sticky="w", padx=4, pady=2)
        ttk.Label(self.db_card, text="用户名:", style="Card.TLabel").grid(
            row=0, column=6, sticky="e", padx=4, pady=2)
        ttk.Entry(self.db_card, textvariable=self.user_var, width=12).grid(
            row=0, column=7, sticky="w", padx=4, pady=2)

        ttk.Label(self.db_card, text="密码:", style="Card.TLabel").grid(
            row=1, column=0, sticky="e", padx=4, pady=2)
        ttk.Entry(self.db_card, textvariable=self.pwd_var, width=12, show="*").grid(
            row=1, column=1, sticky="w", padx=4, pady=2)
        ttk.Label(self.db_card, text="数据库:", style="Card.TLabel").grid(
            row=1, column=2, sticky="e", padx=4, pady=2)
        ttk.Entry(self.db_card, textvariable=self.database_var, width=15).grid(
            row=1, column=3, sticky="w", padx=4, pady=2)
        ttk.Label(self.db_card, text="超时(秒):", style="Card.TLabel").grid(
            row=1, column=4, sticky="e", padx=4, pady=2)
        ttk.Entry(self.db_card, textvariable=self.timeout_var, width=8).grid(
            row=1, column=5, sticky="w", padx=4, pady=2)

        btn_box = ttk.Frame(self.db_card, style="Card.TFrame")
        btn_box.grid(row=1, column=6, columnspan=2, sticky="w", padx=4, pady=2)
        ttk.Button(btn_box, text="测试连接", command=self.on_test_connection).pack(
            side="left", padx=2)
        ttk.Button(btn_box, text="连接并加载表", style="Blue.TButton",
                   command=self.on_connect).pack(side="left", padx=2)
        ttk.Button(btn_box, text="断开", command=self.disconnect).pack(
            side="left", padx=2)

    def _build_output_card(self, parent):
        self.output_card = ttk.LabelFrame(parent, text="③ 输出设置", padding=8)
        self.output_card.grid(row=2, column=0, sticky="ew", pady=(6, 0))
        self.output_card.columnconfigure(0, weight=1)

        # ---- 输出 = SQL ----
        self.out_db_frame = ttk.Frame(self.output_card, style="Card.TFrame")
        self.out_db_frame.columnconfigure(1, weight=1)
        ttk.Label(self.out_db_frame, text="目标表名:",
                  style="Card.TLabel").grid(row=0, column=0, sticky="e", padx=4, pady=4)
        ttk.Entry(self.out_db_frame, textvariable=self.out_table_var,
                  width=28).grid(row=0, column=1, sticky="w", padx=4, pady=4)
        ttk.Label(self.out_db_frame, text="写入模式:",
                  style="Card.TLabel").grid(row=0, column=2, sticky="e", padx=4, pady=4)
        ttk.Combobox(self.out_db_frame, textvariable=self.out_mode_var,
                     values=MODE_LABELS, state="readonly",
                     width=36).grid(row=0, column=3, sticky="w", padx=4, pady=4)
        ttk.Label(self.out_db_frame, text="每批行数:",
                  style="Card.TLabel").grid(row=0, column=4, sticky="e", padx=4, pady=4)
        ttk.Entry(self.out_db_frame, textvariable=self.batch_var,
                  width=9).grid(row=0, column=5, sticky="w", padx=4, pady=4)

        # ---- 输出 = 文件 ----
        self.out_file_frame = ttk.Frame(self.output_card, style="Card.TFrame")
        self.out_file_frame.columnconfigure(1, weight=1)
        ttk.Label(self.out_file_frame, text="Sheet名/JSON键:",
                  style="Card.TLabel").grid(row=0, column=0, sticky="e", padx=4, pady=4)
        self.out_sheet_entry = ttk.Entry(self.out_file_frame,
                                         textvariable=self.out_sheet_var, width=28)
        self.out_sheet_entry.grid(row=0, column=1, sticky="w", padx=4, pady=4)
        ttk.Label(self.out_file_frame,
                  text="(Excel 用 Sheet 名,JSON 用一级键,CSV 忽略此项)",
                  style="Muted.Card.TLabel").grid(row=0, column=2, padx=4, pady=4)

        self.out_csv_opts = ttk.Frame(self.output_card, style="Card.TFrame")
        ttk.Label(self.out_csv_opts, text="分隔符:",
                  style="Card.TLabel").grid(row=0, column=0, sticky="e", padx=4, pady=2)
        ttk.Combobox(self.out_csv_opts, textvariable=self.out_delim_var,
                     values=list(DELIM_LABELS), state="readonly",
                     width=18).grid(row=0, column=1, sticky="w", padx=4, pady=2)
        ttk.Label(self.out_csv_opts, text="编码:",
                  style="Card.TLabel").grid(row=0, column=2, sticky="e", padx=4, pady=2)
        ttk.Combobox(self.out_csv_opts, textvariable=self.out_enc_var,
                     values=list(ENC_LABELS), state="readonly",
                     width=24).grid(row=0, column=3, sticky="w", padx=4, pady=2)

    # ------------------------------------------------------- 线程 / 消息队列
    def log(self, message, level="info"):
        """线程安全日志:worker 线程调用,主线程统一渲染。"""
        self.msg_queue.put(("log", (str(message), level)))

    def set_status(self, message, connected=False):
        self.msg_queue.put(("status", (str(message), connected)))

    def run_task(self, fn):
        """在后台线程执行耗时任务,避免界面卡死。"""
        if self._busy:
            messagebox.showinfo("提示", "已有任务正在执行,请等待完成后再试。")
            return False
        self._busy = True
        self.status_var.set("运行中...")
        self.status_chip.configure(style="Warn.Status.TLabel")

        def worker():
            try:
                fn()
            except Exception as exc:  # noqa: BLE001
                self.log(f"发生错误: {exc}", "err")
                self.log(traceback.format_exc(), "err")
                self.msg_queue.put(("error", str(exc)))
            finally:
                self.msg_queue.put(("done", None))

        threading.Thread(target=worker, daemon=True).start()
        return True

    def _poll_queue(self):
        try:
            while True:
                kind, payload = self.msg_queue.get_nowait()
                if kind == "log":
                    message, level = payload
                    self.log_text.configure(state="normal")
                    self.log_text.insert("end", message + "\n", level)
                    self.log_text.see("end")
                    self.log_text.configure(state="disabled")
                elif kind == "status":
                    message, connected = payload
                    self.status_var.set(message)
                    self.status_chip.configure(
                        style="Ok.Status.TLabel" if connected else "Status.TLabel")
                elif kind == "tables":
                    self.table_dropdown.set_items(payload)
                    self.root.after(10, self.on_tables_confirmed,
                                    self.table_dropdown.get_selected())
                elif kind == "columns":
                    self.column_dropdown.set_items(payload, checked=payload)
                elif kind == "sheets":
                    self.in_sheet_combo["values"] = payload
                    if payload:
                        self.in_sheet_combo.current(0)
                        self.in_sheet_var.set(payload[0])
                        self.on_in_sheet_selected()
                elif kind == "json_keys":
                    self.in_key_combo["values"] = payload
                    if payload:
                        self.in_key_combo.current(0)
                        self.in_key_var.set(payload[0])
                        self.on_in_key_selected()
                elif kind == "error":
                    messagebox.showerror("错误", payload)
                elif kind == "info":
                    messagebox.showinfo("提示", payload)
                elif kind == "done":
                    self._busy = False
                    connected = self.conn is not None
                    self.status_var.set(
                        f"{'已连接数据库' if connected else '未连接数据库'} | 就绪")
                    self.status_chip.configure(
                        style="Ok.Status.TLabel" if connected else "Status.TLabel")
        except queue.Empty:
            pass
        self.root.after(100, self._poll_queue)

    # ------------------------------------------------------------- 模式切换
    @staticmethod
    def _key_of(value, mapping):
        return next((k for k, v in mapping.items() if v == value), None)

    def input_key(self):
        return self._key_of(self.input_mode_var.get(), INPUT_LABELS)

    def output_key(self):
        return self._key_of(self.output_mode_var.get(), OUTPUT_LABELS)

    def on_input_mode_changed(self, _event=None):
        key = self.input_key()
        self.input_card.configure(style=f"CardIn{key}.TLabelframe")
        if key == "SQL":
            self.src_sql_frame.grid(row=0, column=0, sticky="ew")
        else:
            self.src_sql_frame.grid_remove()
            self.src_file_frame.grid(row=0, column=0, sticky="ew")
            self.file_excel_opts.grid_remove()
            self.file_json_opts.grid_remove()
            self.file_csv_opts.grid_remove()
            if key == "Excel":
                self.file_excel_opts.grid(row=1, column=0, columnspan=3, sticky="w")
            elif key == "JSON":
                self.file_json_opts.grid(row=1, column=0, columnspan=3, sticky="w")
            elif key == "CSV":
                self.file_csv_opts.grid(row=1, column=0, columnspan=3, sticky="w")
        self._sync_db_card()
        self._auto_fill_output_names()
        self.log(f"输入格式切换为: {INPUT_LABELS[key]}", "info")

    def on_output_mode_changed(self, _event=None):
        key = self.output_key()
        self.output_card.configure(style=f"CardOut{key}.TLabelframe")
        if key == "SQL":
            self.out_db_frame.grid(row=0, column=0, sticky="ew")
            self.out_file_frame.grid_remove()
            self.out_csv_opts.grid_remove()
        else:
            self.out_db_frame.grid_remove()
            self.out_file_frame.grid(row=0, column=0, sticky="ew")
            self.out_csv_opts.grid_remove()
            if key == "CSV":
                self.out_csv_opts.grid(row=1, column=0, columnspan=3, sticky="w")
        self._sync_db_card()
        self._auto_fill_output_names()
        self.log(f"输出格式切换为: {OUTPUT_LABELS[key]}", "info")

    def _sync_db_card(self):
        need_db = self.input_key() == "SQL" or self.output_key() == "SQL"
        if need_db:
            self.db_card.grid(row=1, column=0, sticky="ew", pady=(0, 6))
        else:
            self.db_card.grid_remove()

    def _auto_fill_output_names(self):
        """根据输入选择自动填充输出侧的 Sheet/JSON 键/目标表名(可手改)。"""
        in_key = self.input_key()
        base = ""
        if in_key == "SQL":
            base = (self.selected_table or "").split(".")[-1]
        elif in_key == "Excel":
            base = self.in_sheet_var.get() or Path(self.in_file_var.get()).stem
        elif in_key == "JSON":
            base = self.in_key_var.get() or Path(self.in_file_var.get()).stem
        elif in_key == "CSV":
            base = Path(self.in_file_var.get()).stem
        if not base:
            return
        if not self.out_sheet_var.get() or self.out_sheet_var.get() == self._auto_sheet:
            self.out_sheet_var.set(base)
            self._auto_sheet = base
        if not self.out_table_var.get() or self.out_table_var.get() == self._auto_table:
            self.out_table_var.set(base)
            self._auto_table = base

    # ------------------------------------------------------------- 连接管理
    def get_cfg(self):
        return {
            "host": self.host_var.get().strip(),
            "port": self.port_var.get().strip() or str(self.adapter.DEFAULT_PORT),
            "user": self.user_var.get().strip(),
            "password": self.pwd_var.get(),
            "database": self.database_var.get().strip(),
            "charset": "utf8mb4",
            "timeout": self.timeout_var.get().strip() or "10",
        }

    def connect_db(self):
        cfg = self.get_cfg()
        if not cfg["host"] or not cfg["user"] or not cfg["database"]:
            raise ValueError("请填写【主机】【用户名】【数据库】后再连接")
        self._disconnect_conn()
        self.conn = self.adapter.connect(cfg)
        return self.adapter.server_version(self.conn)

    def _disconnect_conn(self):
        if self.conn is not None:
            try:
                self.conn.close()
            except Exception:
                pass
            self.conn = None

    def disconnect(self):
        self._disconnect_conn()
        self.status_var.set("未连接数据库 | 就绪")
        self.status_chip.configure(style="Status.TLabel")

    def _ping(self):
        try:
            if self.adapter.NAME == "MySQL":
                self.conn.ping(reconnect=False)
                return True
            return not self.conn.closed
        except Exception:
            return False

    def ensure_conn(self):
        """确保有可用连接(worker 线程内调用)。"""
        try:
            if self.conn is not None and self._ping():
                return self.conn
        except Exception:
            pass
        self._disconnect_conn()
        version = self.connect_db()
        self.log(f"已自动连接 {self.adapter.NAME} {version}", "ok")
        self.set_status(f"已连接 {self.adapter.NAME} {version}", connected=True)
        return self.conn

    # ------------------------------------------------------------- UI 事件
    def on_db_type_changed(self, _event=None):
        name = self.db_type_var.get()
        self.adapter = ADAPTERS[name]
        self.port_var.set(str(self.adapter.DEFAULT_PORT))
        self.user_var.set("root" if name == "MySQL" else "postgres")
        self.disconnect()
        self.log(f"已切换到 {name},默认端口 {self.adapter.DEFAULT_PORT}", "info")

    def on_test_connection(self):
        def work():
            version = self.connect_db()
            self.log(f"连接成功: {self.adapter.NAME} {version}", "ok")
            self.msg_queue.put(("info", f"连接成功!\n{self.adapter.NAME} {version}"))
            self._disconnect_conn()
        self.run_task(work)

    def on_connect(self):
        def work():
            version = self.connect_db()
            tables = self.adapter.list_tables(self.conn)
            self.log(f"连接成功: {self.adapter.NAME} {version},共 {len(tables)} 张表", "ok")
            self.msg_queue.put(("tables", tables))
            self.set_status(f"已连接 {self.adapter.NAME} {version}", connected=True)
        self.run_task(work)

    def on_refresh_tables(self):
        def work():
            conn = self.ensure_conn()
            tables = self.adapter.list_tables(conn)
            self.log(f"共发现 {len(tables)} 张表", "info")
            self.msg_queue.put(("tables", tables))
        self.run_task(work)

    def on_tables_confirmed(self, selected):
        if not selected:
            self.selected_table = None
            self.column_dropdown.set_items([])
            return
        table = selected[0]
        self.selected_table = table
        self._auto_fill_output_names()

        def work():
            conn = self.ensure_conn()
            columns = self.adapter.list_columns(conn, table)
            self.log(f"表【{table}】共 {len(columns)} 个字段,已默认全选", "ok")
            self.msg_queue.put(("columns", columns))
        self.run_task(work)

    def on_pick_input_file(self):
        key = self.input_key()
        filetypes = {
            "Excel": [("Excel 工作簿", "*.xlsx"), ("所有文件", "*.*")],
            "JSON": [("JSON 文件", "*.json"), ("所有文件", "*.*")],
            "CSV": [("CSV 文件", "*.csv"), ("所有文件", "*.*")],
        }[key]
        path = filedialog.askopenfilename(
            title=f"选择输入文件 ({INPUT_LABELS[key]})", filetypes=filetypes)
        if not path:
            return
        self.in_file_var.set(path)
        self._auto_fill_output_names()

        if key == "Excel":
            def work():
                wb = load_workbook(path, read_only=True)
                try:
                    names = wb.sheetnames
                finally:
                    wb.close()
                self.log(f"工作簿包含工作表: {', '.join(names)}", "info")
                self.msg_queue.put(("sheets", names))
            self.run_task(work)
        elif key == "JSON":
            def work():
                with open(path, "r", encoding="utf-8") as fp:
                    doc = json.load(fp)
                keys = list(doc.keys()) if isinstance(doc, dict) else []
                self.log(f"JSON 包含一级键: {', '.join(map(str, keys))}", "info")
                self.msg_queue.put(("json_keys", keys))
            self.run_task(work)
        elif key == "CSV":
            def work():
                header = read_csv_header(
                    path, DELIM_LABELS[self.in_delim_var.get()],
                    ENC_LABELS[self.in_enc_var.get()])
                self.log(f"CSV 表头列: {header}", "info")
            self.run_task(work)

    def on_in_sheet_selected(self, _event=None):
        self._auto_fill_output_names()

    def on_in_key_selected(self, _event=None):
        self._auto_fill_output_names()

    @staticmethod
    def _parse_int(var, default):
        try:
            return max(1, int(str(var.get()).strip()))
        except (TypeError, ValueError):
            return default

    def ask_conflict_mode(self, existing_paths):
        """同名文件冲突纠错对话框,返回 'replace' / 'merge' / None(取消)。"""
        dlg = tk.Toplevel(self.root)
        dlg.title("检测到同名文件")
        dlg.transient(self.root)
        dlg.configure(bg=CARD)
        dlg.grab_set()
        dlg.resizable(False, False)
        lines = "\n".join(f"  • {p}" for p in existing_paths)
        ttk.Label(
            dlg, justify="left", style="Card.TLabel",
            text=(f"以下文件已存在:\n{lines}\n\n请选择处理方式:"),
        ).pack(padx=14, pady=(14, 8))
        box = ttk.Frame(dlg, style="Card.TFrame")
        box.pack(padx=14, pady=(0, 14))
        result = {"mode": None}

        def pick(mode):
            result["mode"] = mode
            dlg.destroy()

        ttk.Button(box, text="覆盖整个文件", style="Warn.TButton",
                   command=lambda: pick("replace")).pack(side="left", padx=4)
        ttk.Button(box, text="合并写入(覆盖同名Sheet/JSON键;CSV追加行)",
                   style="Blue.TButton",
                   command=lambda: pick("merge")).pack(side="left", padx=4)
        ttk.Button(box, text="取消",
                   command=lambda: pick(None)).pack(side="left", padx=4)
        dlg.wait_window()
        return result["mode"]

    # ------------------------------------------------------------- 主转换流程
    def _ask_output_path(self, out_key, default_base):
        """输出为文件时:弹出保存对话框 + 同名文件纠错。"""
        ext = {"Excel": ".xlsx", "JSON": ".json", "CSV": ".csv"}[out_key]
        label = OUTPUT_LABELS[out_key]
        path = filedialog.asksaveasfilename(
            title=f"保存输出文件 ({label})", defaultextension=ext,
            initialdir=str(Path.home() / "Documents"),
            initialfile=f"{safe_filename(default_base)}{ext}",
            filetypes=[(label, f"*{ext}")])
        if not path:
            return None, None
        if os.path.exists(path):
            mode = self.ask_conflict_mode([path])
            if mode is None:
                self.log("已取消转换(存在同名文件)", "warn")
                return None, None
        else:
            mode = "replace"
        return path, mode

    def _default_output_base(self, in_key):
        if in_key == "SQL":
            db = self.database_var.get().strip() or (self.selected_table or "export")
            return db
        return Path(self.in_file_var.get()).stem or "export"

    def _output_mode_key(self):
        label = self.out_mode_var.get()
        if MODE_REPLACE in label:
            return "replace"
        if MODE_CREATE in label:
            return "create_if_missing"
        return "append"

    def on_convert(self):
        if self._busy:
            messagebox.showinfo("提示", "已有任务正在执行,请等待完成后再试。")
            return
        in_key = self.input_key()
        out_key = self.output_key()
        adapter = self.adapter
        batch = self._parse_int(self.batch_var, DEFAULT_BATCH)

        # ---------- 校验输入 ----------
        source = {}
        if in_key == "SQL":
            table = self.selected_table
            if not table:
                messagebox.showwarning("提示", "请连接数据库,并在“选择数据表”下拉框中勾选数据表")
                return
            columns = self.column_dropdown.get_selected()
            if not columns:
                messagebox.showwarning("提示", "请在“选择字段”下拉框中勾选至少一个字段")
                return
            source = {"table": table, "columns": columns}
        elif in_key == "Excel":
            if not self.in_file_var.get() or not os.path.isfile(self.in_file_var.get()):
                messagebox.showwarning("提示", "请先选择有效的 Excel 输入文件")
                return
            if not self.in_sheet_var.get():
                messagebox.showwarning("提示", "请选择要读取的工作表")
                return
        elif in_key == "JSON":
            if not self.in_file_var.get() or not os.path.isfile(self.in_file_var.get()):
                messagebox.showwarning("提示", "请先选择有效的 JSON 输入文件")
                return
            if not self.in_key_var.get():
                messagebox.showwarning("提示", "请选择 JSON 一级键(表名)")
                return
        elif in_key == "CSV":
            if not self.in_file_var.get() or not os.path.isfile(self.in_file_var.get()):
                messagebox.showwarning("提示", "请先选择有效的 CSV 输入文件")
                return

        # ---------- 校验/询问输出 ----------
        output = {}
        if out_key == "SQL":
            table_out = self.out_table_var.get().strip()
            if not table_out:
                messagebox.showwarning("提示", "请填写目标表名")
                return
            mode = self._output_mode_key()
            if mode == "replace" and not messagebox.askyesno(
                    "危险操作", f"将先删除表【{table_out}】(若存在)再重建并写入,确定继续吗?"):
                return
            output = {"table": table_out, "mode": mode, "batch": batch}
        else:
            default_base = self._default_output_base(in_key)
            path, conflict_mode = self._ask_output_path(out_key, default_base)
            if path is None:
                return
            sheet = safe_sheet_title(self.out_sheet_var.get() or default_base)
            key_name = (self.out_sheet_var.get() or "").strip() or default_base
            delim = DELIM_LABELS[self.out_delim_var.get()]
            enc = ENC_LABELS[self.out_enc_var.get()]
            output = {"path": path, "mode": conflict_mode, "sheet": sheet,
                      "key": key_name, "delimiter": delim, "encoding": enc}

        # ---------- 后台执行管道 ----------
        def work():
            conn = self.ensure_conn() if (in_key == "SQL" or out_key == "SQL") else None
            cur = None
            header = rows = None
            try:
                # 1) 载入输入源
                if in_key == "SQL":
                    self.log(f"读取数据表: {source['table']}(字段: {', '.join(source['columns'])})", "head")
                    header, rows, cur = load_sql_stream(
                        conn, adapter, source["table"], source["columns"])
                elif in_key == "Excel":
                    self.log(f"读取 Excel: {self.in_file_var.get()} → [{self.in_sheet_var.get()}]", "head")
                    header, rows = load_excel_stream(
                        self.in_file_var.get(), self.in_sheet_var.get())
                elif in_key == "JSON":
                    self.log(f"读取 JSON: {self.in_file_var.get()} → 键[{self.in_key_var.get()}]", "head")
                    header, rows = load_json_stream(
                        self.in_file_var.get(), self.in_key_var.get())
                elif in_key == "CSV":
                    delim = DELIM_LABELS[self.in_delim_var.get()]
                    enc = ENC_LABELS[self.in_enc_var.get()]
                    self.log(f"读取 CSV: {self.in_file_var.get()}(分隔符 {delim!r},编码 {enc})", "head")
                    header = read_csv_header(self.in_file_var.get(), delim, enc)
                    rows = iter_csv_rows(self.in_file_var.get(), delim, enc,
                                         len(header), parse=(out_key == "SQL"))
                self.log(f"共 {len(header)} 列: {header}", "info")

                # 2) 写出输出目标
                if out_key == "SQL":
                    if in_key == "SQL":
                        # SQL→SQL:先采样前 SAMPLE_SIZE 行(游标可继续流式读),
                        # 建表/插入走独立连接,避免 MySQL SSCursor 挂起结果集
                        # 时执行 DDL、以及 PG 命名游标被 rollback 失效的问题。
                        sample, buffered = [], []
                        for row in rows:
                            buffered.append(list(row))
                            if len(sample) < SAMPLE_SIZE:
                                sample.append(list(row))
                            if len(buffered) >= SAMPLE_SIZE:
                                break

                        def chained():
                            for row in buffered:
                                yield row
                            for row in rows:  # rows 生成器保留剩余未读数据
                                yield list(row)

                        conn_out = adapter.connect(self.get_cfg())
                        try:
                            total = write_sql_table(
                                conn_out, adapter, output["table"], header,
                                chained(), output["mode"], output["batch"],
                                self.log, sample=sample)
                        finally:
                            conn_out.close()
                    else:
                        total = write_sql_table(
                            conn, adapter, output["table"], header, rows,
                            output["mode"], output["batch"], self.log)
                    self.log(f"✅ 写入数据库完成: 表【{output['table']}】共 {total:,} 行", "ok")
                    self.msg_queue.put(
                        ("info", f"写入完成,共 {total:,} 行 → 表【{output['table']}】"))
                elif out_key == "Excel":
                    total = write_excel_file(
                        output["path"], output["sheet"], header, rows,
                        mode=output["mode"], log=self.log)
                    self.log(f"✅ Excel 导出完成: {output['path']}(Sheet [{output['sheet']}],共 {total:,} 行)", "ok")
                    self.msg_queue.put(
                        ("info", f"导出完成,共 {total:,} 行:\n{output['path']}"))
                elif out_key == "JSON":
                    row_list = list(rows)
                    write_json_file(output["path"], output["key"], header, row_list,
                                    mode=output["mode"], log=self.log)
                    self.log(f"✅ JSON 导出完成: {output['path']}(键 [{output['key']}],共 {len(row_list):,} 行)", "ok")
                    self.msg_queue.put(
                        ("info", f"导出完成,共 {len(row_list):,} 行:\n{output['path']}"))
                elif out_key == "CSV":
                    total = write_csv_file(
                        output["path"], header, rows, delimiter=output["delimiter"],
                        encoding=output["encoding"], mode=output["mode"], log=self.log)
                    self.log(f"✅ CSV 导出完成: {output['path']}(共 {total:,} 行)", "ok")
                    self.msg_queue.put(
                        ("info", f"导出完成,共 {total:,} 行:\n{output['path']}"))
            finally:
                if cur is not None:
                    try:
                        cur.close()
                    except Exception:
                        pass
                    # 结束源查询事务(PG 命名游标必须 rollback;
                    # MySQL 也 rollback 避免残留快照影响后续读取)
                    try:
                        conn.rollback()
                    except Exception:
                        pass

        self.run_task(work)

    # ------------------------------------------------------------- 退出
    def on_close(self):
        self._disconnect_conn()
        self.root.destroy()


def run_self_test():
    """--selftest: 无界面自检(打包后验证驱动/引擎是否齐全),结果写入当前目录 selftest.log。"""
    log_path = Path.cwd() / "selftest.log"
    lines = []

    for module_name in ("pymysql", "psycopg2", "openpyxl"):
        try:
            __import__(module_name)
            lines.append(f"[OK]   import {module_name}")
        except Exception as exc:
            lines.append(f"[FAIL] import {module_name}: {exc}")

    # Excel 写读往返(覆盖 openpyxl + et_xmlfile)
    try:
        tmp = log_path.parent / "_selftest.xlsx"
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title="自检")
        write_rows_to_worksheet(ws, ["id", "name"], iter([[1, "张三"], [2, "李四"]]))
        wb.save(tmp)
        wb.close()
        wb2 = load_workbook(tmp, read_only=True, data_only=True)
        rows = list(wb2[wb2.sheetnames[0]].iter_rows(values_only=True))
        wb2.close()
        tmp.unlink(missing_ok=True)
        ok = len(rows) == 3 and rows[1][0] == 1
        lines.append(f"[{'OK' if ok else 'FAIL'}]   openpyxl 写读往返: {len(rows)} 行")
    except Exception as exc:
        lines.append(f"[FAIL] openpyxl 写读往返: {exc}")

    # JSON 写读往返(覆盖列数组结构 + 同名键合并逻辑)
    try:
        tmp = log_path.parent / "_selftest.json"
        rows = [[1, "张三"], [2, "李四"]]
        write_json_file(str(tmp), "info", ["Id", "Name"], rows)
        with open(tmp, "r", encoding="utf-8") as fp:
            doc = json.load(fp)
        ok = doc["info"]["Id"] == [1, 2] and doc["info"]["Name"][-1] == "李四"
        write_json_file(str(tmp), "info", ["Id", "Name"], [[3, "王五"]], mode="merge")
        write_json_file(str(tmp), "other", ["X"], [[9]], mode="merge")
        with open(tmp, "r", encoding="utf-8") as fp:
            doc = json.load(fp)
        ok = ok and doc["info"]["Id"] == [3] and "other" in doc
        tmp.unlink(missing_ok=True)
        lines.append(f"[{'OK' if ok else 'FAIL'}]   JSON 写读/合并往返")
    except Exception as exc:
        lines.append(f"[FAIL] JSON 写读/合并往返: {exc}")

    # CSV 写读往返(覆盖 csv 引擎 + 类型解析)
    try:
        tmp = log_path.parent / "_selftest.csv"
        write_csv_file(str(tmp), ["Id", "Name"], iter([[1, "张三"], [2, "李四"]]))
        header = read_csv_header(str(tmp), ",", "utf-8-sig")
        rows = list(iter_csv_rows(str(tmp), ",", "utf-8-sig", len(header), parse=True))
        ok = header == ["Id", "Name"] and rows[0] == [1, "张三"] and rows[1][0] == 2
        tmp.unlink(missing_ok=True)
        lines.append(f"[{'OK' if ok else 'FAIL'}]   CSV 写读往返: {len(rows)} 行")
    except Exception as exc:
        lines.append(f"[FAIL] CSV 写读往返: {exc}")

    # 若提供数据库连接环境变量,则做真实连接自检
    for env_prefix, adapter in (("SELFTEST_MYSQL", ADAPTERS["MySQL"]),
                                ("SELFTEST_PG", ADAPTERS["PostgreSQL"])):
        env = {
            "host": os.environ.get(f"{env_prefix}_HOST", "127.0.0.1"),
            "port": os.environ.get(f"{env_prefix}_PORT") or str(adapter.DEFAULT_PORT),
            "user": os.environ.get(f"{env_prefix}_USER", ""),
            "password": os.environ.get(f"{env_prefix}_PASSWORD", ""),
            "database": os.environ.get(f"{env_prefix}_DATABASE", ""),
            "charset": "utf8mb4",
            "timeout": "10",
        }
        if not env["user"] or not env["database"]:
            lines.append(f"[SKIP] {adapter.NAME} 连接自检(未设置 {env_prefix}_* 环境变量)")
            continue
        try:
            conn = adapter.connect(env)
            with conn.cursor() as cur:
                cur.execute("SELECT 1")
                cur.fetchone()
            conn.close()
            lines.append(f"[OK]   {adapter.NAME} 连接 SELECT 1")
        except Exception as exc:
            lines.append(f"[FAIL] {adapter.NAME} 连接: {exc}")

    log_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return all(not line.startswith("[FAIL]") for line in lines)


def main():
    if "--selftest" in sys.argv:
        sys.exit(0 if run_self_test() else 1)
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
