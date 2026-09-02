'''建表语句
mysql> CREATE TABLE VirtualProfile (
    Id INT NOT NULL AUTO_INCREMENT COMMENT '主键ID，自增',
    Name VARCHAR(50) COMMENT '姓名',
    Province VARCHAR(50) COMMENT '省份',
    PhoneNumber VARCHAR(20) COMMENT '电话号码',
    IdCard VARCHAR(30) COMMENT '身份证号',
    Email VARCHAR(100) COMMENT '电子邮箱',
    Birthday DATE COMMENT '生日',
    Company VARCHAR(100) COMMENT '公司名称',
    Bank VARCHAR(100) COMMENT '银行',
    BankAccount VARCHAR(100) COMMENT '银行账户',
    LicensePlate VARCHAR(100) COMMENT '车牌号',
    Address VARCHAR(200) COMMENT '地址',
    CreateTime TIMESTAMP DEFAULT CURRENT_TIMESTAMP COMMENT '创建时间',
    PRIMARY KEY (Id)
 ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci COMMENT='虚拟信息表';
Query OK, 0 rows affected (0.01 sec)

'''



from faker import Faker
from pymysql import connect
from typing import List, Tuple
from sys import argv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from colorama import init, Fore, Style

init(autoreset=True)

num = int(argv[1]) if len(argv) > 1 and argv[1].isdigit() else 15

fk = Faker(locale="zh_CN")
data: List[Tuple] = [
    (
        fk.name(),
        fk.province(),
        fk.phone_number(),
        fk.ssn(),
        fk.ascii_free_email(),
        fk.date_of_birth(minimum_age=18, maximum_age=65),
        fk.company(),
        fk.bank(),
        fk.bban(),
        fk.license_plate(),
        fk.address()
    ) for _ in range(num)
]

# ==================== Excel 导出（不变） ====================
headers = ["Name", "Province", "PhoneNumber", "IdCard", "Email", "Birthday", "Company", "Bank","BankAccount", "LicensePlate", "Address"]
excel_folder = Path(__file__).parent / "Excel"
excel_folder.mkdir(exist_ok=True)

excel_filename = f"{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.xlsx"
excel_path = excel_folder / excel_filename

try:
    wb = Workbook()
    ws = wb.active
    ws.title = "VirtualProfile"
    ws.append(headers)
    for row in data:
        ws.append(row)
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = cell_len
        ws.column_dimensions[col_letter].width = max_len + 2
    wb.save(excel_path)
    print(f"{Fore.GREEN}数据已成功导出到Excel文件：<{excel_path}>。{Style.RESET_ALL}")
except Exception as e:
    print(f"导出 Excel 文件时出错：{e}")

# ==================== 数据库连接 ====================
connection = connect(
    host="127.0.0.1",
    user="root",
    password="123",
    database="db",
    charset="utf8mb4"
)

# ==================== 插入 Virtual_Profile ====================
try:
    with connection.cursor() as cursor:
        sql = """
        INSERT INTO VirtualProfile
        (Name, Province, PhoneNumber, IdCard, Email, Birthday, Company, Bank, BankAccount, LicensePlate, Address)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        affected_rows = cursor.executemany(sql, data)
        print(f"{Fore.BLUE}{Style.BRIGHT}成功插入了 <{affected_rows}> 条数据。{Style.RESET_ALL}")
    connection.commit()
except Exception as err:
    print(f"插入数据时发生错误：<{err}>")
    connection.rollback()



# ==================== 关闭连接 ====================
connection.close()