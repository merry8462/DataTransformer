# -*- coding: utf-8 -*-
"""
SQL ↔ Excel / JSON 数据互转工具 (Tkinter GUI)
==============================================
功能:
    1. SQL -> Excel / JSON : 从 MySQL / PostgreSQL 读取表或自定义 SQL 查询结果,
       导出为 .xlsx 和/或 .json(两者数据一致,可单选或同时导出)
    2. Excel -> SQL : 读取 .xlsx 工作表,批量插入 MySQL / PostgreSQL 数据表

命名约定(与库-表-字段一一对应):
    数据库名  == Excel 文件名  == JSON 文件名(如 information.xlsx / information.json)
    表名      == Excel Sheet名 == JSON 一级参数(如 "info")
    字段名    == Excel 首行单元格 == JSON 二级参数(如 "Id"、"Name" ...)

设计说明:
    * 数据库适配:MySQL 使用 PyMySQL,PostgreSQL 使用 psycopg2,两者参数占位符均为 %s,
      上层统一批量 executemany 写入。
    * Excel 读写:选择 openpyxl(而非 pandas)。pandas 的 read_excel/to_excel 底层仍然
      调用 openpyxl 引擎,并且会先把整张表物化进内存 DataFrame;而 openpyxl 的
      read_only / write_only 流式模式可以逐行读取、逐批写入,配合数据库 executemany
      分批提交,内存占用更低、大批量数据转换效率更高。
    * JSON 导出:按示例采用“列数组”结构(一级键=表名,二级键=字段名,值为该列全部数据
      组成的数组),写文件前会检测同名文件并让用户选择覆盖/合并/取消。
"""

import json
import os
import queue
import re
import sys
import threading
import traceback
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path

import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

try:
    import pymysql
except Exception:  # pragma: no cover
    pymysql = None

try:
    import psycopg2
except Exception:  # pragma: no cover
    psycopg2 = None

# ---------------------------------------------------------------------------
# 常量
# ---------------------------------------------------------------------------
APP_TITLE = "SQL ↔ Excel / JSON 数据转换工具"
SAMPLE_SIZE = 200            # 建表时用于推断字段类型的采样行数
DEFAULT_BATCH = 1000         # 默认每批读写行数
FONT = ("Microsoft YaHei", 10)
FONT_BOLD = ("Microsoft YaHei", 10, "bold")

IDENT_RE = re.compile(r"^[^\W\d]\w*$")  # 标识符:字母/下划线开头,允许中文、数字、下划线

MODE_APPEND = "追加到已有表 (append)"
MODE_CREATE = "不存在则创建,存在则追加 (create if missing)"
MODE_REPLACE = "删除重建 (replace)"
MODE_LABELS = [MODE_APPEND, MODE_CREATE, MODE_REPLACE]

FORMAT_EXCEL = "Excel (.xlsx)"
FORMAT_JSON = "JSON (.json)"
FORMAT_BOTH = "Excel + JSON (.xlsx + .json)"
FORMAT_LABELS = [FORMAT_EXCEL, FORMAT_JSON, FORMAT_BOTH]


# ---------------------------------------------------------------------------
# 数据库适配层:MySQL / PostgreSQL
# ---------------------------------------------------------------------------
class DBAdapter:
    """数据库适配器基类,MySQL 与 PostgreSQL 各自实现差异点。"""

    NAME = ""
    DEFAULT_PORT = 3306
    QUOTE = '"'

    def connect(self, cfg):
        raise NotImplementedError

    def open_query_cursor(self, conn):
        """打开一个流式查询游标(避免一次性把全部结果缓冲到客户端内存)。"""
        raise NotImplementedError

    def list_tables(self, conn):
        raise NotImplementedError

    def is_missing_table_error(self, exc):
        raise NotImplementedError

    # -- 通用能力 ----------------------------------------------------------
    def quote_ident(self, name):
        """校验并引用标识符,支持 schema.table 形式,防止 SQL 注入。"""
        name = str(name).strip()
        if not name:
            raise ValueError("表名/字段名不能为空")
        parts = [p.strip() for p in name.split(".") if p.strip()]
        for part in parts:
            if not IDENT_RE.match(part):
                raise ValueError(f"非法标识符(仅允许字母/下划线/数字): {name!r}")
        q = self.QUOTE
        return ".".join(f"{q}{p}{q}" for p in parts)

    def table_exists(self, conn, quoted_table):
        """通过 SELECT 探测表是否存在(跨库、跨 schema 均可靠)。"""
        try:
            with conn.cursor() as cur:
                cur.execute(f"SELECT 1 FROM {quoted_table} WHERE 1=0")
            return True
        except Exception as exc:
            # PostgreSQL 中任何出错语句都会使当前事务进入 aborted 状态,
            # 必须先 rollback 才能继续后续 SQL。
            try:
                conn.rollback()
            except Exception:
                pass
            if self.is_missing_table_error(exc):
                return False
            raise

    def server_version(self, conn):
        with conn.cursor() as cur:
            cur.execute("SELECT VERSION()")
            return str(cur.fetchone()[0])


class MySQLAdapter(DBAdapter):
    NAME = "MySQL"
    DEFAULT_PORT = 3306
    QUOTE = "`"

    def connect(self, cfg):
        if pymysql is None:
            raise RuntimeError("未安装 PyMySQL,请先执行: pip install pymysql")
        return pymysql.connect(
            host=cfg["host"],
            port=int(cfg["port"]),
            user=cfg["user"],
            password=cfg["password"],
            database=cfg["database"],
            charset=cfg.get("charset") or "utf8mb4",
            connect_timeout=int(cfg.get("timeout", 10)),
        )

    def open_query_cursor(self, conn):
        # SSCursor = 服务端无缓冲游标,大数据量导出不占客户端内存
        return conn.cursor(pymysql.cursors.SSCursor)

    def list_tables(self, conn):
        with conn.cursor() as cur:
            cur.execute("SHOW TABLES")
            return [str(row[0]) for row in cur.fetchall()]

    def is_missing_table_error(self, exc):
        errno = exc.args[0] if exc.args else None
        return isinstance(exc, pymysql.err.ProgrammingError) and errno == 1146


class PostgresAdapter(DBAdapter):
    NAME = "PostgreSQL"
    DEFAULT_PORT = 5432
    QUOTE = '"'

    def connect(self, cfg):
        if psycopg2 is None:
            raise RuntimeError("未安装 psycopg2,请先执行: pip install psycopg2-binary")
        return psycopg2.connect(
            host=cfg["host"],
            port=int(cfg["port"]),
            user=cfg["user"],
            password=cfg["password"],
            dbname=cfg["database"],
            connect_timeout=int(cfg.get("timeout", 10)),
        )

    def open_query_cursor(self, conn):
        # 服务端命名游标:结果按 itersize 分批从服务器拉取
        cur = conn.cursor(name=f"sql_excel_cur_{os.getpid()}_{id(conn)}")
        cur.itersize = 5000
        return cur

    def list_tables(self, conn):
        with conn.cursor() as cur:
            cur.execute(
                "SELECT table_schema, table_name FROM information_schema.tables "
                "WHERE table_schema NOT IN ('pg_catalog', 'information_schema') "
                "ORDER BY table_schema, table_name"
            )
            return [f"{s}.{t}" if s != "public" else str(t) for s, t in cur.fetchall()]

    def is_missing_table_error(self, exc):
        return getattr(exc, "pgcode", None) == "42P01"


ADAPTERS = {
    "MySQL": MySQLAdapter(),
    "PostgreSQL": PostgresAdapter(),
}


# ---------------------------------------------------------------------------
# Excel 工具函数(openpyxl 流式模式)
# ---------------------------------------------------------------------------
def safe_sheet_title(name):
    """工作表名:去掉非法字符,长度 <= 31。"""
    name = re.sub(r'[\\/*?:\[\]]', "_", str(name)).strip()
    return (name[:31] or "Sheet1")


def display_width(text):
    """按显示宽度估算列宽(中文等全角字符算 2)。"""
    return sum(2 if ord(ch) > 0x7F else 1 for ch in str(text))


def to_excel_value(value):
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, bytes):
        try:
            return value.decode("utf-8")
        except UnicodeDecodeError:
            return value.decode("utf-8", "replace")
    return value  # int / float / Decimal / str / bool 等


def write_rows_to_worksheet(ws, header, row_iterator, log=None):
    """把数据行逐行写入 write_only 工作表,顺便自动计算列宽。"""
    ws.append([to_excel_value(h) for h in header])
    widths = {i + 1: min(display_width(h) + 2, 60) for i, h in enumerate(header)}
    total = 0
    for row in row_iterator:
        values = [to_excel_value(v) for v in row]
        ws.append(values)
        for idx, value in enumerate(values, start=1):
            if value is None:
                continue
            width = min(display_width(value) + 2, 60)
            if width > widths.get(idx, 0):
                widths[idx] = width
        total += 1
        if log and total % 20000 == 0:
            log(f"  已写出 {total:,} 行 ...")
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width
    return total


def is_blank_row(row):
    return all(
        v is None or (isinstance(v, str) and not v.strip())
        for v in row
    )


def normalize_row(row, ncols):
    """把 Excel 行对齐到表头列数:短了补 None,长了截断。"""
    row = list(row)
    if len(row) > ncols:
        return row[:ncols]
    return row + [None] * (ncols - len(row))


def to_db_value(value):
    if value is None:
        return None
    if isinstance(value, str):
        s = value.strip()
        return s or None
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    return value  # datetime / date / Decimal / bool / int / float 直接交给驱动


def parse_header(raw_row):
    """解析 Excel 第一行表头:去掉尾部空列,空列自动命名,重名自动加后缀。"""
    cells = list(raw_row)
    last = len(cells)
    while last > 0 and (
        cells[last - 1] is None
        or (isinstance(cells[last - 1], str) and not cells[last - 1].strip())
    ):
        last -= 1
    if last == 0:
        raise ValueError("Excel 第一行(表头)为空,无法导入")

    header, counts = [], {}
    for i in range(last):
        name = str(cells[i]).strip() if cells[i] is not None else ""
        if not name:
            name = f"column_{i + 1}"
        if name in counts:
            counts[name] += 1
            name = f"{name}_{counts[name]}"
        else:
            counts[name] = 1
        header.append(name)
    return header


def collect_stats(header, sample_rows):
    """统计采样行中各列出现的数据类型,用于 CREATE TABLE 的类型推断。"""
    stats = {
        i: {
            "str": False, "max_len": 0, "int": False, "big": False,
            "float": False, "bool": False, "datetime": False, "date": False,
        }
        for i in range(len(header))
    }
    for row in sample_rows:
        for i, value in enumerate(row):
            st = stats[i]
            if value is None:
                continue
            if isinstance(value, bool):
                st["bool"] = True
            elif isinstance(value, int):
                st["int"] = True
                if value > 2 ** 31 - 1 or value < -(2 ** 31):
                    st["big"] = True
            elif isinstance(value, (float, Decimal)):
                st["float"] = True
            elif isinstance(value, datetime):
                st["datetime"] = True
            elif isinstance(value, date):
                st["date"] = True
            else:
                st["str"] = True
                st["max_len"] = max(st["max_len"], len(str(value)))
    return stats


def decide_type(adapter, st):
    is_mysql = adapter.NAME == "MySQL"
    if st.get("str"):                   # 含文本(混合类型也按文本兜底)
        max_len = st.get("max_len", 0) or 1
        return "TEXT" if max_len > 4000 else f"VARCHAR({max_len})"
    if st.get("datetime"):
        return "TIMESTAMP"
    if st.get("date"):
        return "DATE"
    if st.get("float"):
        return "DOUBLE" if is_mysql else "DOUBLE PRECISION"
    if st.get("bool") and not st.get("int"):
        return "TINYINT(1)" if is_mysql else "BOOLEAN"
    if st.get("int"):
        return "BIGINT" if st.get("big") else "INT"
    return "VARCHAR(255)"


def build_create_table(adapter, table, header, stats):
    columns = [
        f"{adapter.quote_ident(col)} {decide_type(adapter, stats[idx])}"
        for idx, col in enumerate(header)
    ]
    return f"CREATE TABLE {adapter.quote_ident(table)} ({', '.join(columns)})"


# ---------------------------------------------------------------------------
# JSON / Excel 文件写出工具(含同名文件冲突处理)
# ---------------------------------------------------------------------------
def safe_filename(name):
    """去掉文件名中 Windows 不允许的字符。"""
    name = re.sub(r'[\\/:*?"<>|\r\n]+', "_", str(name)).strip()
    return name or "export"


def to_json_value(value):
    """把数据库返回值转换成 JSON 兼容值(时间统一为字符串,与示例文件一致)。"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, bytes):
        try:
            return value.decode("utf-8")
        except UnicodeDecodeError:
            return value.decode("utf-8", "replace")
    return value  # int / float / str / bool


def build_json_document(top_key, columns, rows):
    """构造 {表名: {字段名: [值...]}} 的列数组式 JSON 结构。"""
    doc = {top_key: {col: [] for col in columns}}
    for row in rows:
        norm = normalize_row(row, len(columns))
        for col, value in zip(columns, norm):
            doc[top_key][col].append(to_json_value(value))
    return doc


def write_json_file(path, top_key, columns, rows, mode="replace", log=None):
    """写出 JSON 文件。mode=merge 时保留原文件其他一级键,仅覆盖同名键。"""
    doc = build_json_document(top_key, columns, rows)
    if mode == "merge" and os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as fp:
                old = json.load(fp)
        except Exception:
            old = None
        if isinstance(old, dict):
            old[top_key] = doc[top_key]
            doc = old
    with open(path, "w", encoding="utf-8") as fp:
        json.dump(doc, fp, ensure_ascii=False, indent=4)
    return len(rows)


def write_excel_file(path, sheet, columns, row_iterator, mode="replace", log=None):
    """写出 Excel 文件。

    mode=replace : 用流式 write_only 模式重建整个工作簿(大数据量首选);
    mode=merge   : 保留原工作簿其他 Sheet,覆盖同名 Sheet(无同名则新增)。
    """
    if mode == "replace":
        wb = Workbook(write_only=True)
        try:
            ws = wb.create_sheet(title=sheet)
            total = write_rows_to_worksheet(ws, columns, row_iterator, log)
            wb.save(path)
        finally:
            try:
                wb.close()
            except Exception:
                pass
        return total

    # merge:打开已有工作簿(常规模式),只动同名 Sheet
    wb = load_workbook(path)
    try:
        if sheet in wb.sheetnames:
            del wb[sheet]
        ws = wb.create_sheet(title=sheet)
        total = write_rows_to_worksheet(ws, columns, row_iterator, log)
        wb.save(path)
    finally:
        wb.close()
    return total


# ---------------------------------------------------------------------------
# Tkinter 主界面
# ---------------------------------------------------------------------------
class App:
    def __init__(self, root):
        self.root = root
        root.title(APP_TITLE)
        root.geometry("1080x720")
        root.minsize(980, 660)

        self.conn = None
        self.adapter = ADAPTERS["MySQL"]
        self.msg_queue = queue.Queue()
        self._busy = False

        self._build_vars()
        self._build_ui()
        root.protocol("WM_DELETE_WINDOW", self.on_close)
        root.after(100, self._poll_queue)

    # ------------------------------------------------------------------ UI
    def _build_vars(self):
        self.db_type_var = tk.StringVar(value="MySQL")
        self.host_var = tk.StringVar(value="127.0.0.1")
        self.port_var = tk.StringVar(value="3306")
        self.user_var = tk.StringVar(value="root")
        self.pwd_var = tk.StringVar(value="")
        self.database_var = tk.StringVar(value="")
        self.timeout_var = tk.StringVar(value="10")

        self.table_var = tk.StringVar(value="")
        self.custom_sql_var = tk.BooleanVar(value=False)
        self.out_format_var = tk.StringVar(value=FORMAT_EXCEL)
        self.sheet_out_var = tk.StringVar(value="")

        self.excel_path_var = tk.StringVar(value="")
        self.sheet_in_var = tk.StringVar(value="")
        self.target_table_var = tk.StringVar(value="")
        self.mode_var = tk.StringVar(value=MODE_APPEND)
        self.batch_var = tk.StringVar(value=str(DEFAULT_BATCH))

        self.status_var = tk.StringVar(value="未连接 | 就绪")

    def _build_ui(self):
        pad = {"padx": 4, "pady": 3}

        # ---------- ① 数据库连接 ----------
        frm_conn = ttk.LabelFrame(self.root, text="① 数据库连接")
        frm_conn.grid(row=0, column=0, sticky="ew", padx=8, pady=(8, 4))
        for col in (0, 2, 3, 5, 6, 8, 9):
            frm_conn.columnconfigure(col, weight=1)

        ttk.Label(frm_conn, text="数据库类型:").grid(row=0, column=0, sticky="e", **pad)
        self.db_type_combo = ttk.Combobox(
            frm_conn, textvariable=self.db_type_var, values=list(ADAPTERS),
            state="readonly", width=12)
        self.db_type_combo.grid(row=0, column=1, sticky="w", **pad)
        self.db_type_combo.bind("<<ComboboxSelected>>", self.on_db_type_changed)

        ttk.Label(frm_conn, text="主机:").grid(row=0, column=2, sticky="e", **pad)
        ttk.Entry(frm_conn, textvariable=self.host_var, width=14).grid(
            row=0, column=3, sticky="w", **pad)

        ttk.Label(frm_conn, text="端口:").grid(row=0, column=4, sticky="e", **pad)
        ttk.Entry(frm_conn, textvariable=self.port_var, width=8).grid(
            row=0, column=5, sticky="w", **pad)

        ttk.Label(frm_conn, text="用户名:").grid(row=0, column=6, sticky="e", **pad)
        ttk.Entry(frm_conn, textvariable=self.user_var, width=12).grid(
            row=0, column=7, sticky="w", **pad)

        ttk.Label(frm_conn, text="密码:").grid(row=1, column=0, sticky="e", **pad)
        ttk.Entry(frm_conn, textvariable=self.pwd_var, width=12, show="*").grid(
            row=1, column=1, sticky="w", **pad)

        ttk.Label(frm_conn, text="数据库:").grid(row=1, column=2, sticky="e", **pad)
        ttk.Entry(frm_conn, textvariable=self.database_var, width=14).grid(
            row=1, column=3, sticky="w", **pad)

        ttk.Label(frm_conn, text="超时(秒):").grid(row=1, column=4, sticky="e", **pad)
        ttk.Entry(frm_conn, textvariable=self.timeout_var, width=8).grid(
            row=1, column=5, sticky="w", **pad)

        btn_box = ttk.Frame(frm_conn)
        btn_box.grid(row=1, column=6, columnspan=4, sticky="w", **pad)
        ttk.Button(btn_box, text="测试连接", command=self.on_test_connection).pack(
            side="left", padx=4)
        ttk.Button(btn_box, text="连接并加载表", command=self.on_connect).pack(
            side="left", padx=4)
        ttk.Button(btn_box, text="断开连接", command=self.disconnect).pack(
            side="left", padx=4)

        # ---------- ② SQL -> Excel / JSON ----------
        frm_export = ttk.LabelFrame(
            self.root, text="② SQL → Excel / JSON(数据库导出,文件名 = 数据库名)")
        frm_export.grid(row=1, column=0, sticky="ew", padx=8, pady=4)
        frm_export.columnconfigure(2, weight=1)
        frm_export.columnconfigure(4, weight=1)

        ttk.Label(frm_export, text="数据表:").grid(row=0, column=0, sticky="e", **pad)
        self.table_combo = ttk.Combobox(
            frm_export, textvariable=self.table_var, state="readonly", width=36)
        self.table_combo.grid(row=0, column=1, sticky="w", **pad)
        self.table_combo.bind("<<ComboboxSelected>>", self.on_table_selected)
        ttk.Button(frm_export, text="刷新表列表", command=self.on_refresh_tables).grid(
            row=0, column=2, sticky="w", **pad)

        self.sql_check = ttk.Checkbutton(
            frm_export, text="使用自定义 SQL(忽略上面的表选择):",
            variable=self.custom_sql_var, command=self.on_custom_sql_toggle)
        self.sql_check.grid(row=1, column=0, columnspan=2, sticky="w", **pad)

        ttk.Label(frm_export, text="输出格式:").grid(row=2, column=0, sticky="e", **pad)
        ttk.Combobox(
            frm_export, textvariable=self.out_format_var, values=FORMAT_LABELS,
            state="readonly", width=26).grid(row=2, column=1, sticky="w", **pad)

        ttk.Label(frm_export, text="Sheet名/JSON一级键:").grid(
            row=2, column=2, sticky="e", **pad)
        ttk.Entry(frm_export, textvariable=self.sheet_out_var, width=18).grid(
            row=2, column=3, sticky="w", **pad)

        self.sql_text = tk.Text(frm_export, height=3, width=80, font=FONT,
                                state="disabled", wrap="none")
        self.sql_text.grid(row=1, column=2, columnspan=3, rowspan=1,
                           sticky="ew", padx=4, pady=3)

        self.export_btn = ttk.Button(frm_export, text="SQL → 导出(Excel/JSON)",
                                     command=self.on_export)
        self.export_btn.grid(row=2, column=4, sticky="e", **pad)

        # ---------- ③ Excel -> SQL ----------
        frm_import = ttk.LabelFrame(self.root, text="③ Excel → SQL(Excel 导入到数据库)")
        frm_import.grid(row=2, column=0, sticky="ew", padx=8, pady=4)
        frm_import.columnconfigure(1, weight=1)

        ttk.Label(frm_import, text="Excel 文件:").grid(row=0, column=0, sticky="e", **pad)
        self.excel_entry = ttk.Entry(frm_import, textvariable=self.excel_path_var)
        self.excel_entry.grid(row=0, column=1, sticky="ew", **pad)
        ttk.Button(frm_import, text="浏览...", command=self.on_pick_excel).grid(
            row=0, column=2, sticky="w", **pad)

        ttk.Label(frm_import, text="工作表:").grid(row=1, column=0, sticky="e", **pad)
        self.sheet_in_combo = ttk.Combobox(
            frm_import, textvariable=self.sheet_in_var, state="readonly", width=30)
        self.sheet_in_combo.grid(row=1, column=1, sticky="w", **pad)

        ttk.Label(frm_import, text="目标表名:").grid(row=1, column=2, sticky="e", **pad)
        ttk.Entry(frm_import, textvariable=self.target_table_var, width=24).grid(
            row=1, column=3, sticky="w", **pad)

        ttk.Label(frm_import, text="写入模式:").grid(row=2, column=0, sticky="e", **pad)
        ttk.Combobox(frm_import, textvariable=self.mode_var, values=MODE_LABELS,
                     state="readonly", width=38).grid(
            row=2, column=1, sticky="w", **pad)
        ttk.Label(frm_import, text="每批行数:").grid(row=2, column=2, sticky="e", **pad)
        ttk.Entry(frm_import, textvariable=self.batch_var, width=10).grid(
            row=2, column=3, sticky="w", **pad)
        ttk.Button(frm_import, text="Excel → SQL 导入", command=self.on_import).grid(
            row=2, column=4, sticky="e", padx=8)

        # ---------- ④ 日志 ----------
        frm_log = ttk.LabelFrame(self.root, text="④ 运行日志")
        frm_log.grid(row=3, column=0, sticky="nsew", padx=8, pady=(4, 2))
        frm_log.columnconfigure(0, weight=1)
        frm_log.rowconfigure(0, weight=1)
        self.log_text = scrolledtext.ScrolledText(
            frm_log, height=12, font=("Consolas", 9), state="disabled", wrap="word")
        self.log_text.grid(row=0, column=0, sticky="nsew")

        # ---------- 状态栏 ----------
        self.status_var.set("未连接 | 就绪")
        status_bar = ttk.Label(self.root, textvariable=self.status_var,
                               anchor="w", relief="sunken")
        status_bar.grid(row=4, column=0, sticky="ew", padx=8, pady=(0, 6))

        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(3, weight=1)

    # ------------------------------------------------------- 线程 / 消息队列
    def log(self, message):
        """线程安全日志:worker 线程调用,主线程统一渲染。"""
        self.msg_queue.put(("log", str(message)))

    def set_status(self, message):
        self.msg_queue.put(("status", str(message)))

    def run_task(self, fn):
        """在后台线程执行耗时任务,避免界面卡死。"""
        if self._busy:
            messagebox.showinfo("提示", "已有任务正在执行,请等待完成后再试。")
            return False
        self._busy = True
        self.status_var.set("运行中...")

        def worker():
            try:
                fn()
            except Exception as exc:  # noqa: BLE001
                self.log(f"发生错误: {exc}")
                self.log(traceback.format_exc())
                self.msg_queue.put(("error", str(exc)))
            finally:
                self.msg_queue.put(("done", None))

        threading.Thread(target=worker, daemon=True).start()
        return True

    def _poll_queue(self):
        try:
            while True:
                kind, payload = self.msg_queue.get_nowait()
                if kind == "log":
                    self.log_text.configure(state="normal")
                    self.log_text.insert("end", payload + "\n")
                    self.log_text.see("end")
                    self.log_text.configure(state="disabled")
                elif kind == "status":
                    self.status_var.set(payload)
                elif kind == "tables":
                    self.table_combo["values"] = payload
                    if payload:
                        self.table_combo.current(0)
                        self.on_table_selected()
                elif kind == "sheets":
                    self.sheet_in_combo["values"] = payload
                    if payload:
                        self.sheet_in_combo.current(0)
                elif kind == "error":
                    messagebox.showerror("错误", payload)
                elif kind == "info":
                    messagebox.showinfo("提示", payload)
                elif kind == "done":
                    self._busy = False
                    self.status_var.set(
                        f"{'已连接' if self.conn else '未连接'} | 就绪")
        except queue.Empty:
            pass
        self.root.after(100, self._poll_queue)

    # ------------------------------------------------------------- 连接管理
    def get_cfg(self):
        return {
            "host": self.host_var.get().strip(),
            "port": self.port_var.get().strip() or str(self.adapter.DEFAULT_PORT),
            "user": self.user_var.get().strip(),
            "password": self.pwd_var.get(),
            "database": self.database_var.get().strip(),
            "charset": "utf8mb4",
            "timeout": self.timeout_var.get().strip() or "10",
        }

    def connect_db(self):
        cfg = self.get_cfg()
        if not cfg["host"] or not cfg["user"] or not cfg["database"]:
            raise ValueError("请填写【主机】【用户名】【数据库】后再连接")
        self._disconnect_conn()
        self.conn = self.adapter.connect(cfg)
        return self.adapter.server_version(self.conn)

    def _disconnect_conn(self):
        if self.conn is not None:
            try:
                self.conn.close()
            except Exception:
                pass
            self.conn = None

    def disconnect(self):
        self._disconnect_conn()
        self.status_var.set("未连接 | 就绪")

    def _ping(self):
        try:
            if self.adapter.NAME == "MySQL":
                self.conn.ping(reconnect=False)
                return True
            return not self.conn.closed
        except Exception:
            return False

    def ensure_conn(self):
        """确保有可用连接(worker 线程内调用)。"""
        try:
            if self.conn is not None and self._ping():
                return self.conn
        except Exception:
            pass
        self._disconnect_conn()
        version = self.connect_db()
        self.log(f"已自动连接 {self.adapter.NAME} {version}")
        self.set_status(f"已连接 {self.adapter.NAME} {version}")
        return self.conn

    # ------------------------------------------------------------- UI 事件
    def on_db_type_changed(self, _event=None):
        name = self.db_type_var.get()
        self.adapter = ADAPTERS[name]
        self.port_var.set(str(self.adapter.DEFAULT_PORT))
        self.user_var.set("root" if name == "MySQL" else "postgres")
        self.disconnect()
        self.log(f"已切换到 {name},默认端口 {self.adapter.DEFAULT_PORT}")

    def on_test_connection(self):
        def work():
            version = self.connect_db()
            self.log(f"连接成功: {self.adapter.NAME} {version}")
            self.msg_queue.put(("info", f"连接成功!\n{self.adapter.NAME} {version}"))
            self._disconnect_conn()
        self.run_task(work)

    def on_connect(self):
        def work():
            version = self.connect_db()
            tables = self.adapter.list_tables(self.conn)
            self.log(f"连接成功: {self.adapter.NAME} {version},共 {len(tables)} 张表")
            self.msg_queue.put(("tables", tables))
            self.set_status(f"已连接 {self.adapter.NAME} {version}")
        self.run_task(work)

    def on_refresh_tables(self):
        def work():
            conn = self.ensure_conn()
            tables = self.adapter.list_tables(conn)
            self.log(f"共发现 {len(tables)} 张表")
            self.msg_queue.put(("tables", tables))
        self.run_task(work)

    def on_custom_sql_toggle(self):
        state = "normal" if self.custom_sql_var.get() else "disabled"
        self.sql_text.configure(state=state)

    def on_table_selected(self, _event=None):
        """选择数据表后,自动把 Sheet 名 / JSON 一级键填成表名。"""
        table = self.table_var.get().strip()
        if table and not self.custom_sql_var.get():
            self.sheet_out_var.set(table.split(".")[-1])

    def ask_conflict_mode(self, existing_paths):
        """同名文件冲突纠错对话框,返回 'replace' / 'merge' / None(取消)。"""
        dlg = tk.Toplevel(self.root)
        dlg.title("检测到同名文件")
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.resizable(False, False)
        lines = "\n".join(f"  • {p}" for p in existing_paths)
        ttk.Label(
            dlg, justify="left",
            text=(f"以下文件已存在:\n{lines}\n\n请选择处理方式:"),
        ).pack(padx=14, pady=(14, 8))
        box = ttk.Frame(dlg)
        box.pack(padx=14, pady=(0, 14))
        result = {"mode": None}

        def pick(mode):
            result["mode"] = mode
            dlg.destroy()

        ttk.Button(box, text="覆盖整个文件",
                   command=lambda: pick("replace")).pack(side="left", padx=4)
        ttk.Button(box, text="合并写入(仅覆盖同名Sheet/JSON键)",
                   command=lambda: pick("merge")).pack(side="left", padx=4)
        ttk.Button(box, text="取消",
                   command=lambda: pick(None)).pack(side="left", padx=4)
        dlg.wait_window()
        return result["mode"]

    @staticmethod
    def _parse_int(var, default):
        try:
            return max(1, int(str(var.get()).strip()))
        except (TypeError, ValueError):
            return default

    # ------------------------------------------- SQL -> Excel / JSON 导出
    def on_export(self):
        if self._busy:
            messagebox.showinfo("提示", "已有任务正在执行,请等待完成后再试。")
            return
        adapter = self.adapter
        custom = self.custom_sql_var.get()

        if custom:
            sql = self.sql_text.get("1.0", "end").strip()
            if not sql:
                messagebox.showwarning("提示", "请填写自定义 SQL 语句")
                return
            default_key = "query_result"
        else:
            table = self.table_var.get().strip()
            if not table:
                messagebox.showwarning("提示", "请先选择数据表,或勾选“使用自定义 SQL”")
                return
            sql = f"SELECT * FROM {adapter.quote_ident(table)}"
            default_key = table.split(".")[-1]

        if not re.match(r"^\s*(SELECT|WITH)\b", sql, re.I):
            messagebox.showwarning("提示", "仅支持 SELECT / WITH 开头的查询语句")
            return

        # ---- 输出格式:Excel / JSON / 两者 ----
        fmt = self.out_format_var.get()
        want_excel = fmt in (FORMAT_EXCEL, FORMAT_BOTH)
        want_json = fmt in (FORMAT_JSON, FORMAT_BOTH)

        # ---- Sheet 名 == JSON 一级键 == 表名 ----
        key_name = (self.sheet_out_var.get() or "").strip() or default_key
        sheet = safe_sheet_title(key_name)

        # ---- 文件名 == 数据库名 ----
        db_name = self.database_var.get().strip() or default_key
        file_base = safe_filename(db_name)

        try:
            initial_dir = str(Path.home() / "Documents")
        except Exception:
            initial_dir = str(Path(__file__).parent)

        excel_path = None
        json_path = None
        if fmt == FORMAT_EXCEL:
            excel_path = filedialog.asksaveasfilename(
                title="保存 Excel 文件(文件名 = 数据库名)",
                defaultextension=".xlsx", initialdir=initial_dir,
                initialfile=f"{file_base}.xlsx",
                filetypes=[("Excel 工作簿", "*.xlsx")])
            if not excel_path:
                return
        elif fmt == FORMAT_JSON:
            json_path = filedialog.asksaveasfilename(
                title="保存 JSON 文件(文件名 = 数据库名)",
                defaultextension=".json", initialdir=initial_dir,
                initialfile=f"{file_base}.json",
                filetypes=[("JSON 文件", "*.json")])
            if not json_path:
                return
        else:  # Excel + JSON
            excel_path = filedialog.asksaveasfilename(
                title="保存 Excel + JSON(文件名 = 数据库名)",
                defaultextension=".xlsx", initialdir=initial_dir,
                initialfile=f"{file_base}.xlsx",
                filetypes=[("Excel 工作簿", "*.xlsx")])
            if not excel_path:
                return
            json_path = str(Path(excel_path).with_suffix(".json"))

        # ---- 同名文件纠错:覆盖整个文件 / 合并写入 / 取消 ----
        existing = [p for p in (excel_path, json_path) if p and os.path.exists(p)]
        conflict_mode = "replace"
        if existing:
            conflict_mode = self.ask_conflict_mode(existing)
            if conflict_mode is None:
                self.log("已取消导出(存在同名文件)")
                return

        batch = self._parse_int(self.batch_var, DEFAULT_BATCH)

        def work():
            conn = self.ensure_conn()
            if conn is None:
                return
            cur = adapter.open_query_cursor(conn)
            total = 0
            json_rows = []  # JSON 是列数组结构,需要整表数据暂存内存
            try:
                cur.execute(sql)
                # PostgreSQL 命名游标的 description 要等首次 fetch 后才就绪
                first_rows = cur.fetchmany(batch)
                columns = [str(desc[0]) for desc in cur.description]
                self.log(f"开始导出: {sql}")
                self.log(f"共 {len(columns)} 列 | Sheet名/JSON键: {key_name}"
                         f" | 文件名基准: {file_base}(冲突处理: {conflict_mode})")
                if want_json:
                    self.log("提示:JSON 为“列数组”结构,导出期间会把整表数据暂存内存")

                def fetch_more():
                    while True:
                        rows = cur.fetchmany(batch)
                        if not rows:
                            break
                        for row in rows:
                            yield row

                def base_iter():
                    for row in first_rows:
                        yield row
                    yield from fetch_more()

                def collecting_iter():
                    for row in base_iter():
                        json_rows.append(row)
                        yield row

                if want_excel:
                    row_iter = collecting_iter() if want_json else base_iter()
                    total = write_excel_file(
                        excel_path, sheet, columns, row_iter,
                        mode=conflict_mode, log=self.log)
                    self.log(f"✅ Excel 导出完成: {excel_path}(共 {total:,} 行)")

                if want_json:
                    if not json_rows:  # 仅导出 JSON 时
                        json_rows = list(base_iter())
                    write_json_file(json_path, key_name, columns, json_rows,
                                    mode=conflict_mode, log=self.log)
                    total = len(json_rows)
                    self.log(f"✅ JSON 导出完成: {json_path}(共 {total:,} 行)")
            finally:
                try:
                    cur.close()
                except Exception:
                    pass
                if adapter.NAME == "PostgreSQL":
                    try:
                        conn.rollback()  # 结束服务端命名游标事务
                    except Exception:
                        pass

            outputs = ", ".join(p for p in (excel_path, json_path) if p)
            self.log(f"✅ 导出完成: {outputs}(共 {total:,} 行)")
            self.msg_queue.put(("info", f"导出完成,共 {total:,} 行:\n{outputs}"))

        self.run_task(work)

    # ----------------------------------------------------- Excel -> SQL
    def on_pick_excel(self):
        path = filedialog.askopenfilename(
            title="选择 Excel 文件",
            filetypes=[("Excel 工作簿", "*.xlsx"), ("所有文件", "*.*")])
        if not path:
            return
        self.excel_path_var.set(path)
        stem = re.sub(r"[^\w\u4e00-\u9fff]+", "_", Path(path).stem) or "imported_table"
        self.target_table_var.set(stem)

        def work():
            wb = load_workbook(path, read_only=True)
            try:
                names = wb.sheetnames
            finally:
                wb.close()
            self.log(f"工作簿包含工作表: {', '.join(names)}")
            self.msg_queue.put(("sheets", names))
        self.run_task(work)

    def _import_mode_key(self):
        label = self.mode_var.get()
        if MODE_REPLACE in label:
            return "replace"
        if MODE_CREATE in label:
            return "create_if_missing"
        return "append"

    def on_import(self):
        if self._busy:
            messagebox.showinfo("提示", "已有任务正在执行,请等待完成后再试。")
            return
        path = self.excel_path_var.get().strip()
        sheet = self.sheet_in_var.get().strip()
        table = self.target_table_var.get().strip()
        mode = self._import_mode_key()

        if not path or not os.path.isfile(path):
            messagebox.showwarning("提示", "请先选择有效的 Excel 文件")
            return
        if not sheet:
            messagebox.showwarning("提示", "请选择要导入的工作表")
            return
        if not table:
            messagebox.showwarning("提示", "请填写目标表名")
            return
        if mode == "replace" and not messagebox.askyesno(
                "危险操作", f"将先删除表【{table}】(若存在)再重建并导入,确定继续吗?"):
            return

        adapter = self.adapter
        batch = self._parse_int(self.batch_var, DEFAULT_BATCH)

        def work():
            conn = self.ensure_conn()
            if conn is None:
                return

            # ---- 第一遍:读表头 + 采样前 SAMPLE_SIZE 行用于类型推断 ----
            self.log(f"读取 Excel: {path} → 工作表 [{sheet}]")
            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                if sheet not in wb.sheetnames:
                    raise ValueError(
                        f"工作簿中不存在工作表 {sheet!r},可用: {', '.join(wb.sheetnames)}")
                ws = wb[sheet]
                header = None
                sample = []
                for idx, row in enumerate(ws.iter_rows(values_only=True)):
                    if idx == 0:
                        header = parse_header(row)
                        continue
                    if is_blank_row(row):
                        continue
                    if len(sample) < SAMPLE_SIZE:
                        sample.append(normalize_row(row, len(header)))
                    if len(sample) >= SAMPLE_SIZE and mode != "append":
                        break
            finally:
                wb.close()
            self.log(f"表头列: {header}")

            # ---- 建表策略 ----
            quoted_table = adapter.quote_ident(table)
            exists = adapter.table_exists(conn, quoted_table)
            need_create = (mode == "replace") or (mode == "create_if_missing" and not exists)
            if mode == "append" and not exists:
                raise ValueError(
                    f"目标表【{table}】不存在。请改用“{MODE_CREATE}”或“{MODE_REPLACE}”模式")
            if need_create:
                stats = collect_stats(header, sample)
                if mode == "replace" and exists:
                    with conn.cursor() as cur:
                        cur.execute(f"DROP TABLE {quoted_table}")
                    self.log(f"已删除旧表: {table}")
                ddl = build_create_table(adapter, table, header, stats)
                with conn.cursor() as cur:
                    cur.execute(ddl)
                self.log(f"已创建表: {table}")

            # ---- 第二遍:流式逐行读取并分批插入 ----
            quoted_cols = [adapter.quote_ident(col) for col in header]
            insert_sql = (
                f"INSERT INTO {quoted_table} ({', '.join(quoted_cols)}) "
                f"VALUES ({', '.join(['%s'] * len(header))})"
            )
            total, skipped = 0, 0
            chunk = []
            try:
                wb = load_workbook(path, read_only=True, data_only=True)
                try:
                    ws = wb[sheet]
                    with conn.cursor() as cur:
                        for idx, row in enumerate(ws.iter_rows(values_only=True)):
                            if idx == 0:
                                continue
                            if is_blank_row(row):
                                skipped += 1
                                continue
                            norm = normalize_row(row, len(header))
                            chunk.append(tuple(to_db_value(v) for v in norm))
                            if len(chunk) >= batch:
                                cur.executemany(insert_sql, chunk)
                                total += len(chunk)
                                chunk = []
                                self.log(f"  已写入 {total:,} 行 ...")
                        if chunk:
                            cur.executemany(insert_sql, chunk)
                            total += len(chunk)
                    conn.commit()
                finally:
                    wb.close()
            except Exception:
                try:
                    conn.rollback()
                except Exception:
                    pass
                raise

            self.log(f"✅ 导入完成: 共 {total:,} 行写入表【{table}】"
                     + (f",跳过空行 {skipped} 行" if skipped else ""))
            self.msg_queue.put(
                ("info", f"导入完成,共 {total:,} 行写入表【{table}】"))

        self.run_task(work)

    # ------------------------------------------------------------- 退出
    def on_close(self):
        self._disconnect_conn()
        self.root.destroy()


def run_self_test():
    """--selftest: 无界面自检(打包后验证驱动/引擎是否齐全),结果写入当前目录 selftest.log。"""
    log_path = Path.cwd() / "selftest.log"
    lines = []

    for module_name in ("pymysql", "psycopg2", "openpyxl"):
        try:
            __import__(module_name)
            lines.append(f"[OK]   import {module_name}")
        except Exception as exc:
            lines.append(f"[FAIL] import {module_name}: {exc}")

    # Excel 写读往返(覆盖 openpyxl + et_xmlfile)
    try:
        tmp = log_path.parent / "_selftest.xlsx"
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title="自检")
        write_rows_to_worksheet(
            ws, ["id", "name"], iter([[1, "张三"], [2, "李四"]]))
        wb.save(tmp)
        wb.close()
        wb2 = load_workbook(tmp, read_only=True, data_only=True)
        rows = list(wb2[wb2.sheetnames[0]].iter_rows(values_only=True))
        wb2.close()
        tmp.unlink(missing_ok=True)
        ok = len(rows) == 3 and rows[1][0] == 1
        lines.append(f"[{'OK' if ok else 'FAIL'}]   openpyxl 写读往返: {len(rows)} 行")
    except Exception as exc:
        lines.append(f"[FAIL] openpyxl 写读往返: {exc}")

    # JSON 写读往返(覆盖列数组结构 + 同名键合并逻辑)
    try:
        tmp = log_path.parent / "_selftest.json"
        rows = [[1, "张三"], [2, "李四"]]
        write_json_file(str(tmp), "info", ["Id", "Name"], rows)
        with open(tmp, "r", encoding="utf-8") as fp:
            doc = json.load(fp)
        ok = doc["info"]["Id"] == [1, 2] and doc["info"]["Name"][-1] == "李四"
        # 合并模式:覆盖同名键、保留其他一级键
        write_json_file(str(tmp), "info", ["Id", "Name"], [[3, "王五"]], mode="merge")
        write_json_file(str(tmp), "other", ["X"], [[9]], mode="merge")
        with open(tmp, "r", encoding="utf-8") as fp:
            doc = json.load(fp)
        ok = ok and doc["info"]["Id"] == [3] and "other" in doc
        tmp.unlink(missing_ok=True)
        lines.append(f"[{'OK' if ok else 'FAIL'}]   JSON 写读/合并往返")
    except Exception as exc:
        lines.append(f"[FAIL] JSON 写读/合并往返: {exc}")

    # 若提供数据库连接环境变量,则做真实连接自检
    for env_prefix, adapter in (("SELFTEST_MYSQL", ADAPTERS["MySQL"]),
                                ("SELFTEST_PG", ADAPTERS["PostgreSQL"])):
        env = {
            "host": os.environ.get(f"{env_prefix}_HOST", "127.0.0.1"),
            "port": os.environ.get(f"{env_prefix}_PORT") or str(adapter.DEFAULT_PORT),
            "user": os.environ.get(f"{env_prefix}_USER", ""),
            "password": os.environ.get(f"{env_prefix}_PASSWORD", ""),
            "database": os.environ.get(f"{env_prefix}_DATABASE", ""),
            "charset": "utf8mb4",
            "timeout": "10",
        }
        if not env["user"] or not env["database"]:
            lines.append(f"[SKIP] {adapter.NAME} 连接自检(未设置 {env_prefix}_* 环境变量)")
            continue
        try:
            conn = adapter.connect(env)
            with conn.cursor() as cur:
                cur.execute("SELECT 1")
                cur.fetchone()
            conn.close()
            lines.append(f"[OK]   {adapter.NAME} 连接 SELECT 1")
        except Exception as exc:
            lines.append(f"[FAIL] {adapter.NAME} 连接: {exc}")

    log_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return all(not line.startswith("[FAIL]") for line in lines)


def main():
    if "--selftest" in sys.argv:
        sys.exit(0 if run_self_test() else 1)
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
